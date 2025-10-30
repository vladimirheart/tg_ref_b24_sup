# panel/app.py
from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    session,
    Response,
    send_file,
    send_from_directory,
    abort,
    g,
)
import sqlite3
import datetime
import requests
import json
import os
import logging
import re
import unicodedata
import time, sqlite3
from uuid import uuid4
import sys
import shutil
from pathlib import Path
import secrets
from functools import wraps

PARENT_DIR = Path(__file__).resolve().parent.parent
if str(PARENT_DIR) not in sys.path:
    sys.path.insert(0, str(PARENT_DIR))

from bot_settings_utils import (
    DEFAULT_BOT_PRESET_DEFINITIONS,
    build_location_presets,
    sanitize_bot_settings,
)
def exec_with_retry(fn, retries=5, base_delay=0.15):
    for i in range(retries):
        try:
            return fn()
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e).lower() and i < retries - 1:
                time.sleep(base_delay * (i + 1))  # backoff
                continue
            raise
from zoneinfo import ZoneInfo
from apscheduler.schedulers.background import BackgroundScheduler
from threading import Timer
from datetime import datetime as dt, timedelta, timezone
from werkzeug.security import generate_password_hash, check_password_hash
import io
import mimetypes
from typing import Any
from werkzeug.utils import secure_filename

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
SETTINGS_PATH = os.path.join(BASE_DIR, "settings.json")
ATTACHMENTS_DIR = os.path.join(BASE_DIR, "attachments")
KNOWLEDGE_BASE_ATTACHMENTS_DIR = os.path.join(ATTACHMENTS_DIR, "knowledge_base")
TICKETS_DB_PATH = os.path.join(BASE_DIR, "tickets.db")
USERS_DB_PATH = os.path.join(BASE_DIR, "users.db")
LOCATIONS_PATH = os.path.join(BASE_DIR, "locations.json")
OBJECT_PASSPORT_DB_PATH = os.path.join(BASE_DIR, "object_passports.db")
ORG_STRUCTURE_PATH = os.path.join(BASE_DIR, "org_structure.json")
OBJECT_PASSPORT_UPLOADS_DIR = os.path.join(BASE_DIR, "object_passport_uploads")
USER_PHOTOS_DIR = os.path.join(os.path.dirname(__file__), "static", "user_photos")
ALLOWED_USER_PHOTO_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
MAX_USER_PHOTO_SIZE = 5 * 1024 * 1024
WEB_FORM_SESSIONS_TABLE = "web_form_sessions"

os.makedirs(OBJECT_PASSPORT_UPLOADS_DIR, exist_ok=True)
os.makedirs(KNOWLEDGE_BASE_ATTACHMENTS_DIR, exist_ok=True)
os.makedirs(USER_PHOTOS_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY")

# === НАСТРОЙКИ ===
from dotenv import load_dotenv
import os

load_dotenv()

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
GROUP_CHAT_ID = int(os.getenv("GROUP_CHAT_ID"))

PARAMETER_TYPES = {
    "business": "Бизнес",
    "partner_type": "Тип партнёра",
    "country": "Страна",
    "legal_entity": "ЮЛ",
    "partner_contact": "Контакты партнёров и КА",
    "department": "Департамент",
    "network": "Внутренняя сеть",
    "it_connection": "Подключения IT-блока",
    "remote_access": "Параметры удалённого доступа",
    "iiko_server": "Адреса серверов iiko",
}

REMOTE_ACCESS_DEFAULTS = (
    "RMSviewer",
    "Anydes",
    "Ассистент",
    "VNC",
)

DEFAULT_IT_CONNECTION_CATEGORIES = {
    "equipment_type": "Тип оборудования",
    "equipment_vendor": "Производитель оборудования",
    "equipment_model": "Модель оборудования",
    "equipment_status": "Статус оборудования",
}

DEFAULT_IT_CONNECTION_CATEGORY_FIELDS = {
    "equipment_type": "equipment_type",
    "equipment_vendor": "equipment_vendor",
    "equipment_model": "equipment_model",
    "equipment_status": "equipment_status",
}


def slugify_it_connection_category(label: str, existing: set[str] | None = None) -> str:
    """Create a slug identifier for an IT-connection category."""
    normalized = unicodedata.normalize("NFKD", str(label or "")).strip()
    # Replace all non-word characters with underscores
    base = re.sub(r"[^\w]+", "_", normalized, flags=re.UNICODE).strip("_").lower()
    if not base:
        base = "category"
    if base[0].isdigit():
        base = f"cat_{base}"
    existing_keys = set(existing or set())
    candidate = base
    counter = 2
    while candidate in existing_keys:
        candidate = f"{base}_{counter}"
        counter += 1
    return candidate


def normalize_it_connection_categories(raw) -> dict[str, str]:
    """Normalize user-provided IT-connection categories into a {slug: label} dict."""
    categories: dict[str, str] = {}
    if isinstance(raw, dict):
        for key, label in raw.items():
            key_clean = str(key or "").strip()
            label_clean = str(label or "").strip()
            if not key_clean or not label_clean:
                continue
            categories[key_clean] = label_clean
        return categories
    if isinstance(raw, list):
        existing = set(DEFAULT_IT_CONNECTION_CATEGORIES.keys())
        for item in raw:
            label = None
            key = None
            if isinstance(item, dict):
                label = str(item.get("label") or "").strip()
                key = str(item.get("key") or "").strip()
            else:
                label = str(item or "").strip()
            if not label:
                continue
            if not key:
                key = slugify_it_connection_category(label, existing)
            categories[key] = label
            existing.add(key)
        return categories
    return {}


def get_custom_it_connection_categories(settings: dict | None = None) -> dict[str, str]:
    """Return custom categories stored in settings.json."""
    if settings is None:
        settings = load_settings()
    raw = settings.get("it_connection_categories") if isinstance(settings, dict) else {}
    normalized = normalize_it_connection_categories(raw)
    return normalized


def get_it_connection_categories(
    settings: dict | None = None, *, include: dict[str, str] | None = None
) -> dict[str, str]:
    """Return full list of IT-connection categories including defaults and custom ones."""
    if settings is None:
        settings = load_settings()
    custom = get_custom_it_connection_categories(settings)
    categories: dict[str, str] = dict(DEFAULT_IT_CONNECTION_CATEGORIES)
    for key, label in custom.items():
        if not key or not label:
            continue
        categories[key] = label
    if include:
        for key, label in include.items():
            key_clean = str(key or "").strip()
            if not key_clean:
                continue
            label_clean = str(label or "").strip() or key_clean
            categories.setdefault(key_clean, label_clean)
    return categories


def save_it_connection_categories(
    categories: dict[str, str], settings: dict | None = None
) -> dict[str, str]:
    """Persist custom IT-connection categories to settings.json."""
    if settings is None:
        settings = load_settings()
    normalized = normalize_it_connection_categories(categories)
    settings["it_connection_categories"] = normalized
    save_settings(settings)
    return settings

PARAMETER_STATE_TYPES = {"partner_type", "country", "legal_entity", "partner_contact"}
PARAMETER_ALLOWED_STATES = (
    "Активен",
    "Подписание",
    "Заблокирован",
    "Black_List",
    "Закрыт",
)

PARTNER_CONTACT_TYPES = {
    "partner": "Партнёр",
    "ka": "КА",
}

PARTNER_CONTACT_PHONE_TYPES = {
    "work": "Рабочий",
    "personal": "Личный",
    "fax": "Факс",
    "support": "Саппорт",
}

PARTNER_CONTACT_EMAIL_TYPES = {
    "work": "Рабочий",
    "personal": "Личный",
    "common": "Общий",
    "support": "Саппорт",
}

PARAMETER_USAGE_MAPPING = {
    "business": "business",
    "partner_type": "partner_type",
    "country": "country",
    "legal_entity": "legal_entity",
    "department": "department",
    "network": "network",
    "it_connection": "it_connection_type",
    "iiko_server": "it_iiko_server",
}

def _normalize_partner_contact_channel_list(raw_list, allowed_types: dict[str, str], default_key: str):
    allowed_keys = list(allowed_types.keys())
    fallback = default_key if default_key in allowed_types else (allowed_keys[0] if allowed_keys else "")
    result: list[dict[str, str]] = []
    if not isinstance(raw_list, (list, tuple)):
        return result
    for item in raw_list:
        if isinstance(item, dict):
            raw_type = (item.get("type") or "").strip()
            value = (item.get("value") or "").strip()
        else:
            raw_type = ""
            value = str(item or "").strip()
        if not value:
            continue
        normalized_type = raw_type if raw_type in allowed_types else fallback
        result.append({"type": normalized_type, "value": value})
    return result


def _normalize_partner_contact_served_entities(raw_value):
    """Normalize served legal entity entries into a list of {id, name}."""

    items: list = []
    if isinstance(raw_value, (list, tuple)):
        items = list(raw_value)
    elif raw_value is None:
        items = []
    else:
        items = [raw_value]

    result: list[dict[str, object]] = []
    seen: set[int] = set()

    for item in items:
        raw_id = None
        raw_name = ""
        if isinstance(item, dict):
            raw_id = item.get("id")
            if raw_id is None:
                raw_id = item.get("value")
            if isinstance(item.get("name"), str):
                raw_name = item.get("name") or ""
            elif isinstance(item.get("label"), str):
                raw_name = item.get("label") or ""
            elif isinstance(item.get("title"), str):
                raw_name = item.get("title") or ""
            elif isinstance(item.get("served_legal_entity_name"), str):
                raw_name = item.get("served_legal_entity_name") or ""
        elif isinstance(item, (list, tuple)) and item:
            raw_id = item[0]
            if len(item) > 1:
                raw_name = item[1]
        else:
            raw_id = item

        try:
            parsed_id = int(raw_id)
        except (TypeError, ValueError):
            parsed_id = None
        if not parsed_id or parsed_id <= 0:
            continue
        if parsed_id in seen:
            continue
        seen.add(parsed_id)
        name = str(raw_name or "").strip()
        result.append({"id": parsed_id, "name": name})

    return result


def _normalize_partner_contact_people(raw_value) -> list[dict]:
    """Normalize a list of contact persons."""

    if isinstance(raw_value, dict) and "contacts" in raw_value:
        raw_value = raw_value.get("contacts")

    if not isinstance(raw_value, (list, tuple)):
        return []

    result: list[dict] = []
    seen_keys: set[str] = set()

    for item in raw_value:
        if not isinstance(item, dict):
            continue

        key = str(item.get("key") or "").strip()
        if key and key in seen_keys:
            key = ""
        if key:
            seen_keys.add(key)

        full_name = str(item.get("full_name") or "").strip()
        position = str(item.get("position") or "").strip()
        phone = re.sub(r"\D+", "", str(item.get("phone") or ""))
        email = str(item.get("email") or "").strip()

        if not (full_name or position or phone or email):
            continue

        entry = {
            "full_name": full_name,
            "position": position,
            "phone": phone,
            "email": email,
        }
        if key:
            entry["key"] = key

        result.append(entry)

    return result


def sanitize_partner_contact_extra(payload, *, existing: dict | None = None, value: str | None = None) -> dict:
    base: dict[str, object] = {}
    if isinstance(existing, dict):
        base.update(existing)

    if value is not None:
        base["legal_entity"] = (value or "").strip()
    else:
        base["legal_entity"] = (base.get("legal_entity") or "").strip()

    if "contact_type" in payload or "contact_type" not in base:
        raw_type = (payload.get("contact_type") if isinstance(payload, dict) else None) or base.get("contact_type") or ""
        raw_type = str(raw_type or "").strip()
        contact_type = raw_type if raw_type in PARTNER_CONTACT_TYPES else "partner"
        base["contact_type"] = contact_type

    if "phones" in payload or "phones" not in base:
        phones_raw = payload.get("phones") if isinstance(payload, dict) else None
        if phones_raw is None:
            phones_raw = base.get("phones")
        base["phones"] = _normalize_partner_contact_channel_list(
            phones_raw, PARTNER_CONTACT_PHONE_TYPES, "work"
        )

    if "emails" in payload or "emails" not in base:
        emails_raw = payload.get("emails") if isinstance(payload, dict) else None
        if emails_raw is None:
            emails_raw = base.get("emails")
        base["emails"] = _normalize_partner_contact_channel_list(
            emails_raw, PARTNER_CONTACT_EMAIL_TYPES, "work"
        )

    base["contacts"] = _normalize_partner_contact_people(base.get("contacts"))

    if isinstance(payload, dict) and "internal_name" in payload:
        internal_name = str(payload.get("internal_name") or "").strip()
    else:
        internal_name = str(base.get("internal_name") or "").strip()
    base["internal_name"] = internal_name

    base_entities = _normalize_partner_contact_served_entities(base.get("served_legal_entities"))
    if not base_entities:
        single_fallback = {
            "id": base.get("served_legal_entity_id"),
            "name": base.get("served_legal_entity_name"),
        }
        base_entities = _normalize_partner_contact_served_entities(
            [] if single_fallback["id"] is None else [single_fallback]
        )

    if isinstance(payload, dict):
        if "served_legal_entities" in payload:
            base_entities = _normalize_partner_contact_served_entities(payload.get("served_legal_entities"))
        elif "served_legal_entity_ids" in payload:
            raw_ids = payload.get("served_legal_entity_ids")
            base_entities = _normalize_partner_contact_served_entities(raw_ids)
            if base_entities:
                existing_lookup = {
                    entry["id"]: entry.get("name", "")
                    for entry in _normalize_partner_contact_served_entities(base.get("served_legal_entities"))
                }
                for entry in base_entities:
                    if not entry["name"]:
                        entry["name"] = existing_lookup.get(entry["id"], "")
        elif "served_legal_entity_id" in payload or "served_legal_entity_name" in payload:
            base_entities = _normalize_partner_contact_served_entities(
                [
                    {
                        "id": payload.get("served_legal_entity_id"),
                        "name": payload.get("served_legal_entity_name"),
                    }
                ]
            )

    if isinstance(payload, dict) and "contacts" in payload:
        base["contacts"] = _normalize_partner_contact_people(payload.get("contacts"))

    base_entities = _normalize_partner_contact_served_entities(base_entities)
    base["served_legal_entities"] = base_entities
    base["served_legal_entity_ids"] = [entry["id"] for entry in base_entities]
    base["served_legal_entity_names"] = [entry["name"] for entry in base_entities if entry["name"]]
    first_entry = base_entities[0] if base_entities else None
    base["served_legal_entity_id"] = first_entry["id"] if first_entry else None
    base["served_legal_entity_name"] = first_entry["name"] if first_entry else ""

    return base


def normalize_partner_contact_extra(extra: dict | None, value: str | None = None) -> dict:
    return sanitize_partner_contact_extra({}, existing=extra, value=value)

LEGACY_PARAMETER_KEY_CANDIDATES = {
    "business": ["business", "businesses", "business_list"],
    "partner_type": ["partner_type", "partner_types", "partnerType"],
    "country": ["country", "countries"],
    "legal_entity": ["legal_entity", "legal_entities", "legalEntities"],
    "department": ["department", "departments", "department_list"],
    "network": ["network", "networks", "network_list"],
    "it_connection": [
        "it_connection",
        "it_connections",
        "itConnection",
        "it_connections_list",
    ],
    "remote_access": ["remote_access", "remote_accesses", "remoteAccess"],
    "iiko_server": ["iiko_server", "iiko_servers", "iikoServers"],
}

LEGACY_PARAMETER_CONTAINER_KEYS = (
    "parameters",
    "parameter_values",
    "parameterValues",
    "parameter_options",
    "parameterOptions",
    "parameter_values_map",
)

LEGACY_IT_CATEGORY_ALIASES = {
    "type": "equipment_type",
    "equipment": "equipment_type",
    "vendor": "equipment_vendor",
    "brand": "equipment_vendor",
    "model": "equipment_model",
    "status": "equipment_status",
}

def _is_missing_table_error(error: Exception, table: str) -> bool:
    """Return True if the sqlite error corresponds to a missing table."""

    if not isinstance(error, sqlite3.OperationalError):
        return False
    message = str(error).lower()
    return f"no such table: {table}".lower() in message

def _normalize_parameter_state(param_type, state):
    default_state = PARAMETER_ALLOWED_STATES[0]
    if param_type in PARAMETER_STATE_TYPES:
        candidate = (state or "").strip()
        if candidate in PARAMETER_ALLOWED_STATES:
            return candidate
        return default_state
    return default_state

# Подключение к базе
def get_db():
    # autocommit (isolation_level=None) + увеличенный timeout
    conn = sqlite3.connect(TICKETS_DB_PATH, timeout=30, isolation_level=None)
    conn.row_factory = sqlite3.Row
    # режим WAL и адекватный busy_timeout на всякий случай
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA busy_timeout=30000;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn


def _extract_operator_name(message: str) -> str:
    if not message:
        return ""
    text = str(message)
    patterns = [
        r"от поддержки\s*\(([^)]+)\)",
        r"^От:\s*(.+)$",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


def _resolve_ticket_responsible(cur: sqlite3.Cursor, ticket_id: str, channel_id: int | None = None) -> dict[str, str | None]:
    assigned_row = cur.execute(
        "SELECT responsible, assigned_at, assigned_by FROM ticket_responsibles WHERE ticket_id = ?",
        (ticket_id,),
    ).fetchone()
    manual = (assigned_row["responsible"].strip() if assigned_row and assigned_row["responsible"] else "")
    assigned_at = assigned_row["assigned_at"] if assigned_row else None
    assigned_by = assigned_row["assigned_by"] if assigned_row else None

    first_support_row = cur.execute(
        """
        SELECT sender, message
        FROM chat_history
        WHERE ticket_id = ?
          AND (channel_id = ? OR channel_id IS NULL OR ? IS NULL)
          AND sender IS NOT NULL
          AND TRIM(sender) != ''
          AND LOWER(sender) != 'user'
        ORDER BY timestamp ASC
        LIMIT 1
        """,
        (ticket_id, channel_id, channel_id),
    ).fetchone()

    auto = ""
    if first_support_row:
        sender = (first_support_row["sender"] or "").strip()
        if sender and sender.lower() not in {"support", "bot"}:
            auto = sender
        else:
            auto = _extract_operator_name(first_support_row["message"] or "")

    effective = manual or auto
    source = "manual" if manual else ("auto" if auto else "")

    return {
        "manual": manual,
        "auto": auto,
        "assigned_at": assigned_at,
        "assigned_by": assigned_by,
        "responsible": effective,
        "source": source,
    }

def _ensure_users_schema(conn: sqlite3.Connection) -> None:
    """Добавляет недостающие поля в таблицу пользователей."""

    cursor = conn.execute("PRAGMA table_info(users)")
    existing_columns = {row[1] for row in cursor.fetchall()}
    schema_updated = False

    def add_column(sql: str) -> None:
        nonlocal schema_updated
        conn.execute(sql)
        schema_updated = True

    if "full_name" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN full_name TEXT")
    if "photo" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN photo TEXT")
    if "registration_date" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN registration_date TEXT")
    if "birth_date" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN birth_date TEXT")
    if "email" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN email TEXT")
    if "department" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN department TEXT")
    if "phones" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN phones TEXT")
    if "is_blocked" not in existing_columns:
        add_column("ALTER TABLE users ADD COLUMN is_blocked INTEGER NOT NULL DEFAULT 0")

    if schema_updated:
        conn.commit()

    # Убедимся, что у существующих записей есть дата регистрации
    needs_registration_update = conn.execute(
        "SELECT COUNT(1) FROM users WHERE registration_date IS NULL"
    ).fetchone()[0]
    if needs_registration_update:
        now_iso = datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
        conn.execute(
            "UPDATE users SET registration_date = ? WHERE registration_date IS NULL",
            (now_iso,),
        )
        conn.commit()


def get_users_db():
    conn = sqlite3.connect(USERS_DB_PATH)
    conn.row_factory = sqlite3.Row
    _ensure_users_schema(conn)
    return conn

PERMISSION_WILDCARD = "*"

AVAILABLE_PAGE_PERMISSIONS = [
    {"key": "dialogs", "label": "Диалоги"},
    {"key": "tasks", "label": "Задачи"},
    {"key": "clients", "label": "Клиенты"},
    {"key": "object_passports", "label": "Паспорта объектов"},
    {"key": "knowledge_base", "label": "База знаний"},
    {"key": "dashboard", "label": "Дашборд"},
    {"key": "analytics", "label": "Аналитика"},
    {"key": "channels", "label": "Каналы"},
    {"key": "settings", "label": "Настройки"},
    {"key": "user_management", "label": "Пользователи и роли"},
]

EDITABLE_FIELD_PERMISSIONS = [
    {"key": "user.create", "label": "Создание пользователя"},
    {"key": "user.username", "label": "Изменение логина пользователя"},
    {"key": "user.password", "label": "Изменение пароля пользователя"},
    {"key": "user.role", "label": "Изменение роли пользователя"},
    {"key": "user.delete", "label": "Удаление пользователя"},
    {"key": "user.block", "label": "Блокировка пользователя"},
    {"key": "role.create", "label": "Создание роли"},
    {"key": "role.name", "label": "Изменение названия роли"},
    {"key": "role.description", "label": "Изменение описания роли"},
    {"key": "role.pages", "label": "Настройка доступа к страницам"},
    {"key": "role.fields.edit", "label": "Настройка прав редактирования полей"},
    {"key": "role.fields.view", "label": "Настройка прав просмотра полей"},
    {"key": "role.delete", "label": "Удаление роли"},
]

VIEWABLE_FIELD_PERMISSIONS = [
    {"key": "user.password", "label": "Просмотр пароля пользователя"},
]

DEFAULT_ROLE_PERMISSIONS = {
    "admin": {
        "pages": [PERMISSION_WILDCARD],
        "fields": {
            "edit": [PERMISSION_WILDCARD],
            "view": [PERMISSION_WILDCARD],
        },
        "description": "Полный доступ ко всем настройкам панели",
    },
    "user": {
        "pages": [
            item["key"]
            for item in AVAILABLE_PAGE_PERMISSIONS
            if item["key"] not in {"settings", "user_management"}
        ],
        "fields": {"edit": [], "view": []},
        "description": "Базовая роль без доступа к настройкам",
    },
}


def _normalize_permission_list(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        raw = [value]
    elif isinstance(value, (set, tuple)):
        raw = list(value)
    elif isinstance(value, list):
        raw = value
    else:
        return []
    result: list[str] = []
    for item in raw:
        key = str(item or "").strip()
        if key:
            result.append(key)
    return result


def _normalize_permissions_payload(raw) -> dict:
    if not raw:
        return {"pages": [], "fields": {"edit": [], "view": []}}
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            raw = {}
    if not isinstance(raw, dict):
        return {"pages": [], "fields": {"edit": [], "view": []}}
    pages = _normalize_permission_list(raw.get("pages"))
    fields_payload = raw.get("fields") or {}
    if isinstance(fields_payload, str):
        try:
            fields_payload = json.loads(fields_payload)
        except json.JSONDecodeError:
            fields_payload = {}
    if not isinstance(fields_payload, dict):
        fields_payload = {}
    edit_list = _normalize_permission_list(fields_payload.get("edit"))
    view_list = _normalize_permission_list(fields_payload.get("view"))
    return {"pages": pages, "fields": {"edit": edit_list, "view": view_list}}


def _permission_matches(values: list[str], key: str) -> bool:
    for value in values or []:
        if value == PERMISSION_WILDCARD:
            return True
        candidate = str(value or "").strip()
        if not candidate:
            continue
        if candidate.endswith(".*") and key.startswith(candidate[:-1]):
            return True
        if candidate == key:
            return True
    return False


def _merge_permission_lists(base: list[str], extra: list[str]) -> list[str]:
    result: list[str] = []
    for item in (base or []) + (extra or []):
        key = str(item or "").strip()
        if key and key not in result:
            result.append(key)
    return result


def _sanitize_permission_values(values: list[str], allowed_keys: list[str]) -> list[str]:
    allowed = set(allowed_keys or [])
    result: list[str] = []
    for item in _normalize_permission_list(values):
        if item == PERMISSION_WILDCARD or item in allowed:
            if item not in result:
                result.append(item)
    return result


def _ensure_user_schema():
    with get_users_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT,
                permissions TEXT NOT NULL DEFAULT '{}'
            )
            """
        )
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_roles_name ON roles(name)"
        )

        user_columns = {
            row["name"]: row for row in conn.execute("PRAGMA table_info(users)")
        }
        if "role_id" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN role_id INTEGER")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_users_role_id ON users(role_id)"
            )

        # Если пароль был захэширован и лежит в столбце password, перенесём его в password_hash
        rows = conn.execute(
            "SELECT id, password, password_hash FROM users"
        ).fetchall()
        for row in rows:
            password_value = row["password"] or ""
            password_hash = row["password_hash"] or ""
            if password_value and not password_hash and password_value.startswith(
                ("pbkdf2:", "scrypt:", "sha256:")
            ):
                conn.execute(
                    "UPDATE users SET password_hash = ?, password = NULL WHERE id = ?",
                    (password_value, row["id"]),
                )

        existing_roles = {
            row["name"]: row
            for row in conn.execute("SELECT id, name, description, permissions FROM roles")
        }

        for role_name, payload in DEFAULT_ROLE_PERMISSIONS.items():
            if role_name not in existing_roles:
                permissions_payload = {
                    "pages": payload.get("pages", []),
                    "fields": payload.get("fields", {}),
                }
                conn.execute(
                    "INSERT INTO roles (name, description, permissions) VALUES (?, ?, ?)",
                    (
                        role_name,
                        payload.get("description", ""),
                        json.dumps(permissions_payload, ensure_ascii=False),
                    ),
                )

        conn.commit()

        # После добавления ролей получим их актуальный список
        existing_roles = {
            row["name"]: row
            for row in conn.execute("SELECT id, name, permissions FROM roles")
        }

        def assign_role_if_needed(user_row, target_role_name: str):
            role_row = existing_roles.get(target_role_name)
            if not role_row:
                return
            current_role_id = user_row["role_id"] if "role_id" in user_row.keys() else None
            if current_role_id == role_row["id"]:
                return
            conn.execute(
                "UPDATE users SET role_id = ?, role = ? WHERE id = ?",
                (role_row["id"], target_role_name, user_row["id"]),
            )

        # Убедимся, что существует пользователь admin
        admin_row = conn.execute(
            "SELECT id, username, password, password_hash, role_id, role FROM users WHERE username = 'admin'"
        ).fetchone()
        admin_role_row = existing_roles.get("admin")
        if not admin_row:
            default_password = "admin"
            hashed = generate_password_hash(default_password)
            conn.execute(
                "INSERT INTO users (username, password, password_hash, role_id, role) VALUES (?, ?, ?, ?, ?)",
                (
                    "admin",
                    default_password,
                    hashed,
                    admin_role_row["id"] if admin_role_row else None,
                    "admin",
                ),
            )
        else:
            if admin_role_row:
                assign_role_if_needed(admin_row, "admin")
            if not (admin_row["password_hash"] or "").strip():
                current_plain = admin_row["password"] or "admin"
                conn.execute(
                    "UPDATE users SET password_hash = ? WHERE id = ?",
                    (generate_password_hash(current_plain), admin_row["id"]),
                )

        # Пользователи с ролью user должны получить одноимённую запись
        user_role_row = existing_roles.get("user")
        if user_role_row:
            users_without_role = conn.execute(
                "SELECT id, role, role_id FROM users WHERE role_id IS NULL"
            ).fetchall()
            for row in users_without_role:
                target_name = row["role"] or "user"
                role_row = existing_roles.get(target_name) or user_role_row
                conn.execute(
                    "UPDATE users SET role_id = ?, role = ? WHERE id = ?",
                    (role_row["id"], role_row["name"], row["id"]),
                )

        conn.commit()


_ensure_user_schema()


def _get_permissions_catalog() -> dict:
    return {
        "pages": AVAILABLE_PAGE_PERMISSIONS,
        "fields": {
            "edit": EDITABLE_FIELD_PERMISSIONS,
            "view": VIEWABLE_FIELD_PERMISSIONS,
        },
    }


def _empty_permissions() -> dict:
    return {"pages": [], "fields": {"edit": [], "view": []}}


def _load_user_row(user_id: int | None = None, username: str | None = None):
    if user_id is None and not username:
        return None
    with get_users_db() as conn:
        if user_id is not None:
            row = conn.execute(
                """
                SELECT u.*, r.name AS role_name, r.permissions AS role_permissions
                FROM users u
                LEFT JOIN roles r ON r.id = u.role_id
                WHERE u.id = ?
                """,
                (user_id,),
            ).fetchone()
        else:
            row = conn.execute(
                """
                SELECT u.*, r.name AS role_name, r.permissions AS role_permissions
                FROM users u
                LEFT JOIN roles r ON r.id = u.role_id
                WHERE LOWER(u.username) = LOWER(?)
                """,
                (username,),
            ).fetchone()
    return row


def _load_role_row(role_id: int | None = None, role_name: str | None = None):
    if role_id is None and not role_name:
        return None
    query = "SELECT id, name, description, permissions FROM roles WHERE {} LIMIT 1"
    if role_id is not None:
        clause = "id = ?"
        param = (role_id,)
    else:
        clause = "LOWER(name) = LOWER(?)"
        param = (role_name,)
    with get_users_db() as conn:
        return conn.execute(query.format(clause), param).fetchone()


def _row_value(row: sqlite3.Row | dict, key: str, default=None):
    if isinstance(row, dict):
        return row.get(key, default)
    if hasattr(row, "keys") and key in row.keys():
        return row[key]
    return default


def _serialize_role(row: sqlite3.Row | dict) -> dict:
    if not row:
        return {}
    permissions_raw = _row_value(row, "permissions", {})
    permissions = _normalize_permissions_payload(permissions_raw)
    name = _row_value(row, "name", "")
    description = _row_value(row, "description", "")
    role_id = _row_value(row, "id")
    return {
        "id": role_id,
        "name": name,
        "description": description or "",
        "permissions": permissions,
        "is_admin": (name or "").lower() == "admin",
    }


def _build_permissions_from_row(row: sqlite3.Row | dict | None) -> dict:
    if not row:
        return _empty_permissions()
    if isinstance(row, dict):
        permissions_raw = row.get("role_permissions") or row.get("permissions")
    else:
        if hasattr(row, "keys") and "role_permissions" in row.keys():
            permissions_raw = row["role_permissions"]
        else:
            permissions_raw = _row_value(row, "permissions", {})
    return _normalize_permissions_payload(permissions_raw)


def _permission_values_for_action(action: str) -> list[dict]:
    if action == "edit":
        return EDITABLE_FIELD_PERMISSIONS
    if action == "view":
        return VIEWABLE_FIELD_PERMISSIONS
    return []


def _catalog_keys_for_action(action: str) -> list[str]:
    return [item["key"] for item in _permission_values_for_action(action)]


def _prepare_permissions_payload(data: dict | None) -> dict:
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            data = {}
    if not isinstance(data, dict):
        data = {}
    allowed_pages = [item["key"] for item in AVAILABLE_PAGE_PERMISSIONS]
    pages = _sanitize_permission_values(data.get("pages") or [], allowed_pages)
    fields_payload = data.get("fields") or {}
    if isinstance(fields_payload, str):
        try:
            fields_payload = json.loads(fields_payload)
        except json.JSONDecodeError:
            fields_payload = {}
    if not isinstance(fields_payload, dict):
        fields_payload = {}
    edit_permissions = _sanitize_permission_values(
        fields_payload.get("edit") or [], _catalog_keys_for_action("edit")
    )
    view_permissions = _sanitize_permission_values(
        fields_payload.get("view") or [], _catalog_keys_for_action("view")
    )
    return {
        "pages": pages,
        "fields": {
            "edit": edit_permissions,
            "view": view_permissions,
        },
    }


def _is_builtin_role(name: str | None) -> bool:
    return (name or "").lower() == "admin"


def _get_current_permissions() -> dict:
    permissions = getattr(g, "current_permissions", None)
    if not permissions:
        return _empty_permissions()
    return permissions


def _current_user_id() -> int | None:
    user_row = getattr(g, "current_user", None)
    if user_row is not None:
        return _row_value(user_row, "id")
    return session.get("user_id")


def _can_edit_user_profile(user_id: int) -> bool:
    current_id = _current_user_id()
    if current_id is not None and current_id == user_id:
        return True
    return has_field_edit_permission("user.username")


def has_page_access(page_key: str) -> bool:
    permissions = _get_current_permissions()
    return _permission_matches(permissions.get("pages", []), page_key)


def has_field_edit_permission(field_key: str) -> bool:
    permissions = _get_current_permissions()
    return _permission_matches(permissions.get("fields", {}).get("edit", []), field_key)


def has_field_view_permission(field_key: str) -> bool:
    permissions = _get_current_permissions()
    return _permission_matches(permissions.get("fields", {}).get("view", []), field_key)


def require_field_permission(field_key: str, action: str = "edit"):
    allowed = (
        has_field_edit_permission(field_key)
        if action == "edit"
        else has_field_view_permission(field_key)
    )
    if not allowed:
        abort(403)


def _check_user_password(user_row, candidate: str) -> bool:
    if not user_row:
        return False
    password_hash = _row_value(user_row, "password_hash", "") or ""
    if password_hash:
        try:
            if check_password_hash(password_hash, candidate):
                return True
        except ValueError:
            pass
    stored_password = _row_value(user_row, "password", "") or ""
    if stored_password:
        return secrets.compare_digest(stored_password, candidate)
    return False


def _set_user_password(conn, user_id: int, new_password: str):
    if user_id is None:
        raise ValueError("user_id is required to set password")
    hashed = generate_password_hash(new_password)
    conn.execute(
        "UPDATE users SET password = ?, password_hash = ? WHERE id = ?",
        (new_password, hashed, user_id),
    )


@app.before_request
def _load_current_user_context():
    g.current_user = None
    g.current_permissions = _empty_permissions()
    user_id = session.get("user_id")
    username = session.get("username") or session.get("user")
    row = None
    if user_id is not None:
        row = _load_user_row(user_id=user_id)
    if row is None and username:
        row = _load_user_row(username=username)
    if row is None:
        return
    g.current_user = row
    permissions = _build_permissions_from_row(row)
    g.current_permissions = permissions
    session["user_id"] = _row_value(row, "id")
    session["user"] = _row_value(row, "username") or session.get("user")
    session["username"] = _row_value(row, "username") or session.get("username")
    session["role_id"] = _row_value(row, "role_id")
    session["role"] = (
        _row_value(row, "role_name")
        or _row_value(row, "role")
        or session.get("role")
    )


def page_access_required(page_key: str):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if "user" not in session:
                return redirect(url_for("login"))
            if not has_page_access(page_key):
                abort(403)
            return func(*args, **kwargs)

        return wrapper

    return decorator
def _init_sqlite():
    with get_db() as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA busy_timeout=5000;")
_init_sqlite()

def get_passport_db():
    conn = sqlite3.connect(OBJECT_PASSPORT_DB_PATH, timeout=30, isolation_level=None)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON;")
    conn.execute("PRAGMA busy_timeout=30000;")
    return conn


def ensure_object_passport_schema():
    with get_passport_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS object_passports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                department TEXT NOT NULL UNIQUE,
                business TEXT,
                partner_type TEXT,
                country TEXT,
                legal_entity TEXT,
                city TEXT,
                location_address TEXT,
                status TEXT,
                start_date TEXT,
                end_date TEXT,
                schedule_json TEXT,
                network TEXT,
                network_provider TEXT,
                network_restaurant_id TEXT,
                network_contract_number TEXT,
                network_legal_entity TEXT,
                network_support_phone TEXT,
                network_speed TEXT,
                network_tunnel TEXT,
                network_connection_params TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS object_passport_photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                passport_id INTEGER NOT NULL,
                category TEXT NOT NULL DEFAULT 'archive',
                caption TEXT,
                filename TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY(passport_id) REFERENCES object_passports(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS object_passport_equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                passport_id INTEGER NOT NULL,
                equipment_type TEXT,
                vendor TEXT,
                name TEXT,
                model TEXT,
                serial_number TEXT,
                status TEXT,
                ip_address TEXT,
                connection_type TEXT,
                connection_id TEXT,
                connection_password TEXT,
                description TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY(passport_id) REFERENCES object_passports(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS object_passport_status_periods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                passport_id INTEGER NOT NULL,
                status TEXT,
                started_at TEXT NOT NULL,
                ended_at TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY(passport_id) REFERENCES object_passports(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS object_passport_equipment_photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_id INTEGER NOT NULL,
                caption TEXT,
                filename TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY(equipment_id) REFERENCES object_passport_equipment(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS object_passport_network_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                passport_id INTEGER NOT NULL,
                original_name TEXT NOT NULL,
                filename TEXT NOT NULL,
                content_type TEXT,
                file_size INTEGER,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY(passport_id) REFERENCES object_passports(id) ON DELETE CASCADE
            )
            """
        )

ensure_object_passport_schema()

def _ensure_object_passport_columns():
    expected_columns = {
        "location_address": "TEXT",
        "network": "TEXT",
        "network_provider": "TEXT",
        "network_restaurant_id": "TEXT",
        "network_contract_number": "TEXT",
        "network_legal_entity": "TEXT",
        "network_support_phone": "TEXT",
        "network_speed": "TEXT",
        "network_tunnel": "TEXT",
        "network_connection_params": "TEXT",
        "status_task_id": "INTEGER",
        "suspension_date": "TEXT",
        "resume_date": "TEXT",
        "it_connection_type": "TEXT",
        "it_connection_id": "TEXT",
        "it_connection_password": "TEXT",
        "it_object_phone": "TEXT",
        "it_manager_name": "TEXT",
        "it_manager_phone": "TEXT",
        "it_iiko_server": "TEXT",
    }
    with get_passport_db() as conn:
        existing_columns = {
            row["name"]: row["type"].upper()
            for row in conn.execute("PRAGMA table_info(object_passports)").fetchall()
        }
        for column, column_type in expected_columns.items():
            if column not in existing_columns:
                conn.execute(f"ALTER TABLE object_passports ADD COLUMN {column} {column_type}")


_ensure_object_passport_columns()


def _ensure_object_passport_equipment_columns():
    expected_columns = {
        "equipment_type": "TEXT",
        "vendor": "TEXT",
        "connection_type": "TEXT",
        "connection_id": "TEXT",
        "connection_password": "TEXT",
        "serial_number": "TEXT",
    }
    with get_passport_db() as conn:
        existing_columns = {
            row["name"]: row["type"].upper()
            for row in conn.execute("PRAGMA table_info(object_passport_equipment)").fetchall()
        }
        for column, column_type in expected_columns.items():
            if column not in existing_columns:
                conn.execute(
                    f"ALTER TABLE object_passport_equipment ADD COLUMN {column} {column_type}"
                )


_ensure_object_passport_equipment_columns()


def ensure_knowledge_base_schema():
    with get_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS knowledge_articles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                department TEXT,
                article_type TEXT,
                status TEXT,
                author TEXT,
                direction TEXT,
                direction_subtype TEXT,
                summary TEXT,
                content TEXT,
                attachments TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            )
        """
        )
        existing_columns = {
            row["name"]: row["type"].upper()
            for row in conn.execute("PRAGMA table_info(knowledge_articles)").fetchall()
        }
        expected_columns = {
            "department": "TEXT",
            "article_type": "TEXT",
            "status": "TEXT",
            "author": "TEXT",
            "direction": "TEXT",
            "direction_subtype": "TEXT",
            "summary": "TEXT",
            "content": "TEXT",
            "attachments": "TEXT",
            "created_at": "TEXT DEFAULT (datetime('now'))",
            "updated_at": "TEXT DEFAULT (datetime('now'))",
        }
        for column, column_type in expected_columns.items():
            if column not in existing_columns:
                conn.execute(
                    f"ALTER TABLE knowledge_articles ADD COLUMN {column} {column_type}"
                )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS knowledge_article_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                article_id INTEGER,
                draft_token TEXT,
                stored_path TEXT NOT NULL,
                original_name TEXT,
                mime_type TEXT,
                file_size INTEGER,
                uploaded_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY(article_id) REFERENCES knowledge_articles(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_kb_files_article ON knowledge_article_files(article_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_kb_files_draft ON knowledge_article_files(draft_token)"
        )


ensure_knowledge_base_schema()

PASSPORT_STATUSES = (
    "Стройка",
    "Открыт",
    "Закрыт",
    "Реконструкция",
    "Заморожен",
    "Другое",
)

PASSPORT_STATUSES_REQUIRING_TASK = {"Закрыт", "Реконструкция", "Заморожен", "Другое"}

PASSPORT_STATUSES_WITH_PERIODS = {"Заморожен", "Реконструкция", "Другое", "Закрыт"}

WEEKDAY_SEQUENCE = [
    ("monday", "Понедельник", "Пн"),
    ("tuesday", "Вторник", "Вт"),
    ("wednesday", "Среда", "Ср"),
    ("thursday", "Четверг", "Чт"),
    ("friday", "Пятница", "Пт"),
    ("saturday", "Суббота", "Сб"),
    ("sunday", "Воскресенье", "Вс"),
]

WEEKDAY_FULL_LABELS = {key: full for key, full, _ in WEEKDAY_SEQUENCE}
WEEKDAY_SHORT_LABELS = {key: short for key, _, short in WEEKDAY_SEQUENCE}
WEEKDAY_ORDER = [key for key, *_ in WEEKDAY_SEQUENCE]


def _empty_schedule_template():
    return [
        {"day": key, "from": None, "to": None, "is_24": False}
        for key in WEEKDAY_ORDER
    ]


def _normalize_schedule(data):
    if not isinstance(data, list):
        return _empty_schedule_template()

    seen = {}
    for entry in data:
        day = entry.get("day") if isinstance(entry, dict) else None
        if day not in WEEKDAY_ORDER:
            continue
        start = entry.get("from")
        end = entry.get("to")
        start = str(start).strip() if isinstance(start, str) else start
        end = str(end).strip() if isinstance(end, str) else end
        start = start or None
        end = end or None
        is_24 = bool(entry.get("is_24"))
        if is_24:
            start = None
            end = None
        seen[day] = {"day": day, "from": start, "to": end, "is_24": bool(is_24)}

    normalized = []
    for key in WEEKDAY_ORDER:
        normalized.append(
            seen.get(key, {"day": key, "from": None, "to": None, "is_24": False})
        )
    return normalized


def _load_schedule(raw):
    if not raw:
        return _empty_schedule_template()
    try:
        data = json.loads(raw)
    except Exception:
        return _empty_schedule_template()
    schedule = _normalize_schedule(data)
    return schedule


def _schedule_to_json(schedule):
    normalized = _normalize_schedule(schedule)
    return json.dumps(normalized, ensure_ascii=False)


def _format_schedule(schedule):
    if not schedule:
        return "—"

    parts = []
    for entry in schedule:
        day = entry.get("day")
        if day not in WEEKDAY_ORDER:
            continue
        label = WEEKDAY_SHORT_LABELS.get(day, day)
        if entry.get("is_24"):
            parts.append(f"{label}: 24 часа")
        elif entry.get("from") and entry.get("to"):
            parts.append(f"{label}: {entry['from']}–{entry['to']}")
        elif entry.get("from") or entry.get("to"):
            start = entry.get("from") or "—"
            end = entry.get("to") or "—"
            parts.append(f"{label}: {start}–{end}")
    return "; ".join(parts) if parts else "—"

def _plural_form_ru(value, forms):
    """Russian pluralization helper."""
    try:
        n = abs(int(value))
    except (TypeError, ValueError):
        n = 0
    if n % 10 == 1 and n % 100 != 11:
        return forms[0]
    if 2 <= n % 10 <= 4 and not 12 <= n % 100 <= 14:
        return forms[1]
    return forms[2]

def _format_total_time(start_date, end_date):
    """Возвращает строку с общей длительностью работы.

    Теперь формат включает грубую оценку в годах/месяцах/днях и
    прежний формат (дни/часы/минуты) в скобках, чтобы не потерять
    первоначальную точность отображения."""

    if not start_date:
        return "—"
    try:
        start = dt.fromisoformat(str(start_date))
    except Exception:
        return "—"

    if end_date:
        try:
            end = dt.fromisoformat(str(end_date))
        except Exception:
            end = dt.now()
    else:
        end = dt.now()

    if end < start:
        end = start

    delta = end - start

    # «Старый» формат (дни/часы/минуты) сохраняем для точности в скобках.
    days_total = delta.days
    hours = delta.seconds // 3600
    minutes = (delta.seconds % 3600) // 60

    legacy_parts = []
    if days_total:
        legacy_parts.append(f"{days_total} д")
    if hours:
        legacy_parts.append(f"{hours} ч")
    if minutes:
        legacy_parts.append(f"{minutes} мин")
    if not legacy_parts:
        legacy_parts.append("0 мин")
    legacy_display = " ".join(legacy_parts)

    # Грубое представление в годах/месяцах/днях.
    remaining_days = days_total
    years = remaining_days // 365
    remaining_days -= years * 365
    months = remaining_days // 30
    remaining_days -= months * 30

    extended_parts = []
    if years:
        years_label = _plural_form_ru(years, ("год", "года", "лет"))
        extended_parts.append(f"{years} {years_label}")
    if months:
        extended_parts.append(f"{months} мес")
    if remaining_days or not extended_parts:
        extended_parts.append(f"{remaining_days} д")

    return f"{' '.join(extended_parts)} ({legacy_display})"


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return dt.fromisoformat(str(value))
    except Exception:
        try:
            return dt.strptime(str(value), "%Y-%m-%d")
        except Exception:
            return None


def _format_display_date(value):
    parsed = _parse_iso_date(value)
    if not parsed:
        return ""
    return parsed.strftime("%d.%m.%Y")


def _format_status_history_total(history):
    total = timedelta()
    for entry in history or []:
        start = _parse_iso_date(entry.get("started_at"))
        if not start:
            continue
        end_value = entry.get("ended_at")
        end = _parse_iso_date(end_value) if end_value else dt.now()
        if end < start:
            end = start
        total += end - start

    if not total or (total.days == 0 and total.seconds == 0):
        return "—"

    days = total.days
    hours = total.seconds // 3600
    minutes = (total.seconds % 3600) // 60
    parts = []
    if days:
        parts.append(f"{days} д")
    if hours:
        parts.append(f"{hours} ч")
    if minutes:
        parts.append(f"{minutes} мин")
    return " ".join(parts) if parts else "0 мин"


def _format_display_datetime(value):
    if not value:
        return ""
    try:
        parsed = dt.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        try:
            parsed = datetime.datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)
    return parsed.strftime("%d.%m.%Y %H:%M")

def _format_file_size(value):
    try:
        size = int(value)
    except (TypeError, ValueError):
        return ""

    units = ["Б", "КБ", "МБ", "ГБ", "ТБ"]
    size_float = float(size)
    for unit in units:
        if size_float < 1024 or unit == units[-1]:
            if unit == "Б":
                return f"{int(size_float)} {unit}"
            return f"{size_float:.1f} {unit}".replace(".0", "")
        size_float /= 1024
    return f"{size_float:.1f} ТБ"


def _calculate_department_case_minutes(department):
    if not department:
        return 0
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT SUM(
                CASE
                    WHEN t.status = 'resolved'
                         AND t.resolved_at IS NOT NULL
                         AND m.created_date IS NOT NULL
                         AND m.created_time IS NOT NULL
                    THEN (julianday(t.resolved_at) - julianday(m.created_date || ' ' || m.created_time)) * 24 * 60
                    ELSE 0
                END
            ) AS total_minutes
            FROM tickets t
            JOIN messages m ON m.ticket_id = t.ticket_id
            WHERE LOWER(TRIM(m.location_name)) = LOWER(TRIM(?))
            """,
            (department,),
        ).fetchone()
        total = row["total_minutes"] if row and row["total_minutes"] is not None else 0
        if total is None:
            return 0
        return max(int(round(total)), 0)
    finally:
        conn.close()


def _calculate_department_task_minutes(department):
    if not department:
        return 0
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT SUM(
                CASE
                    WHEN t.closed_at IS NOT NULL AND t.created_at IS NOT NULL
                    THEN (julianday(t.closed_at) - julianday(t.created_at)) * 24 * 60
                    ELSE 0
                END
            ) AS total_minutes
            FROM tasks t
            JOIN task_links tl ON tl.task_id = t.id
            JOIN messages m ON m.ticket_id = tl.ticket_id
            WHERE LOWER(TRIM(m.location_name)) = LOWER(TRIM(?))
            """,
            (department,),
        ).fetchone()
        total = row["total_minutes"] if row and row["total_minutes"] is not None else 0
        if total is None:
            return 0
        return max(int(round(total)), 0)
    finally:
        conn.close()


def _fetch_cases_for_department(department, limit=200):
    if not department:
        return []
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT ticket_id, business, location_type, city, location_name, problem, created_at
            FROM messages
            WHERE LOWER(TRIM(location_name)) = LOWER(TRIM(?))
            ORDER BY datetime(created_at) DESC
            LIMIT ?
            """,
            (department, limit),
        ).fetchall()
        cases = []
        for row in rows:
            cases.append(
                {
                    "ticket_id": row["ticket_id"],
                    "business": row["business"],
                    "location_type": row["location_type"],
                    "city": row["city"],
                    "location_name": row["location_name"],
                    "problem": row["problem"],
                    "created_at": row["created_at"],
                    "created_at_display": _format_display_datetime(row["created_at"]),
                }
            )
        return cases
    finally:
        conn.close()

def _fetch_tasks_for_department(department, search=None, limit=200):
    if not department:
        return []

    try:
        limit_value = int(limit)
    except (TypeError, ValueError):
        limit_value = 200
    limit_value = max(1, min(limit_value, 200))

    conn = get_db()
    try:
        conn.row_factory = sqlite3.Row
        params = [department]
        query = """
            SELECT DISTINCT
                t.id,
                t.seq,
                t.source,
                t.title,
                t.status,
                t.assignee,
                t.tag,
                t.due_at,
                t.last_activity_at
            FROM tasks t
            JOIN task_links tl ON tl.task_id = t.id
            JOIN messages m ON m.ticket_id = tl.ticket_id
            WHERE LOWER(TRIM(m.location_name)) = LOWER(TRIM(?))
        """

        search_value = (search or "").strip().lower()
        if search_value:
            like = f"%{search_value}%"
            query += """
                AND (
                    LOWER(COALESCE(t.title, '')) LIKE ?
                    OR LOWER(COALESCE(t.assignee, '')) LIKE ?
                    OR LOWER(COALESCE(t.tag, '')) LIKE ?
                    OR LOWER(COALESCE(t.status, '')) LIKE ?
                    OR LOWER(COALESCE(t.source, '')) || '_' || CAST(t.seq AS TEXT) LIKE ?
                    OR CAST(t.seq AS TEXT) LIKE ?
                    OR CAST(t.id AS TEXT) LIKE ?
                )
            """
            params.extend([like, like, like, like, like, like, like])

        query += " ORDER BY datetime(COALESCE(t.last_activity_at, t.created_at)) DESC LIMIT ?"
        params.append(limit_value)
        rows = conn.execute(query, params).fetchall()

        tasks = []
        for row in rows:
            tasks.append(
                {
                    "id": row["id"],
                    "display_no": _display_no(row["source"], row["seq"]),
                    "title": row["title"] or "",
                    "status": row["status"] or "",
                    "assignee": row["assignee"] or "",
                    "tag": row["tag"] or "",
                    "due_at": row["due_at"],
                    "due_at_display": _format_display_datetime(row["due_at"]),
                    "last_activity_at": row["last_activity_at"],
                    "last_activity_display": _format_display_datetime(row["last_activity_at"]),
                }
            )
        return tasks
    finally:
        conn.close()


def _ensure_single_title_photo(conn, passport_id, keep_photo_id):
    if keep_photo_id is None:
        return
    conn.execute(
        """
        UPDATE object_passport_photos
        SET category = 'archive', updated_at = datetime('now')
        WHERE passport_id = ?
          AND id != ?
          AND LOWER(COALESCE(category, '')) = 'title'
        """,
        (passport_id, keep_photo_id),
    )


def _fetch_passport_photos(conn, passport_id):
    rows = conn.execute(
        """
        SELECT id, passport_id, category, caption, filename, created_at
        FROM object_passport_photos
        WHERE passport_id = ?
        ORDER BY CASE WHEN LOWER(COALESCE(category, '')) = 'title' THEN 0 ELSE 1 END,
                 datetime(created_at) DESC,
                 id DESC
        """,
        (passport_id,),
    ).fetchall()
    title_ids = [row["id"] for row in rows if (row["category"] or "").lower() == "title"]
    if len(title_ids) > 1:
        _ensure_single_title_photo(conn, passport_id, title_ids[0])
        conn.commit()
        rows = conn.execute(
            """
            SELECT id, passport_id, category, caption, filename, created_at
            FROM object_passport_photos
            WHERE passport_id = ?
            ORDER BY CASE WHEN LOWER(COALESCE(category, '')) = 'title' THEN 0 ELSE 1 END,
                     datetime(created_at) DESC,
                     id DESC
            """,
            (passport_id,),
        ).fetchall()
    items = []
    for row in rows:
        items.append(
            {
                "id": row["id"],
                "passport_id": row["passport_id"],
                "category": row["category"],
                "caption": row["caption"] or "",
                "filename": row["filename"],
                "url": url_for("object_passport_media", filename=row["filename"]),
            }
        )
    return items

def _fetch_network_files(conn, passport_id):
    rows = conn.execute(
        """
        SELECT id, passport_id, original_name, filename, content_type, file_size, created_at
        FROM object_passport_network_files
        WHERE passport_id = ?
        ORDER BY created_at DESC
        """,
        (passport_id,),
    ).fetchall()
    files = []
    for row in rows:
        files.append(
            {
                "id": row["id"],
                "passport_id": row["passport_id"],
                "original_name": row["original_name"],
                "filename": row["filename"],
                "content_type": row["content_type"] or "",
                "size": row["file_size"] or 0,
                "size_display": _format_file_size(row["file_size"]),
                "url": url_for("object_passport_media", filename=row["filename"]),
                "download_url": url_for(
                    "api_object_passport_download_network_file", file_id=row["id"]
                ),
                "created_at": row["created_at"],
                "created_at_display": _format_display_datetime(row["created_at"]),
            }
        )
    return files


def _fetch_equipment_photos(conn, equipment_id):
    rows = conn.execute(
        """
        SELECT id, equipment_id, caption, filename
        FROM object_passport_equipment_photos
        WHERE equipment_id = ?
        ORDER BY created_at DESC
        """,
        (equipment_id,),
    ).fetchall()
    photos = []
    for row in rows:
        photos.append(
            {
                "id": row["id"],
                "equipment_id": row["equipment_id"],
                "caption": row["caption"] or "",
                "filename": row["filename"],
                "url": url_for("object_passport_media", filename=row["filename"]),
            }
        )
    return photos


def _fetch_passport_equipment(conn, passport_id):
    rows = conn.execute(
        """
        SELECT id, passport_id, equipment_type, vendor, name, model, serial_number, status, ip_address,
               connection_type, connection_id, connection_password, description
        FROM object_passport_equipment
        WHERE passport_id = ?
        ORDER BY LOWER(COALESCE(name, ''))
        """,
        (passport_id,),
    ).fetchall()
    equipment = []
    for row in rows:
        equipment.append(
            {
                "id": row["id"],
                "passport_id": row["passport_id"],
                "equipment_type": row["equipment_type"] or "",
                "vendor": row["vendor"] or "",
                "name": row["name"] or "",
                "model": row["model"] or "",
                "serial_number": row["serial_number"] or "",
                "status": row["status"] or "",
                "ip_address": row["ip_address"] or "",
                "connection_type": row["connection_type"] or "",
                "connection_id": row["connection_id"] or "",
                "connection_password": row["connection_password"] or "",
                "description": row["description"] or "",
                "photos": _fetch_equipment_photos(conn, row["id"]),
            }
        )
    return equipment


def _fetch_status_periods(conn, passport_id):
    rows = conn.execute(
        """
        SELECT id, status, started_at, ended_at
        FROM object_passport_status_periods
        WHERE passport_id = ?
        ORDER BY started_at DESC, id DESC
        """,
        (passport_id,),
    ).fetchall()
    history = []
    for row in rows:
        started_at = row["started_at"] or ""
        ended_at = row["ended_at"] or ""
        history.append(
            {
                "id": row["id"],
                "status": row["status"] or "",
                "started_at": started_at,
                "ended_at": ended_at,
                "started_display": _format_display_date(started_at),
                "ended_display": _format_display_date(ended_at) if ended_at else "",
                "duration_display": _format_total_time(started_at, ended_at),
                "is_active": not bool(ended_at),
            }
        )
    return history


def _sync_status_period(conn, passport_id, status, suspension_date, resume_date):
    start = (suspension_date or "").strip()
    end = (resume_date or "").strip()
    if not start and not end:
        return

    open_row = conn.execute(
        """
        SELECT id, started_at, status
        FROM object_passport_status_periods
        WHERE passport_id = ? AND ended_at IS NULL
        ORDER BY started_at DESC, id DESC
        LIMIT 1
        """,
        (passport_id,),
    ).fetchone()

    if end and open_row:
        conn.execute(
            """
            UPDATE object_passport_status_periods
            SET ended_at = ?, status = COALESCE(?, status), updated_at = datetime('now')
            WHERE id = ?
            """,
            (end, status or open_row["status"], open_row["id"]),
        )
        open_row = None

    if not start:
        return

    if open_row and not end:
        assignments = []
        params = {"id": open_row["id"]}
        if open_row["started_at"] != start:
            assignments.append("started_at = :started_at")
            params["started_at"] = start
        if status and open_row["status"] != status:
            assignments.append("status = :status")
            params["status"] = status
        if assignments:
            assignments.append("updated_at = datetime('now')")
            conn.execute(
                f"UPDATE object_passport_status_periods SET {', '.join(assignments)} WHERE id = :id",
                params,
            )
        return

    existing = conn.execute(
        """
        SELECT id, ended_at, status
        FROM object_passport_status_periods
        WHERE passport_id = ? AND started_at = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (passport_id, start),
    ).fetchone()

    if existing:
        assignments = []
        params = {"id": existing["id"]}
        if end and (existing["ended_at"] or "") != end:
            assignments.append("ended_at = :ended_at")
            params["ended_at"] = end
        if status and existing["status"] != status:
            assignments.append("status = :status")
            params["status"] = status
        if assignments:
            assignments.append("updated_at = datetime('now')")
            conn.execute(
                f"UPDATE object_passport_status_periods SET {', '.join(assignments)} WHERE id = :id",
                params,
            )
        return

    if status not in PASSPORT_STATUSES_WITH_PERIODS and not end:
        return

    conn.execute(
        """
        INSERT INTO object_passport_status_periods (passport_id, status, started_at, ended_at)
        VALUES (?, ?, ?, ?)
        """,
        (passport_id, status or "", start, end or None),
    )


def _serialize_passport_row(row):
    schedule = _load_schedule(row["schedule_json"])
    return {
        "id": row["id"],
        "department": row["department"] or "",
        "business": row["business"] or "",
        "partner_type": row["partner_type"] or "",
        "country": row["country"] or "",
        "legal_entity": row["legal_entity"] or "",
        "city": row["city"] or "",
        "location_address": row["location_address"] or "",
        "status": row["status"] or "",
        "network": row["network"] or "",
        "network_provider": row["network_provider"] or "",
        "network_restaurant_id": row["network_restaurant_id"] or "",
        "network_contract_number": row["network_contract_number"] or "",
        "network_legal_entity": row["network_legal_entity"] or "",
        "network_tunnel": row["network_tunnel"] or "",
        "network_speed": row["network_speed"] or "",
        "it_connection_type": row["it_connection_type"] or "",
        "it_connection_id": row["it_connection_id"] or "",
        "it_connection_password": row["it_connection_password"] or "",
        "it_object_phone": row["it_object_phone"] or "",
        "it_manager_name": row["it_manager_name"] or "",
        "it_manager_phone": row["it_manager_phone"] or "",
        "it_iiko_server": row["it_iiko_server"] or "",
        "start_date": row["start_date"] or "",
        "end_date": row["end_date"] or "",
        "suspension_date": row["suspension_date"] or "",
        "resume_date": row["resume_date"] or "",
        "status_task_id": row["status_task_id"],
        "total_work_time": _format_total_time(row["start_date"], row["end_date"]),
        "schedule_display": _format_schedule(schedule),
    }


def _serialize_passport_detail(conn, row):
    schedule_raw = _load_schedule(row["schedule_json"])
    schedule_for_client = []
    for entry in schedule_raw:
        schedule_for_client.append(
            {
                "day": entry.get("day"),
                "from": entry.get("from") or "",
                "to": entry.get("to") or "",
                "is_24": bool(entry.get("is_24")),
                "label": WEEKDAY_FULL_LABELS.get(entry.get("day"), entry.get("day")),
            }
        )

    detail = {
        "id": row["id"],
        "department": row["department"] or "",
        "business": row["business"] or "",
        "partner_type": row["partner_type"] or "",
        "country": row["country"] or "",
        "legal_entity": row["legal_entity"] or "",
        "city": row["city"] or "",
        "location_address": row["location_address"] or "",
        "status": row["status"] or PASSPORT_STATUSES[0],
        "network": row["network"] or "",
        "network_provider": row["network_provider"] or "",
        "network_restaurant_id": row["network_restaurant_id"] or "",
        "network_contract_number": row["network_contract_number"] or "",
        "network_legal_entity": row["network_legal_entity"] or "",
        "network_support_phone": row["network_support_phone"] or "",
        "network_speed": row["network_speed"] or "",
        "network_tunnel": row["network_tunnel"] or "",
        "network_connection_params": row["network_connection_params"] or "",
        "it_connection_type": row["it_connection_type"] or "",
        "it_connection_id": row["it_connection_id"] or "",
        "it_connection_password": row["it_connection_password"] or "",
        "it_object_phone": row["it_object_phone"] or "",
        "it_manager_name": row["it_manager_name"] or "",
        "it_manager_phone": row["it_manager_phone"] or "",
        "it_iiko_server": row["it_iiko_server"] or "",
        "start_date": row["start_date"] or "",
        "end_date": row["end_date"] or "",
        "suspension_date": row["suspension_date"] or "",
        "resume_date": row["resume_date"] or "",
        "status_task_id": row["status_task_id"],
        "total_work_time": _format_total_time(row["start_date"], row["end_date"]),
        "schedule": schedule_for_client,
        "schedule_display": _format_schedule(schedule_raw),
        "photos": _fetch_passport_photos(conn, row["id"]),
        "network_files": _fetch_network_files(conn, row["id"]),
        "equipment": _fetch_passport_equipment(conn, row["id"]),
        "case_time_minutes": 0,
        "case_time_display": "—",
        "task_time_minutes": 0,
        "task_time_display": "—",
    }
    case_minutes = _calculate_department_case_minutes(detail["department"])
    detail["case_time_minutes"] = case_minutes
    detail["case_time_display"] = format_time_duration(case_minutes)
    task_minutes = _calculate_department_task_minutes(detail["department"])
    detail["task_time_minutes"] = task_minutes
    detail["task_time_display"] = format_time_duration(task_minutes)
    status_history = _fetch_status_periods(conn, row["id"])
    if not status_history and (row["suspension_date"] or row["resume_date"]):
        _sync_status_period(
            conn,
            row["id"],
            row["status"],
            row["suspension_date"],
            row["resume_date"],
        )
        status_history = _fetch_status_periods(conn, row["id"])
    detail["status_history"] = status_history
    detail["status_history_total"] = _format_status_history_total(status_history)
    detail["status_task"] = None
    if row["status_task_id"]:
        with get_db() as main_conn:
            main_conn.row_factory = sqlite3.Row
            task_row = main_conn.execute(
                "SELECT id, seq, source, title, status, assignee, due_at, last_activity_at FROM tasks WHERE id = ?",
                (row["status_task_id"],),
            ).fetchone()
            if task_row:
                detail["status_task"] = {
                    "id": task_row["id"],
                    "display_no": _display_no(task_row["source"], task_row["seq"]),
                    "title": task_row["title"] or "",
                    "status": task_row["status"] or "",
                    "assignee": task_row["assignee"] or "",
                    "due_at": task_row["due_at"],
                    "due_at_display": _format_display_datetime(task_row["due_at"]),
                    "last_activity": task_row["last_activity_at"],
                    "last_activity_display": _format_display_datetime(task_row["last_activity_at"]),
                }
    detail["cases"] = _fetch_cases_for_department(detail["department"])
    detail["tasks"] = _fetch_tasks_for_department(detail["department"])
    return detail


def _fetch_passport_rows(filters):
    filters = filters or {}
    if hasattr(filters, "copy"):
        filters_local = filters.copy()
    else:
        filters_local = dict(filters)

    contract_alias = (
        (filters_local.get("network_contract_number") or "").strip()
        or (filters_local.get("contract_number") or "").strip()
        or (filters_local.get("contract") or "").strip()
    )
    if contract_alias and not (filters_local.get("network_contract_number") or "").strip():
        filters_local["network_contract_number"] = contract_alias

    it_connection_alias = (filters_local.get("it_connection") or "").strip()
    if it_connection_alias and not (filters_local.get("it_connection_type") or "").strip():
        filters_local["it_connection_type"] = it_connection_alias

    iiko_alias = (filters_local.get("iiko_server") or "").strip()
    if iiko_alias and not (filters_local.get("it_iiko_server") or "").strip():
        filters_local["it_iiko_server"] = iiko_alias

    equipment_aliases = {
        "equipment_type": ["equipment_type", "equipmentType"],
        "equipment_vendor": ["equipment_vendor", "equipmentVendor", "vendor"],
        "equipment_model": ["equipment_model", "equipmentModel", "model"],
        "equipment_status": ["equipment_status", "equipmentStatus", "status_equipment"],
    }
    for canonical, aliases in equipment_aliases.items():
        current = (filters_local.get(canonical) or "").strip()
        if current:
            continue
        for alias in aliases[1:]:
            alias_value = (filters_local.get(alias) or "").strip()
            if alias_value:
                filters_local[canonical] = alias_value
                break

    conn = get_passport_db()
    try:
        query = "SELECT * FROM object_passports"
        conditions = []
        params = []

        search = (filters_local.get("search") or "").strip().lower()
        if search:
            like = f"%{search}%"
            searchable = [
                "business",
                "partner_type",
                "country",
                "legal_entity",
                "city",
                "department",
                "status",
                "network",
                "network_provider",
                "network_restaurant_id",
                "network_contract_number",
            ]
            conditions.append(
                "(" + " OR ".join([f"LOWER(COALESCE({col}, '')) LIKE ?" for col in searchable]) + ")"
            )
            params.extend([like] * len(searchable))

        for field in [
            "business",
            "partner_type",
            "country",
            "legal_entity",
            "city",
            "department",
            "status",
            "network_contract_number",
            "network_restaurant_id",
            "network_provider",
            "network",
            "it_connection_type",
            "it_iiko_server",
        ]:
            value = (filters_local.get(field) or "").strip()
            if value:
                conditions.append(f"LOWER(COALESCE({field}, '')) = LOWER(?)")
                params.append(value)

        equipment_filters = {
            "equipment_type": "equipment_type",
            "equipment_vendor": "vendor",
            "equipment_model": "model",
            "equipment_status": "status",
        }
        for field, column in equipment_filters.items():
            value = (filters_local.get(field) or "").strip()
            if value:
                conditions.append(
                    "EXISTS ("
                    "SELECT 1 FROM object_passport_equipment eq "
                    "WHERE eq.passport_id = object_passports.id "
                    f"AND LOWER(TRIM(COALESCE(eq.{column}, ''))) = LOWER(?)"
                    ")"
                )
                params.append(value)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY LOWER(COALESCE(department, ''))"
        rows = conn.execute(query, params).fetchall()
        return rows
    finally:
        conn.close()


def _parameter_values_for_passports():
    parameter_values = {key: [] for key in PARAMETER_TYPES.keys()}
    try:
        with get_db() as conn:
            grouped = _fetch_parameters_grouped(conn, include_deleted=False)
        for slug, items in grouped.items():
            values = []
            for item in items:
                if not item.get("value"):
                    continue
                if item.get("is_deleted"):
                    continue
                if slug == "it_connection":
                    category = (item.get("category") or "").strip()
                    if category and category != "equipment_type":
                        continue
                values.append(item["value"])
            parameter_values[slug] = values
    except Exception:
        pass
    return parameter_values


def _collect_it_equipment_options():
    options = {"types": [], "vendors": [], "models": [], "serials": [], "statuses": []}
    catalog_items = []
    try:
        with get_db() as conn:
            grouped = _fetch_parameters_grouped(conn, include_deleted=False)
            catalog_items = _fetch_it_equipment_catalog(conn)
        items = grouped.get("it_connection", []) if grouped else []
        for item in items:
            if not item or item.get("is_deleted"):
                continue
            category = (item.get("category") or "").strip()
            value = (item.get("value") or "").strip()
            if not value:
                continue
            if category == "equipment_type":
                if value not in options["types"]:
                    options["types"].append(value)
            elif category == "equipment_vendor":
                if value not in options["vendors"]:
                    options["vendors"].append(value)
            elif category == "equipment_model":
                if value not in options["models"]:
                    options["models"].append(value)
            elif category == "equipment_status":
                if value not in options["statuses"]:
                    options["statuses"].append(value)
        for catalog_item in catalog_items:
            equipment_type = (catalog_item.get("equipment_type") or "").strip()
            vendor = (catalog_item.get("equipment_vendor") or "").strip()
            model = (catalog_item.get("equipment_model") or "").strip()
            serial = (catalog_item.get("serial_number") or "").strip()
            if equipment_type and equipment_type not in options["types"]:
                options["types"].append(equipment_type)
            if vendor and vendor not in options["vendors"]:
                options["vendors"].append(vendor)
            if model and model not in options["models"]:
                options["models"].append(model)
            if serial and serial not in options["serials"]:
                options["serials"].append(serial)
    except Exception:
        pass
    return options

def _collect_remote_connection_options(settings: dict | None = None) -> list[str]:
    """Return list of remote connection types configured in settings."""

    try:
        with get_db() as conn:
            grouped = _fetch_parameters_grouped(conn, include_deleted=False)
    except Exception:
        grouped = {}

    seen: set[str] = set()
    options: list[str] = []
    items = grouped.get("remote_access", []) if isinstance(grouped, dict) else []
    for item in items:
        if not isinstance(item, dict):
            continue
        if item.get("is_deleted"):
            continue
        value = (item.get("value") or "").strip()
        if not value:
            continue
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        options.append(value)
    if not options:
        options = list(REMOTE_ACCESS_DEFAULTS)
    return options

def _serialize_it_equipment_row(row):
    return {
        "id": row["id"],
        "equipment_type": (row["equipment_type"] or "").strip(),
        "equipment_vendor": (row["equipment_vendor"] or "").strip(),
        "equipment_model": (row["equipment_model"] or "").strip(),
        "photo_url": (row["photo_url"] or "").strip(),
        "serial_number": (row["serial_number"] or "").strip(),
        "accessories": (row["accessories"] or "").strip(),
    }

def _fetch_it_equipment_catalog(conn):
    try:
        rows = conn.execute(
            """
            SELECT id, equipment_type, equipment_vendor, equipment_model, photo_url, serial_number, accessories
            FROM it_equipment_catalog
            ORDER BY equipment_type COLLATE NOCASE, equipment_vendor COLLATE NOCASE, equipment_model COLLATE NOCASE
            """
        ).fetchall()
    except sqlite3.OperationalError as exc:
        if _is_missing_table_error(exc, "it_equipment_catalog"):
            ensure_it_equipment_catalog_schema()
            return []
        raise
    return [_serialize_it_equipment_row(row) for row in rows]


def _city_options():
    cities = set()
    _, third_level = _collect_departments_from_locations()
    for city in third_level:
        if city:
            cities.add(city)

    try:
        with get_db() as conn:
            rows = conn.execute(
                "SELECT DISTINCT city FROM messages WHERE city IS NOT NULL AND TRIM(city) != ''"
            ).fetchall()
            for row in rows:
                cities.add(row["city"])
    except Exception:
        pass

    try:
        with get_passport_db() as conn:
            rows = conn.execute(
                "SELECT DISTINCT city FROM object_passports WHERE city IS NOT NULL AND TRIM(city) != ''"
            ).fetchall()
            for row in rows:
                cities.add(row["city"])
    except Exception:
        pass

    return sorted(cities, key=lambda name: name.lower())


def _blank_passport_detail():
    return {
        "id": None,
        "department": "",
        "business": "",
        "partner_type": "",
        "country": "",
        "legal_entity": "",
        "city": "",
        "location_address": "",
        "status": PASSPORT_STATUSES[0],
        "network": "",
        "network_provider": "",
        "network_restaurant_id": "",
        "network_contract_number": "",
        "network_legal_entity": "",
        "network_support_phone": "",
        "network_speed": "",
        "network_tunnel": "",
        "network_connection_params": "",
        "it_connection_type": "",
        "it_connection_id": "",
        "it_connection_password": "",
        "it_object_phone": "",
        "it_manager_name": "",
        "it_manager_phone": "",
        "it_iiko_server": "",
        "start_date": "",
        "end_date": "",
        "suspension_date": "",
        "resume_date": "",
        "status_task_id": None,
        "status_task": None,
        "status_history": [],
        "status_history_total": "—",
        "total_work_time": "—",
        "case_time_minutes": 0,
        "case_time_display": "—",
        "task_time_minutes": 0,
        "task_time_display": "—",
        "schedule": [
            {
                "day": key,
                "from": "",
                "to": "",
                "is_24": False,
                "label": WEEKDAY_FULL_LABELS.get(key, key),
            }
            for key in WEEKDAY_ORDER
        ],
        "schedule_display": "—",
        "photos": [],
        "network_files": [],
        "equipment": [],
        "cases": [],
        "tasks": [],
    }


def _render_passport_template(passport_detail, is_new):
    parameter_values = _parameter_values_for_passports()
    cities = _city_options()
    payload = dict(passport_detail)
    payload["is_new"] = is_new
    settings = load_settings()
    network_profiles = settings.get("network_profiles", [])
    it_equipment_options = _collect_it_equipment_options()
    try:
        with get_db() as conn:
            it_equipment_catalog = _fetch_it_equipment_catalog(conn)
    except Exception:
        it_equipment_catalog = []

    def _append_unique(target, value):
        if value and value not in target:
            target.append(value)

    provider_options = []
    contract_options = []
    support_phone_options = []
    speed_options = []
    legal_entity_options = []
    restaurant_id_options = []
    it_connection_options = []
    iiko_server_options = []

    for profile in network_profiles:
        if not isinstance(profile, dict):
            continue
        _append_unique(provider_options, (profile.get("provider") or "").strip())
        _append_unique(contract_options, (profile.get("contract_number") or "").strip())
        _append_unique(support_phone_options, (profile.get("support_phone") or "").strip())
        _append_unique(speed_options, (profile.get("speed") or "").strip())
        _append_unique(legal_entity_options, (profile.get("legal_entity") or "").strip())
        restaurant_ids = profile.get("restaurant_ids")
        appended = False
        if isinstance(restaurant_ids, (list, tuple)):
            for raw_value in restaurant_ids:
                value = (raw_value or "").strip()
                if value:
                    _append_unique(restaurant_id_options, value)
                    appended = True
        if not appended:
            _append_unique(restaurant_id_options, (profile.get("restaurant_id") or "").strip())

    for option in _collect_remote_connection_options(settings):
        _append_unique(it_connection_options, (option or "").strip())
    if not it_connection_options:
        for option in parameter_values.get("it_connection", []):
            _append_unique(it_connection_options, (option or "").strip())
    if not it_connection_options:
        it_connection_options = ["RMSviewer", "Anydes", "Ассистент", "VNC"]

    for option in parameter_values.get("iiko_server", []):
        _append_unique(iiko_server_options, (option or "").strip())
    return render_template(
        "object_passport_detail.html",
        passport_payload=json.dumps(payload, ensure_ascii=False),
        parameter_values=parameter_values,
        cities=cities,
        statuses=list(PASSPORT_STATUSES),
        day_labels=WEEKDAY_SEQUENCE,
        network_profiles=network_profiles,
        network_provider_options=provider_options,
        network_contract_options=contract_options,
        network_support_phone_options=support_phone_options,
        network_speed_options=speed_options,
        network_legal_entity_options=legal_entity_options,
        network_restaurant_id_options=restaurant_id_options,
        it_connection_options=it_connection_options,
        iiko_server_options=iiko_server_options,
        it_equipment_options=it_equipment_options,
        it_equipment_catalog=it_equipment_catalog,
    )


def _prepare_passport_record(payload):
    payload = payload or {}
    department = (payload.get("department") or "").strip()
    if not department:
        return None, None, "Поле «Департамент» обязательно"

    status = (payload.get("status") or PASSPORT_STATUSES[0]).strip()
    if status not in PASSPORT_STATUSES:
        return None, None, "Недопустимый статус"

    status_task_id = payload.get("status_task_id")
    if status_task_id in ("", None):
        status_task_id_int = None
    else:
        try:
            status_task_id_int = int(status_task_id)
        except (TypeError, ValueError):
            return None, None, "Некорректная задача для статуса"

        with get_db() as conn:
            exists = conn.execute(
                "SELECT 1 FROM tasks WHERE id = ?",
                (status_task_id_int,),
            ).fetchone()
        if not exists:
            return None, None, "Выбранная задача не найдена"

    if status in PASSPORT_STATUSES_REQUIRING_TASK and not status_task_id_int:
        return None, None, "Для выбранного статуса необходимо выбрать задачу"

    schedule_data = payload.get("schedule") or []
    schedule_normalized = _normalize_schedule(schedule_data)
    schedule_json = json.dumps(schedule_normalized, ensure_ascii=False)

    def clean(field):
        value = (payload.get(field) or "").strip()
        return value or None

    record = {
        "department": department,
        "business": clean("business"),
        "partner_type": clean("partner_type"),
        "country": clean("country"),
        "legal_entity": clean("legal_entity"),
        "city": clean("city"),
        "location_address": clean("location_address"),
        "status": status,
        "network": clean("network"),
        "network_provider": clean("network_provider"),
        "network_restaurant_id": clean("network_restaurant_id"),
        "network_contract_number": clean("network_contract_number"),
        "network_legal_entity": clean("network_legal_entity"),
        "network_support_phone": clean("network_support_phone"),
        "network_speed": clean("network_speed"),
        "network_tunnel": clean("network_tunnel"),
        "network_connection_params": clean("network_connection_params"),
        "it_connection_type": clean("it_connection_type"),
        "it_connection_id": clean("it_connection_id"),
        "it_connection_password": clean("it_connection_password"),
        "it_object_phone": clean("it_object_phone"),
        "it_manager_name": clean("it_manager_name"),
        "it_manager_phone": clean("it_manager_phone"),
        "it_iiko_server": clean("it_iiko_server"),
        "start_date": clean("start_date"),
        "end_date": clean("end_date"),
        "suspension_date": clean("suspension_date"),
        "resume_date": clean("resume_date"),
        "status_task_id": status_task_id_int,
        "schedule_json": schedule_json,
    }

    return record, schedule_normalized, None

def ensure_tasks_schema():
    with get_db() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS tasks(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq INTEGER NOT NULL,               -- порядковый номер (целый, инкремент)
            source TEXT,                        -- 'DL' или NULL
            title TEXT,
            body_html TEXT,
            creator TEXT,
            assignee TEXT,
            tag TEXT,
            status TEXT DEFAULT 'Новая',
            due_at TEXT,                        -- ISO
            created_at TEXT DEFAULT (datetime('now')),
            closed_at TEXT,                     -- ISO
            last_activity_at TEXT DEFAULT (datetime('now'))
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS task_people(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            role TEXT NOT NULL,                 -- 'co' или 'watcher'
            identity TEXT NOT NULL,
            UNIQUE(task_id, role, identity),
            FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS task_comments(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            author TEXT,
            html TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS task_history(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            at TEXT DEFAULT (datetime('now')),
            text TEXT,
            FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS ticket_active(
            ticket_id TEXT PRIMARY KEY,
            user      TEXT NOT NULL,
            last_seen TEXT DEFAULT (datetime('now'))
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS ticket_responsibles(
            ticket_id   TEXT PRIMARY KEY,
            responsible TEXT NOT NULL,
            assigned_at TEXT DEFAULT (datetime('now')),
            assigned_by TEXT
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS task_links(
            task_id   INTEGER NOT NULL,
            ticket_id TEXT    NOT NULL,
            PRIMARY KEY(task_id, ticket_id)
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS notifications(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT NOT NULL,
            text TEXT NOT NULL,
            url TEXT,
            is_read INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        )
        """)
        # автоинкремент для seq через служебную таблицу
        conn.execute("""
        CREATE TABLE IF NOT EXISTS task_seq(
            id INTEGER PRIMARY KEY CHECK (id=1),
            val INTEGER NOT NULL
        )
        """)
        cur = conn.execute("SELECT val FROM task_seq WHERE id=1").fetchone()
        if not cur:
            conn.execute("INSERT INTO task_seq(id,val) VALUES(1,0)")

ensure_tasks_schema()

def ensure_client_blacklist_schema():
    """
    Таблица блэклиста клиентов:
      - user_id (PK)
      - is_blacklisted: 0/1
      - reason: причина
      - added_at: ISO
      - added_by: оператор
      - unblock_requested: 0/1 (клиент нажал «запросить разблокировку»)
      - unblock_requested_at: ISO
    """
    try:
        with get_db() as conn:
            conn.execute("""
            CREATE TABLE IF NOT EXISTS client_blacklist (
                user_id TEXT PRIMARY KEY,
                is_blacklisted INTEGER NOT NULL DEFAULT 0,
                reason TEXT,
                added_at TEXT,
                added_by TEXT,
                unblock_requested INTEGER NOT NULL DEFAULT 0,
                unblock_requested_at TEXT
            )
            """)
    except Exception as e:
        print(f"ensure_client_blacklist_schema: {e}")

ensure_client_blacklist_schema()


def ensure_client_unblock_requests_schema():
    try:
        with get_db() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS client_unblock_requests (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL,
                    channel_id INTEGER,
                    reason TEXT,
                    created_at TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending',
                    decided_at TEXT,
                    decided_by TEXT,
                    decision_comment TEXT
                )
                """
            )
            conn.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_client_unblock_requests_user
                ON client_unblock_requests(user_id)
                """
            )
    except Exception as e:
        print(f"ensure_client_unblock_requests_schema: {e}")


ensure_client_unblock_requests_schema()


def ensure_settings_parameters_schema():
    try:
        with get_db() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS settings_parameters (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    param_type TEXT NOT NULL,
                    value TEXT NOT NULL,
                    state TEXT NOT NULL DEFAULT 'Активен',
                    is_deleted INTEGER NOT NULL DEFAULT 0,
                    deleted_at TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    UNIQUE(param_type, value)
                )
                """
            )
            columns = {
                row["name"]: row["type"].upper()
                for row in conn.execute("PRAGMA table_info(settings_parameters)").fetchall()
            }
            if "state" not in columns:
                conn.execute(
                    "ALTER TABLE settings_parameters ADD COLUMN state TEXT NOT NULL DEFAULT 'Активен'"
                )
            if "is_deleted" not in columns:
                conn.execute(
                    "ALTER TABLE settings_parameters ADD COLUMN is_deleted INTEGER NOT NULL DEFAULT 0"
                )
            if "deleted_at" not in columns:
                conn.execute(
                    "ALTER TABLE settings_parameters ADD COLUMN deleted_at TEXT"
                )
            if "extra_json" not in columns:
                conn.execute(
                    "ALTER TABLE settings_parameters ADD COLUMN extra_json TEXT"
                )
    except Exception as e:
        print(f"ensure_settings_parameters_schema: {e}")


def ensure_it_equipment_catalog_schema():
    try:
        with get_db() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS it_equipment_catalog (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    equipment_type TEXT NOT NULL,
                    equipment_vendor TEXT NOT NULL,
                    equipment_model TEXT NOT NULL,
                    photo_url TEXT,
                    serial_number TEXT,
                    accessories TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    updated_at TEXT DEFAULT (datetime('now'))
                )
                """
            )
    except Exception as e:
        print(f"ensure_it_equipment_catalog_schema: {e}")


ensure_settings_parameters_schema()
ensure_it_equipment_catalog_schema()


def _load_legacy_settings_payload() -> dict:
    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except FileNotFoundError:
        return {}
    except Exception as exc:
        logging.warning("Не удалось прочитать legacy-настройки: %s", exc)
        return {}
    return data if isinstance(data, dict) else {}


def _iter_legacy_parameter_items(raw):
    if isinstance(raw, dict):
        for key, value in raw.items():
            if isinstance(value, dict):
                item = dict(value)
                if "value" not in item or not str(item.get("value", "")).strip():
                    if isinstance(key, str) and key.strip():
                        item.setdefault("value", key)
                yield item
            elif isinstance(value, (list, tuple)):
                sequence = list(value)
                if not sequence:
                    continue
                payload = {"value": sequence[0]}
                if len(sequence) > 1:
                    payload["state"] = sequence[1]
                yield payload
            else:
                if isinstance(key, str) and key.strip():
                    payload = {"value": key}
                    if isinstance(value, str) and value.strip():
                        payload["state"] = value
                    yield payload
                elif isinstance(value, (str, int, float)):
                    yield {"value": value}
    elif isinstance(raw, (list, tuple, set)):
        for value in raw:
            if isinstance(value, dict):
                yield value
            elif isinstance(value, (list, tuple)):
                sequence = list(value)
                if not sequence:
                    continue
                payload = {"value": sequence[0]}
                if len(sequence) > 1:
                    payload["state"] = sequence[1]
                yield payload
            else:
                yield value
    else:
        yield raw


def _normalize_legacy_parameter_item(slug: str, item, *, it_categories: dict[str, str]):
    if isinstance(item, str) or isinstance(item, (int, float)):
        value = str(item).strip()
        if not value:
            return None
        state = _normalize_parameter_state(slug, None)
        entry: dict[str, Any] = {"value": value, "state": state}
        if slug == "iiko_server":
            entry["extra"] = {"server_name": value}
        elif slug == "it_connection":
            default_label = DEFAULT_IT_CONNECTION_CATEGORIES.get("equipment_type", "Тип оборудования")
            entry["extra"] = {
                "category": "equipment_type",
                "category_label": it_categories.get("equipment_type", default_label),
                "equipment_type": value,
                "equipment_vendor": "",
                "equipment_model": "",
                "equipment_status": "",
            }
        return entry

    if isinstance(item, (list, tuple)):
        sequence = list(item)
        if not sequence:
            return None
        payload = {"value": sequence[0]}
        if len(sequence) > 1:
            payload["state"] = sequence[1]
        return _normalize_legacy_parameter_item(slug, payload, it_categories=it_categories)

    if not isinstance(item, dict):
        return None

    raw_value_candidates = [
        item.get("value"),
        item.get("address"),
        item.get("url"),
        item.get("server"),
        item.get("host"),
        item.get("ip"),
        item.get("name"),
        item.get("title"),
        item.get("label"),
        item.get("login"),
        item.get("text"),
    ]
    if slug == "it_connection":
        raw_value_candidates.extend(
            [
                item.get("equipment_type"),
                item.get("equipment_vendor"),
                item.get("equipment_model"),
                item.get("equipment_status"),
            ]
        )
    value = ""
    for candidate in raw_value_candidates:
        if isinstance(candidate, (str, int, float)):
            value = str(candidate).strip()
            if value:
                break
    if not value:
        return None

    raw_state = None
    for key in ("state", "status", "condition"):
        if key in item:
            raw_state = item.get(key)
            break
    state = _normalize_parameter_state(slug, raw_state)
    entry: dict[str, Any] = {"value": value, "state": state}

    if slug == "iiko_server":
        server_name_candidates = [
            item.get("server_name"),
            item.get("name"),
            item.get("title"),
            item.get("label"),
        ]
        server_name = ""
        for candidate in server_name_candidates:
            if isinstance(candidate, (str, int, float)):
                server_name = str(candidate).strip()
                if server_name:
                    break
        if not server_name:
            server_name = value
        entry["extra"] = {"server_name": server_name}
    elif slug == "it_connection":
        extra_payload: dict[str, str] = {}
        category_raw = item.get("category") or item.get("type")
        if isinstance(category_raw, str):
            category_key = category_raw.strip().lower()
            category = LEGACY_IT_CATEGORY_ALIASES.get(category_key, category_key)
        else:
            category = ""
        for field, aliases in (
            ("equipment_type", ("equipment_type", "type", "equipment")),
            ("equipment_vendor", ("equipment_vendor", "vendor", "brand")),
            ("equipment_model", ("equipment_model", "model")),
            ("equipment_status", ("equipment_status", "status_text")),
        ):
            value_candidate = ""
            for alias in aliases:
                candidate = item.get(alias)
                if isinstance(candidate, (str, int, float)):
                    value_candidate = str(candidate).strip()
                    if value_candidate:
                        break
            extra_payload[field] = value_candidate
        if not category:
            for field in ("equipment_type", "equipment_vendor", "equipment_model", "equipment_status"):
                if extra_payload.get(field):
                    category = field
                    break
        if not category:
            category = "equipment_type"
        category_field = DEFAULT_IT_CONNECTION_CATEGORY_FIELDS.get(category)
        if category_field and not extra_payload.get(category_field):
            extra_payload[category_field] = value
        category_label = item.get("category_label")
        if isinstance(category_label, str):
            category_label = category_label.strip()
        else:
            category_label = ""
        if not category_label:
            category_label = it_categories.get(category) or DEFAULT_IT_CONNECTION_CATEGORIES.get(category) or category
        extra_payload.update(
            {
                "category": category,
                "category_label": category_label,
            }
        )
        entry["extra"] = extra_payload
    return entry


def _extract_legacy_parameter_entries(payload: dict) -> dict[str, list[dict[str, Any]]]:
    if not payload:
        return {}
    candidates = [payload]
    for container_key in LEGACY_PARAMETER_CONTAINER_KEYS:
        container = payload.get(container_key)
        if isinstance(container, dict):
            candidates.append(container)
    it_categories = normalize_it_connection_categories(payload.get("it_connection_categories"))
    extracted: dict[str, list[dict[str, Any]]] = {}
    for slug, keys in LEGACY_PARAMETER_KEY_CANDIDATES.items():
        raw_entries = None
        for container in candidates:
            for key in keys:
                if key in container:
                    raw_entries = container[key]
                    break
            if raw_entries is not None:
                break
        if raw_entries is None:
            continue
        normalized_items: list[dict[str, Any]] = []
        for raw_item in _iter_legacy_parameter_items(raw_entries):
            normalized = _normalize_legacy_parameter_item(slug, raw_item, it_categories=it_categories)
            if not normalized:
                continue
            normalized_items.append(normalized)
        if normalized_items:
            extracted[slug] = normalized_items
    return extracted


def maybe_migrate_legacy_settings_parameters() -> None:
    try:
        with get_db() as conn:
            try:
                count_row = conn.execute("SELECT COUNT(*) FROM settings_parameters").fetchone()
            except sqlite3.OperationalError as exc:
                if _is_missing_table_error(exc, "settings_parameters"):
                    ensure_settings_parameters_schema()
                    count_row = conn.execute("SELECT COUNT(*) FROM settings_parameters").fetchone()
                else:
                    raise
            if count_row and count_row[0]:
                return
    except Exception:
        return

    payload = _load_legacy_settings_payload()
    legacy_entries = _extract_legacy_parameter_entries(payload)
    if not legacy_entries:
        return

    try:
        with get_db() as conn:
            for slug, entries in legacy_entries.items():
                seen: set[tuple] = set()
                for entry in entries:
                    value = str(entry.get("value") or "").strip()
                    if not value:
                        continue
                    state = _normalize_parameter_state(slug, entry.get("state"))
                    extra_payload = entry.get("extra") if isinstance(entry.get("extra"), dict) else {}
                    if slug == "it_connection":
                        extra_payload = dict(extra_payload)
                        category = (extra_payload.get("category") or "equipment_type").strip() or "equipment_type"
                        category_field = DEFAULT_IT_CONNECTION_CATEGORY_FIELDS.get(category)
                        if category_field and not (extra_payload.get(category_field) or "").strip():
                            extra_payload[category_field] = value
                        extra_payload.setdefault(
                            "category_label",
                            DEFAULT_IT_CONNECTION_CATEGORIES.get(category, category),
                        )
                        extra_payload.setdefault("equipment_type", "")
                        extra_payload.setdefault("equipment_vendor", "")
                        extra_payload.setdefault("equipment_model", "")
                        extra_payload.setdefault("equipment_status", "")
                        dedup_key = (slug, category, value.lower())
                    else:
                        dedup_key = (slug, value.lower())
                    if dedup_key in seen:
                        continue
                    seen.add(dedup_key)
                    extra_json = (
                        json.dumps(extra_payload, ensure_ascii=False) if extra_payload else None
                    )
                    try:
                        conn.execute(
                            """
                            INSERT OR IGNORE INTO settings_parameters
                                (param_type, value, state, is_deleted, extra_json)
                            VALUES (?, ?, ?, 0, ?)
                            """,
                            (slug, value, state, extra_json),
                        )
                    except sqlite3.Error as exc:
                        logging.warning(
                            "Не удалось мигрировать значение настройки %s=%s: %s",
                            slug,
                            value,
                            exc,
                        )
            conn.commit()
    except Exception as exc:
        logging.warning("Ошибка миграции legacy-настроек: %s", exc)


maybe_migrate_legacy_settings_parameters()

def _collect_departments_from_locations() -> tuple[list[str], set[str]]:
    """
    Возвращает:
    - отсортированный список названий департаментов (четвёртая ветка дерева
      locations.json: бизнес → тип → город → точка);
    - множество названий третьего уровня (городов/веток) для последующей
      очистки устаревших значений.
    """

    try:
       locations_payload = load_locations()
    except Exception:
        locations_payload = {"tree": {}, "statuses": {}}

    locations = locations_payload.get("tree") if isinstance(locations_payload, dict) else {}
    departments: set[str] = set()
    third_level_names: set[str] = set()

    for brand in (locations or {}).values():
        if not isinstance(brand, dict):
            continue

        for partner_type in brand.values():
            if not isinstance(partner_type, dict):
                continue

            for city_name, city_departments in partner_type.items():
                if isinstance(city_name, str) and city_name.strip():
                    third_level_names.add(city_name.strip())

                if isinstance(city_departments, dict):
                    iterable = city_departments.values()
                elif isinstance(city_departments, list):
                    iterable = city_departments
                else:
                    continue

                for department_name in iterable:
                    if isinstance(department_name, str) and department_name.strip():
                        departments.add(department_name.strip())

    return (
        sorted(departments, key=lambda name: name.lower()),
        third_level_names,
    )


def ensure_departments_seeded():
    """Добавляет значения департаментов на основе locations.json, если их нет."""

    departments, obsolete_candidates = _collect_departments_from_locations()
    if not departments:
        return

    department_set = set(departments)

    try:
        with get_db() as conn:
            existing_rows = conn.execute(
                "SELECT value FROM settings_parameters WHERE param_type = ? AND is_deleted = 0",
                ("department",),
            ).fetchall()
            existing = {row["value"] for row in existing_rows if row["value"]}

            obsolete = [
                value
                for value in existing
                if value in obsolete_candidates and value not in department_set
            ]
            if obsolete:
                conn.executemany(
                    "UPDATE settings_parameters SET is_deleted = 1, deleted_at = datetime('now') WHERE param_type = ? AND value = ?",
                    [("department", value) for value in obsolete],
                )

            missing = [dep for dep in departments if dep not in existing]
            if not missing:
                return

            conn.executemany(
                """
                INSERT OR IGNORE INTO settings_parameters (param_type, value, state, is_deleted)
                VALUES (?, ?, 'Активен', 0)
                """,
                [("department", dep) for dep in missing],
            )
            conn.commit()
    except Exception as e:
        logging.warning("Не удалось заполнить департаменты: %s", e)


ensure_departments_seeded()

def create_unblock_request_task(user_id: str, reason: str = ""):
    """
    Создаёт задачу оператору о запросе разблокировки клиента.
    """
    try:
        with get_db() as conn:
            cur = conn.cursor()
            # аккуратно вычислим следующий seq
            row = cur.execute("SELECT COALESCE(MAX(seq),0)+1 AS n FROM tasks").fetchone()
            seq = (row["n"] if row and "n" in row.keys() else 1)
            title = f"Запрос разблокировки клиента {user_id}"
            body_html = f"<p>Клиент {user_id} запросил разблокировку.</p>" + (f"<p>Комментарий: {reason}</p>" if reason else "")
            cur.execute("""
                INSERT INTO tasks (seq, source, title, body_html, creator, tag, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (seq, "DL", title, body_html, "system", "unblock_request", "Новая"))
            conn.commit()
    except Exception as e:
        print(f"create_unblock_request_task: {e}")

def _next_task_seq(conn):
    row = conn.execute("SELECT val FROM task_seq WHERE id=1").fetchone()
    cur = int(row["val"])
    conn.execute("UPDATE task_seq SET val = ? WHERE id=1", (cur+1,))
    return cur + 1

def _display_no(source, seq):
    return f"{source}_{seq}" if source else str(seq)

def _touch_activity(conn, task_id):
    conn.execute("UPDATE tasks SET last_activity_at = datetime('now') WHERE id=?", (task_id,))

def _generate_unique_channel_public_id(cur: sqlite3.Cursor, used: set[str] | None = None) -> str:
    """Возвращает уникальный идентификатор для канала (hex UUID)."""
    used_ids = set(used or set())
    while True:
        candidate = uuid4().hex.lower()
        if candidate in used_ids:
            continue
        try:
            row = cur.execute(
                "SELECT 1 FROM channels WHERE public_id = ?",
                (candidate,),
            ).fetchone()
        except sqlite3.OperationalError:
            row = None
        if row:
            continue
        return candidate


def ensure_channels_schema():
    """Добавляет недостающие колонки в channels, если база старая."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(channels)").fetchall()}
        # ожидаемые поля: id, token, bot_name, bot_username, channel_name,
        # questions_cfg, max_questions, is_active, question_template_id, rating_template_id
        if 'bot_name' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN bot_name TEXT")
        if 'bot_username' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN bot_username TEXT")
        if 'channel_name' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN channel_name TEXT")
        if 'questions_cfg' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN questions_cfg TEXT DEFAULT '{}'")
        if 'max_questions' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN max_questions INTEGER DEFAULT 0")
        if 'is_active' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN is_active INTEGER DEFAULT 1")
        if 'question_template_id' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN question_template_id TEXT")
        if 'rating_template_id' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN rating_template_id TEXT")
        if 'public_id' not in cols:
            cur.execute("ALTER TABLE channels ADD COLUMN public_id TEXT")
        existing_ids = set()
        try:
            rows = cur.execute("SELECT id, public_id FROM channels").fetchall()
        except sqlite3.OperationalError:
            rows = []
        for row in rows:
            cid = row[0]
            raw_value = row[1] or ''
            current = raw_value.strip().lower()
            if current:
                if raw_value != current:
                    cur.execute(
                        "UPDATE channels SET public_id = ? WHERE id = ?",
                        (current, cid),
                    )
                if current not in existing_ids:
                    existing_ids.add(current)
                continue
            new_id = _generate_unique_channel_public_id(cur, existing_ids)
            cur.execute("UPDATE channels SET public_id = ? WHERE id = ?", (new_id, cid))
            existing_ids.add(new_id)
        try:
            cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_channels_public_id ON channels(public_id)")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute(
                "UPDATE channels SET bot_username = COALESCE(bot_username, bot_name) WHERE bot_username IS NULL OR bot_username = ''"
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        print(f"ensure_channels_schema: {e}")
    finally:
        try: conn.close()
        except: pass
ensure_channels_schema()

def _load_sanitized_bot_settings_payload():
    settings_payload = load_settings()
    locations_payload = load_locations()
    location_tree = (
        locations_payload.get("tree", {})
        if isinstance(locations_payload, dict)
        else {}
    )
    definitions = build_location_presets(
        location_tree, base_definitions=DEFAULT_BOT_PRESET_DEFINITIONS
    )
    source = settings_payload.get("bot_settings") if isinstance(settings_payload, dict) else None
    return sanitize_bot_settings(source, definitions=definitions)

def _fetch_bot_identity(token: str) -> tuple[str, str]:
    if not token:
        raise ValueError("Пустой токен")
    response = requests.get(f"https://api.telegram.org/bot{token}/getMe", timeout=10)
    data = response.json()
    if not data.get("ok"):
        raise ValueError(data.get("description") or "Не удалось получить данные бота")
    result = data.get("result") or {}
    username = (result.get("username") or "").strip()
    first_name = (result.get("first_name") or "").strip()
    last_name = (result.get("last_name") or "").strip()
    full_name = " ".join(part for part in [first_name, last_name] if part).strip()
    display_name = full_name or username
    return display_name, username


def _refresh_bot_identity_if_needed(conn: sqlite3.Connection, channel_row: dict) -> dict:
    if not channel_row:
        return channel_row
    token = (channel_row.get("token") or "").strip()
    if not token:
        return channel_row
    existing_name = (channel_row.get("bot_name") or "").strip()
    existing_username = (channel_row.get("bot_username") or "").strip()
    needs_refresh = (not existing_name) or (not existing_username) or (existing_name == existing_username)
    if not needs_refresh:
        return channel_row
    try:
        display_name, username = _fetch_bot_identity(token)
    except Exception as exc:
        logging.warning("Не удалось обновить имя бота #%s: %s", channel_row.get("id"), exc)
        return channel_row

    updated_name = display_name or existing_name
    updated_username = username or existing_username
    if updated_name == existing_name and updated_username == existing_username:
        return channel_row

    channel_row["bot_name"] = updated_name
    channel_row["bot_username"] = updated_username
    try:
        conn.execute(
            "UPDATE channels SET bot_name = ?, bot_username = ? WHERE id = ?",
            (updated_name, updated_username, channel_row.get("id")),
        )
        conn.commit()
    except Exception as exc:
        logging.warning("Не удалось сохранить имя бота #%s: %s", channel_row.get("id"), exc)
    return channel_row

def ensure_feedback_requests_table():
    try:
        conn = get_db()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pending_feedback_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                channel_id INTEGER NOT NULL,
                ticket_id TEXT,
                source TEXT,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                UNIQUE(user_id, channel_id, ticket_id, source)
            )
        """)
        conn.commit()
    except Exception as e:
        print(f"ensure_feedback_requests_table: {e}")
    finally:
        try: conn.close()
        except: pass

ensure_feedback_requests_table()

def ensure_client_profile_schema():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS client_usernames (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                username TEXT NOT NULL,
                seen_at TEXT NOT NULL,
                UNIQUE(user_id, username)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS client_phones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                phone TEXT NOT NULL,
                label TEXT,
                source TEXT NOT NULL CHECK (source IN ('telegram','manual')),
                is_active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                created_by TEXT
            )
        """)
        conn.commit()
    except Exception as e:
        print(f"ensure_client_profile_schema: {e}")
    finally:
        try: conn.close()
        except: pass

ensure_client_profile_schema()

def ensure_history_mark_columns():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(chat_history)")
        cols = {r['name'] for r in cur.fetchall()}
        if 'edited_at' not in cols:
            cur.execute("ALTER TABLE chat_history ADD COLUMN edited_at TEXT")
        if 'deleted_at' not in cols:
            cur.execute("ALTER TABLE chat_history ADD COLUMN deleted_at TEXT")
        conn.commit()
    except Exception as e:
        print(f"ensure_history_mark_columns: {e}")
    finally:
        try: conn.close()
        except: pass

# === Миграция: столбцы для reply/thread ===
def ensure_history_reply_columns():
    try:
        conn = get_db()
        cur = conn.cursor()
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(chat_history)").fetchall()}
        if "tg_message_id" not in cols:
            cur.execute("ALTER TABLE chat_history ADD COLUMN tg_message_id INTEGER")
        if "reply_to_tg_id" not in cols:
            cur.execute("ALTER TABLE chat_history ADD COLUMN reply_to_tg_id INTEGER")
        conn.commit()
    except Exception as e:
        print(f"ensure_history_reply_columns: {e}")
    finally:
        try: conn.close()
        except: pass
        commit()

    ensure_history_mark_columns()
    ensure_history_reply_columns()

def ensure_ticket_time_schema():
    """ticket_spans + счётчики reopen/closed в tickets."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS ticket_spans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id TEXT NOT NULL,
                span_no INTEGER NOT NULL,
                started_at TEXT NOT NULL,
                ended_at   TEXT,
                duration_seconds INTEGER,
                UNIQUE(ticket_id, span_no)
            )
        """)
        # Добавляем счётчики в tickets
        cols = {r['name'] for r in cur.execute("PRAGMA table_info(tickets)").fetchall()}
        if 'reopen_count' not in cols:
            cur.execute("ALTER TABLE tickets ADD COLUMN reopen_count INTEGER DEFAULT 0")
        if 'closed_count' not in cols:
            cur.execute("ALTER TABLE tickets ADD COLUMN closed_count INTEGER DEFAULT 0")
        conn.commit()
    except Exception as e:
        print(f"ensure_ticket_time_schema: {e}")
    finally:
        try: conn.close()
        except: pass

ensure_ticket_time_schema()

def _active_span(conn, ticket_id: str):
    return conn.execute(
        "SELECT id, span_no, started_at FROM ticket_spans WHERE ticket_id=? AND ended_at IS NULL",
        (ticket_id,)
    ).fetchone()

def _next_span_no(conn, ticket_id: str) -> int:
    row = conn.execute("SELECT MAX(span_no) AS mx FROM ticket_spans WHERE ticket_id=?", (ticket_id,)).fetchone()
    return (row['mx'] or 0) + 1

def start_span_if_missing(ticket_id: str):
    conn = get_db()
    try:
        cur = conn.cursor()
        if not _active_span(conn, ticket_id):
            span_no = _next_span_no(conn, ticket_id)
            cur.execute(
                "INSERT INTO ticket_spans(ticket_id, span_no, started_at) VALUES(?,?,?)",
                (ticket_id, span_no, dt.now().isoformat())
            )
            conn.commit()
    finally:
        try: conn.close()
        except: pass

def close_active_span(ticket_id: str):
    conn = get_db()
    try:
        cur = conn.cursor()
        sp = _active_span(conn, ticket_id)
        if sp:
            ended = dt.now()
            dur = int((ended - datetime.datetime.fromisoformat(sp['started_at'])).total_seconds())
            cur.execute(
                "UPDATE ticket_spans SET ended_at=?, duration_seconds=? WHERE id=?",
                (ended.isoformat(), dur, sp['id'])
            )
            conn.commit()
    finally:
        try: conn.close()
        except: pass

def open_new_span(ticket_id: str):
    conn = get_db()
    try:
        cur = conn.cursor()
        span_no = _next_span_no(conn, ticket_id)
        cur.execute(
            "INSERT INTO ticket_spans(ticket_id, span_no, started_at) VALUES(?,?,?)",
            (ticket_id, span_no, dt.now().isoformat())
        )
        conn.commit()
    finally:
        try: conn.close()
        except: pass

def get_latest_client_name(user_id):
    """Получить последнее имя клиента из таблицы messages"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT client_name FROM messages
            WHERE user_id = ? AND client_name IS NOT NULL AND client_name != ''
            ORDER BY created_at DESC
            LIMIT 1
        """, (user_id,))
        row = cur.fetchone()
        conn.close()
        return row['client_name'] if row else None
    except Exception as e:
        print(f"Ошибка при получении имени клиента: {e}")
        return None

def _generate_template_id(prefix: str) -> str:
    return f"{prefix}_{uuid4().hex[:8]}"


def _sanitize_str(value) -> str:
    return str(value).strip() if isinstance(value, str) else str(value).strip() if value is not None else ""


def _sanitize_category_templates(raw, fallback=None):
    templates = []
    seen_ids: set[str] = set()
    source = raw if isinstance(raw, list) else []
    for item in source:
        if not isinstance(item, dict):
            continue
        template_id = _sanitize_str(item.get("id"))
        if not template_id:
            template_id = _generate_template_id("cat")
        if template_id in seen_ids:
            template_id = _generate_template_id("cat")
        seen_ids.add(template_id)

        name = _sanitize_str(item.get("name"))
        categories = []
        for raw_cat in item.get("categories", []):
            cat = _sanitize_str(raw_cat)
            if cat:
                categories.append(cat)

        if not name:
            name = "Общий шаблон" if not templates else "Шаблон без названия"

        templates.append({
            "id": template_id,
            "name": name,
            "categories": categories,
        })

    if not templates:
        fallback_categories = []
        if isinstance(fallback, list):
            fallback_categories = [_sanitize_str(cat) for cat in fallback if _sanitize_str(cat)]
        template_id = _generate_template_id("cat")
        templates.append({
            "id": template_id,
            "name": "Общий шаблон",
            "categories": fallback_categories,
        })

    return templates


def _sanitize_question_templates(raw):
    templates = []
    seen_ids: set[str] = set()
    source = raw if isinstance(raw, list) else []
    for item in source:
        if not isinstance(item, dict):
            continue
        template_id = _sanitize_str(item.get("id"))
        if not template_id:
            template_id = _generate_template_id("q")
        if template_id in seen_ids:
            template_id = _generate_template_id("q")
        seen_ids.add(template_id)

        name = _sanitize_str(item.get("name")) or "Новый шаблон вопросов"
        questions = []
        for raw_question in item.get("questions", []):
            question = _sanitize_str(raw_question)
            if question:
                questions.append(question)

        if not questions and not name:
            continue

        templates.append({
            "id": template_id,
            "name": name,
            "questions": questions,
        })

    return templates


def _sanitize_completion_templates(raw):
    templates = []
    seen_ids: set[str] = set()
    source = raw if isinstance(raw, list) else []
    for item in source:
        if not isinstance(item, dict):
            continue
        template_id = _sanitize_str(item.get("id"))
        if not template_id:
            template_id = _generate_template_id("act")
        if template_id in seen_ids:
            template_id = _generate_template_id("act")
        seen_ids.add(template_id)

        name = _sanitize_str(item.get("name")) or "Новый шаблон действий"
        entries = []
        for raw_entry in item.get("items", []):
            if not isinstance(raw_entry, dict):
                continue
            question = _sanitize_str(raw_entry.get("question"))
            action = _sanitize_str(raw_entry.get("action"))
            if not (question or action):
                continue
            entries.append({"question": question, "action": action})

        if not entries and not name:
            continue

        templates.append({
            "id": template_id,
            "name": name,
            "items": entries,
        })

    return templates


def _sanitize_hex_color(value: Any, fallback: str) -> str:
    if isinstance(value, str):
        candidate = value.strip()
        if re.fullmatch(r"#([0-9a-fA-F]{3}|[0-9a-fA-F]{6})", candidate):
            return candidate
    return fallback


def _sanitize_time_metrics(raw: Any) -> dict[str, Any]:
    defaults = {
        "good_limit": 30,
        "warning_limit": 60,
        "colors": {
            "good": "#d1f7d1",
            "warning": "#fff4d6",
            "danger": "#f8d7da",
        },
    }
    payload = raw if isinstance(raw, dict) else {}

    try:
        good_limit = int(payload.get("good_limit", defaults["good_limit"]))
    except (TypeError, ValueError):
        good_limit = defaults["good_limit"]
    if good_limit <= 0:
        good_limit = defaults["good_limit"]

    try:
        warning_limit = int(payload.get("warning_limit", defaults["warning_limit"]))
    except (TypeError, ValueError):
        warning_limit = defaults["warning_limit"]
    if warning_limit <= good_limit:
        warning_limit = max(good_limit + 1, defaults["warning_limit"])

    raw_colors = payload.get("colors") if isinstance(payload.get("colors"), dict) else {}
    colors = {
        "good": _sanitize_hex_color(raw_colors.get("good"), defaults["colors"]["good"]),
        "warning": _sanitize_hex_color(raw_colors.get("warning"), defaults["colors"]["warning"]),
        "danger": _sanitize_hex_color(raw_colors.get("danger"), defaults["colors"]["danger"]),
    }

    return {
        "good_limit": good_limit,
        "warning_limit": warning_limit,
        "colors": colors,
    }


def _sanitize_auto_close_templates(raw, *, fallback_hours: int = 24):
    templates: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    source = raw if isinstance(raw, list) else []
    try:
        base_hours = int(fallback_hours)
    except (TypeError, ValueError):
        base_hours = 24
    if base_hours <= 0:
        base_hours = 24
    for item in source:
        if not isinstance(item, dict):
            continue
        template_id = _sanitize_str(item.get("id"))
        if not template_id:
            template_id = _generate_template_id("auto")
        if template_id in seen_ids:
            template_id = _generate_template_id("auto")
        seen_ids.add(template_id)

        name = _sanitize_str(item.get("name")) or "Шаблон автозакрытия"
        description = _sanitize_str(item.get("description"))

        hours_raw = None
        for key in ("hours", "timeout_hours", "auto_close_hours"):
            value = item.get(key)
            if value is None:
                continue
            try:
                hours_raw = int(float(value))
            except (TypeError, ValueError):
                continue
            if hours_raw > 0:
                break
        if not isinstance(hours_raw, int) or hours_raw <= 0:
            hours_raw = base_hours
        hours = max(1, min(hours_raw, 720))

        templates.append(
            {
                "id": template_id,
                "name": name or "Шаблон автозакрытия",
                "description": description,
                "hours": hours,
            }
        )

    if not templates:
        template_id = _generate_template_id("auto")
        templates.append(
            {
                "id": template_id,
                "name": "Стандартный сценарий",
                "description": "",
                "hours": max(1, min(base_hours, 720)),
            }
        )

    return templates


def sanitize_auto_close_config(raw, *, fallback_hours: int = 24):
    config = raw if isinstance(raw, dict) else {}
    templates = _sanitize_auto_close_templates(config.get("templates"), fallback_hours=fallback_hours)
    active_id = _sanitize_str(config.get("active_template_id"))
    template_ids = {tpl["id"] for tpl in templates}
    if active_id not in template_ids:
        active_id = next(iter(template_ids), None)
    return {
        "templates": templates,
        "active_template_id": active_id,
    }

def ensure_dialog_config(settings):
    if not isinstance(settings, dict):
        settings = {}
    config = settings.get("dialog_config")
    if not isinstance(config, dict):
        config = {}

    fallback_categories = settings.get("categories") if isinstance(settings.get("categories"), list) else []
    category_templates = _sanitize_category_templates(config.get("category_templates"), fallback=fallback_categories)
    question_templates = _sanitize_question_templates(config.get("question_templates"))
    completion_templates = _sanitize_completion_templates(config.get("completion_templates"))
    time_metrics = _sanitize_time_metrics(config.get("time_metrics"))

    config["category_templates"] = category_templates
    config["question_templates"] = question_templates
    config["completion_templates"] = completion_templates
    config["time_metrics"] = time_metrics

    settings["dialog_config"] = config
    settings["categories"] = category_templates[0]["categories"] if category_templates else []
    return settings

def ensure_auto_close_config(settings):
    if not isinstance(settings, dict):
        settings = {}
    try:
        fallback_hours = int(settings.get("auto_close_hours", 24))
    except (TypeError, ValueError):
        fallback_hours = 24
    if fallback_hours <= 0:
        fallback_hours = 24
    raw_config = settings.get("auto_close_config")
    auto_close_config = sanitize_auto_close_config(raw_config, fallback_hours=fallback_hours)
    templates = auto_close_config.get("templates", [])
    active_id = auto_close_config.get("active_template_id")
    active_template = next((tpl for tpl in templates if tpl.get("id") == active_id), templates[0] if templates else None)
    if active_template:
        settings["auto_close_hours"] = active_template.get("hours", fallback_hours)
    else:
        settings["auto_close_hours"] = fallback_hours
    settings["auto_close_config"] = auto_close_config
    return settings


# Функция для загрузки настроек
def load_settings():
    settings = {"auto_close_hours": 24, "categories": ["Консультация"], "client_statuses": ["Новый", "Постоянный", "VIP"]}
    if os.path.exists(SETTINGS_PATH):
        try:
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                settings.update(json.load(f))
        except:
            pass
    if not isinstance(settings.get("network_profiles"), list):
        settings["network_profiles"] = []
    settings["it_connection_categories"] = normalize_it_connection_categories(
        settings.get("it_connection_categories")
    )
    settings["bot_settings"] = sanitize_bot_settings(
        settings.get("bot_settings"), definitions=DEFAULT_BOT_PRESET_DEFINITIONS
    )
    settings = ensure_auto_close_config(settings)
    return ensure_dialog_config(settings)

# Функция для сжатия изображений на лету
import io
import mimetypes
def resize_image(image_path, max_size=600):
    """Сжимает изображение до максимального размера"""
    try:
        from PIL import Image
            
        with Image.open(image_path) as img:
            # Конвертируем в RGB если нужно
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            
            # Проверяем, нужно ли изменять размер
            if img.width <= max_size and img.height <= max_size:
                # Если изображение уже меньше максимального размера, возвращаем оригинал
                buffer = io.BytesIO()
                img.save(buffer, format='JPEG', quality=85, optimize=True)
                buffer.seek(0)
                return buffer
            
            # Изменяем размер сохраняя пропорции
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            
            # Сохраняем в буфер
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85, optimize=True)
            buffer.seek(0)
            
            return buffer
    except ImportError:
        # Pillow не установлен — просто вернём оригинал, без ресайза
        return None
    except Exception as e:
        print(f"Ошибка при обработке изображения: {e}")
        return None

# Функция для сохранения настроек
def save_settings(settings):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

# Функция для загрузки локаций
def load_locations():
    tree = {}
    statuses = {}
    if not os.path.exists(LOCATIONS_PATH):
        return {"tree": {}, "statuses": {}}

    try:
        with open(LOCATIONS_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return {"tree": {}, "statuses": {}}

    if isinstance(raw, dict) and "tree" in raw:
        tree = raw.get("tree") or {}
        statuses = raw.get("statuses") or {}
    elif isinstance(raw, dict):
        tree = raw

    return {
        "tree": tree if isinstance(tree, dict) else {},
        "statuses": statuses if isinstance(statuses, dict) else {},
    }


# Функция для сохранения локаций
def save_locations(locations):
    payload = {"tree": {}, "statuses": {}}
    if isinstance(locations, dict):
        if "tree" in locations or "statuses" in locations:
            if isinstance(locations.get("tree"), dict):
                payload["tree"] = locations["tree"]
            if isinstance(locations.get("statuses"), dict):
                payload["statuses"] = locations["statuses"]
        else:
            payload["tree"] = locations

    with open(LOCATIONS_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def sanitize_org_structure_payload(source: dict | list | None) -> dict:
    nodes_source = []
    if isinstance(source, dict):
        raw_nodes = source.get("nodes")
        if isinstance(raw_nodes, list):
            nodes_source = raw_nodes
    elif isinstance(source, list):
        nodes_source = source

    sanitized: list[dict] = []
    if not isinstance(nodes_source, list):
        return {"nodes": sanitized}

    used_ids: set[str] = set()
    raw_to_sanitized: dict[str, str] = {}

    def _sanitize_members(payload) -> list[int]:
        if not isinstance(payload, (list, tuple, set)):
            return []
        result: list[int] = []
        seen: set[int] = set()
        for value in payload:
            try:
                member_id = int(value)
            except (TypeError, ValueError):
                continue
            if member_id not in seen:
                seen.add(member_id)
                result.append(member_id)
        result.sort()
        return result

    temp_entries: list[tuple[dict, dict]] = []
    for entry in nodes_source:
        if not isinstance(entry, dict):
            continue
        raw_id = str(entry.get("id") or "").strip()
        node_id = raw_id or uuid4().hex
        while not node_id or node_id in used_ids:
            node_id = uuid4().hex
        if raw_id:
            raw_to_sanitized.setdefault(raw_id, node_id)
        name = str(entry.get("name") or "").strip() or "Новая ветка"
        members = _sanitize_members(entry.get("members"))
        sanitized_entry = {
            "id": node_id,
            "name": name,
            "members": members,
            "parent_id": None,
        }
        temp_entries.append((entry, sanitized_entry))
        used_ids.add(node_id)

    valid_ids = {item[1]["id"] for item in temp_entries}
    for original, sanitized_entry in temp_entries:
        parent_raw = None
        if isinstance(original, dict):
            parent_raw = original.get("parent_id")
        parent_id: str | None = None
        if parent_raw is not None:
            parent_key = str(parent_raw).strip()
            if parent_key:
                if parent_key in raw_to_sanitized and raw_to_sanitized[parent_key] in valid_ids:
                    parent_id = raw_to_sanitized[parent_key]
                elif parent_key in valid_ids:
                    parent_id = parent_key
        if parent_id == sanitized_entry["id"] or parent_id not in valid_ids:
            parent_id = None
        sanitized_entry["parent_id"] = parent_id
        sanitized.append(sanitized_entry)

    sanitized.sort(key=lambda item: ((item.get("parent_id") or ""), item.get("name", "").lower(), item.get("id")))
    return {"nodes": sanitized}


def load_org_structure() -> dict:
    if not os.path.exists(ORG_STRUCTURE_PATH):
        return {"nodes": []}
    try:
        with open(ORG_STRUCTURE_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return {"nodes": []}
    return sanitize_org_structure_payload(raw)


def save_org_structure(structure: dict | list | None) -> dict:
    sanitized = sanitize_org_structure_payload(structure)
    with open(ORG_STRUCTURE_PATH, "w", encoding="utf-8") as f:
        json.dump(sanitized, f, ensure_ascii=False, indent=2)
    return sanitized


def _normalize_department_path(value: str | None) -> str:
    if not value:
        return ""
    if not isinstance(value, str):
        value = str(value)
    parts = [part.strip() for part in value.split("/") if part and part.strip()]
    return " / ".join(parts)


def _find_org_node_id_by_department(nodes: list[dict], department: str | None) -> str | None:
    normalized_value = _normalize_department_path(department)
    if not normalized_value:
        return None
    normalized_lower = normalized_value.lower()
    id_map: dict[str, dict] = {}
    for node in nodes:
        node_id = node.get("id")
        if node_id:
            id_map[str(node_id)] = node

    def build_path(node_id: str) -> str:
        path_parts: list[str] = []
        current_id = node_id
        visited: set[str] = set()
        while current_id and current_id not in visited:
            visited.add(current_id)
            node = id_map.get(current_id)
            if not node:
                break
            name = str(node.get("name") or "").strip()
            if name:
                path_parts.append(name)
            parent_id = node.get("parent_id")
            current_id = str(parent_id) if parent_id and str(parent_id) in id_map else None
        return " / ".join(reversed(path_parts))

    fallback_id: str | None = None
    for node_id, node in id_map.items():
        path = _normalize_department_path(build_path(node_id))
        if path and path.lower() == normalized_lower:
            return node_id
        if fallback_id is None:
            name_normalized = _normalize_department_path(node.get("name"))
            if name_normalized and name_normalized.lower() == normalized_lower:
                fallback_id = node_id
    return fallback_id


def sync_org_structure_user_department(
    user_id: int | str | None,
    previous_department: str | None,
    new_department: str | None,
) -> None:
    try:
        numeric_user_id = int(user_id)
    except (TypeError, ValueError):
        return
    if numeric_user_id <= 0:
        return

    previous_normalized = _normalize_department_path(previous_department)
    new_normalized = _normalize_department_path(new_department)
    if not previous_normalized and not new_normalized:
        return

    structure = load_org_structure()
    nodes = structure.get("nodes") if isinstance(structure, dict) else None
    if not isinstance(nodes, list) or not nodes:
        return

    target_add = _find_org_node_id_by_department(nodes, new_normalized) if new_normalized else None
    target_remove = None
    if previous_normalized and (not new_normalized or previous_normalized != new_normalized):
        target_remove = _find_org_node_id_by_department(nodes, previous_normalized)

    changed = False
    for node in nodes:
        node_id = node.get("id")
        if not node_id:
            continue
        raw_members = node.get("members") if isinstance(node.get("members"), list) else []
        member_set: set[int] = set()
        for value in raw_members:
            try:
                member_set.add(int(value))
            except (TypeError, ValueError):
                continue

        updated_members = set(member_set)
        if target_remove and str(node_id) == str(target_remove) and numeric_user_id in updated_members:
            if not target_add or str(target_add) != str(node_id):
                updated_members.discard(numeric_user_id)
        if target_add and str(node_id) == str(target_add):
            if numeric_user_id not in updated_members:
                updated_members.add(numeric_user_id)

        if updated_members != member_set:
            node["members"] = sorted(updated_members)
            changed = True

    if changed:
        save_org_structure(structure)


# === Публичная веб-форма обращений ===

def ensure_web_form_schema():
    """Создаёт таблицу с сессиями веб-форм, если её ещё нет."""

    create_sql = f"""
        CREATE TABLE IF NOT EXISTS {WEB_FORM_SESSIONS_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token TEXT NOT NULL UNIQUE,
            ticket_id TEXT NOT NULL,
            channel_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            answers_json TEXT,
            client_name TEXT,
            client_contact TEXT,
            username TEXT,
            created_at TEXT NOT NULL,
            last_active_at TEXT NOT NULL
        )
    """

    with get_db() as conn:
        exec_with_retry(lambda: conn.execute(create_sql))


ensure_web_form_schema()


def _allocate_web_user_id(conn) -> int:
    """Возвращает уникальный отрицательный user_id для обращений с веб-формы."""

    row = exec_with_retry(
        lambda: conn.execute(
            "SELECT MIN(user_id) AS min_id FROM messages WHERE user_id < 0"
        )
    ).fetchone()
    min_id = row["min_id"] if row and row["min_id"] is not None else None
    if min_id is None or int(min_id) >= 0:
        return -1
    return int(min_id) - 1


def _prepare_public_form_config(channel_id: int) -> dict | None:
    """Собирает метаданные для публичной формы (канал, вопросы, пресеты)."""

    with get_db() as conn:
        channel_row = exec_with_retry(
            lambda: conn.execute(
                "SELECT id, channel_name, is_active, public_id FROM channels WHERE id = ?",
                (channel_id,),
            )
        ).fetchone()

    if not channel_row:
        return None

    is_active_raw = channel_row["is_active"]
    if is_active_raw is not None and str(is_active_raw).strip().lower() not in {"1", "true"}:
        return None

    settings = load_settings()
    locations_payload = load_locations()
    location_tree = locations_payload.get("tree") if isinstance(locations_payload, dict) else {}
    definitions = build_location_presets(
        location_tree if isinstance(location_tree, dict) else {},
        base_definitions=DEFAULT_BOT_PRESET_DEFINITIONS,
    )

    bot_settings = settings.get("bot_settings") if isinstance(settings, dict) else {}
    question_flow = []
    if isinstance(bot_settings, dict):
        active_template_id = bot_settings.get("active_template_id")
        templates = bot_settings.get("question_templates") or []
        if active_template_id and isinstance(templates, list):
            for tpl in templates:
                if isinstance(tpl, dict) and tpl.get("id") == active_template_id:
                    question_flow = tpl.get("question_flow") or []
                    break
        if not question_flow:
            question_flow = bot_settings.get("question_flow") or []

    prepared_questions: list[dict] = []
    for index, item in enumerate(question_flow, start=1):
        if not isinstance(item, dict):
            continue
        q_type = str(item.get("type") or "custom").strip().lower()
        if q_type not in {"preset", "custom"}:
            q_type = "custom"
        raw_text = item.get("text") or item.get("label") or ""
        text = str(raw_text or "").strip()
        q_id = str(item.get("id") or "").strip() or uuid4().hex
        try:
            order = int(item.get("order") or index)
        except (TypeError, ValueError):
            order = index

        entry = {
            "id": q_id,
            "text": text,
            "type": "preset" if q_type == "preset" else "custom",
            "order": order,
        }

        excluded = item.get("excluded_options") or item.get("excludedOptions")
        if isinstance(excluded, (list, tuple)):
            entry["excluded_options"] = [
                str(value).strip() for value in excluded if str(value or "").strip()
            ]

        if entry["type"] == "preset":
            preset_payload = item.get("preset") if isinstance(item.get("preset"), dict) else {}
            group = str(preset_payload.get("group") or item.get("group") or "").strip()
            field = str(preset_payload.get("field") or item.get("field") or "").strip()
            entry["preset"] = {"group": group, "field": field}
            meta = definitions.get(group, {}).get("fields", {}).get(field, {})
            options = meta.get("options") if isinstance(meta, dict) else []
            if isinstance(options, list):
                entry["options"] = options
            deps = meta.get("option_dependencies") if isinstance(meta, dict) else None
            if isinstance(deps, dict) and deps:
                entry["option_dependencies"] = deps
            tree = meta.get("tree") if isinstance(meta, dict) else None
            if tree:
                entry["tree"] = tree
            label = str(meta.get("label") or "").strip()
            if not entry["text"] and label:
                entry["text"] = label

        if not entry["text"]:
            entry["text"] = f"Вопрос #{order}"

        prepared_questions.append(entry)

    prepared_questions.sort(key=lambda q: q.get("order") or 0)
    lookup = {item["id"]: item for item in prepared_questions}

    channel_name = channel_row["channel_name"] or f"Канал #{channel_row['id']}"

    public_id = (channel_row["public_id"] or "").strip().lower()
    if not public_id:
        with get_db() as ensure_conn:
            ensure_cur = ensure_conn.cursor()
            public_id = _generate_unique_channel_public_id(ensure_cur)
            ensure_cur.execute(
                "UPDATE channels SET public_id = ? WHERE id = ?",
                (public_id, channel_row["id"]),
            )
            ensure_conn.commit()

    return {
        "channel": {
            "id": channel_row["id"],
            "public_id": public_id,
            "name": channel_name,
        },
        "questions": prepared_questions,
        "question_lookup": lookup,
        "answer_order": [item["id"] for item in prepared_questions],
    }


def _resolve_channel_id(channel_ref: str | int | None) -> int | None:
    """Возвращает внутренний numeric id канала по внешнему идентификатору."""
    if channel_ref is None:
        return None
    if isinstance(channel_ref, int):
        return channel_ref
    ref = str(channel_ref).strip()
    if not ref:
        return None
    if ref.isdigit():
        try:
            return int(ref)
        except ValueError:
            return None
    lookup = ref.lower()
    with get_db() as conn:
        row = exec_with_retry(
            lambda: conn.execute(
                "SELECT id FROM channels WHERE LOWER(public_id) = ?",
                (lookup,),
            )
        ).fetchone()
    if row:
        return int(row["id"])
    return None


def _serialize_public_form_config(config: dict) -> dict:
    questions_payload = []
    for question in config.get("questions", []):
        payload = {
            "id": question["id"],
            "text": question.get("text", ""),
            "type": question.get("type", "custom"),
            "order": question.get("order", 0),
        }
        for key in ("preset", "options", "option_dependencies", "tree", "excluded_options"):
            if key in question and question[key]:
                payload[key] = question[key]
        questions_payload.append(payload)

    return {
        "success": True,
        "channel": config.get("channel", {}),
        "questions": questions_payload,
    }


def _build_answers_summary(config: dict, answers: dict[str, str]) -> list[str]:
    summary: list[str] = []
    order = config.get("answer_order") or []
    lookup = config.get("question_lookup") or {}
    for question_id in order:
        value = str(answers.get(question_id) or "").strip()
        if not value:
            continue
        question = lookup.get(question_id) or {}
        label = str(question.get("text") or "").strip() or "Вопрос"
        summary.append(f"{label}: {value}")
    return summary


def _insert_ticket_message(
    conn,
    *,
    group_msg_id,
    user_id,
    business,
    location_type,
    city,
    location_name,
    problem,
    created_at,
    username,
    category,
    ticket_id,
    created_date,
    created_time,
    client_name,
    client_status,
    updated_at,
    updated_by,
    channel_id,
):
    exec_with_retry(
        lambda: conn.execute(
            """
      INSERT INTO messages(
        group_msg_id, user_id, business, location_type, city, location_name,
        problem, created_at, username, category, ticket_id, created_date,
        created_time, client_status, client_name, updated_at, updated_by, channel_id
      ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """,
            (
                group_msg_id,
                user_id,
                business,
                location_type,
                city,
                location_name,
                problem,
                created_at,
                username,
                category,
                ticket_id,
                created_date,
                created_time,
                client_status,
                client_name,
                updated_at,
                updated_by,
                channel_id,
            ),
        )
    )

    try:
        row = exec_with_retry(
            lambda: conn.execute(
                "SELECT user FROM ticket_active WHERE ticket_id=? ORDER BY last_seen DESC LIMIT 1",
                (ticket_id,),
            )
        ).fetchone()
        target_user = row["user"] if row else None
        if target_user:
            exec_with_retry(
                lambda: conn.execute(
                    "INSERT INTO notifications(user, text, url) VALUES(?, ?, ?)",
                    (
                        target_user,
                        f"Новое сообщение в диалоге {ticket_id}",
                        f"/#open=ticket:{ticket_id}",
                    ),
                )
            )
        else:
            exec_with_retry(
                lambda: conn.execute(
                    "INSERT INTO notifications(user, text, url) VALUES(?, ?, ?)",
                    (
                        "all",
                        f"Новое сообщение в диалоге {ticket_id}",
                        f"/#open=ticket:{ticket_id}",
                    ),
                )
            )
    except Exception:
        pass


def _add_chat_history_entry(
    conn,
    *,
    ticket_id,
    sender,
    message,
    timestamp,
    message_type="text",
    attachment=None,
    channel_id=None,
    user_id=None,
):
    exec_with_retry(
        lambda: conn.execute(
            """
        INSERT INTO chat_history
          (user_id, ticket_id, sender, message, timestamp, message_type, attachment,
           tg_message_id, reply_to_tg_id, channel_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, NULL, NULL, ?)
    """,
            (
                user_id,
                ticket_id,
                sender,
                message,
                timestamp,
                message_type,
                attachment,
                channel_id,
            ),
        )
    )


def _create_web_form_session(
    channel_id: int,
    config: dict,
    answers: dict[str, str],
    message: str,
    *,
    client_name: str | None = None,
    client_contact: str | None = None,
    username: str | None = None,
) -> dict:
    ensure_web_form_schema()

    sanitized_answers: dict[str, str] = {}
    for key, value in (answers or {}).items():
        key_str = str(key)
        value_str = str(value or "").strip()
        if value_str:
            sanitized_answers[key_str] = value_str

    summary = _build_answers_summary(config, sanitized_answers)
    message_body = str(message or "").strip()
    combined_message = message_body
    if summary:
        combined_message = "Ответы формы:\n" + "\n".join(summary)
        if message_body:
            combined_message += "\n\n" + message_body

    field_values: dict[str, str] = {}
    for q_id, question in (config.get("question_lookup") or {}).items():
        if not isinstance(question, dict):
            continue
        if question.get("type") == "preset":
            preset = question.get("preset") or {}
            field = preset.get("field") if isinstance(preset, dict) else None
            field = str(field or "").strip()
            if field:
                value = sanitized_answers.get(q_id)
                if value:
                    field_values[field] = value

    for fallback_key in (
        "business",
        "location_type",
        "city",
        "location_name",
        "category",
        "client_status",
        "client_name",
    ):
        if fallback_key not in field_values and fallback_key in sanitized_answers:
            field_values[fallback_key] = sanitized_answers[fallback_key]

    final_client_name = (client_name or "").strip() or field_values.get("client_name")
    if not final_client_name:
        final_client_name = sanitized_answers.get("client_name")
    if not final_client_name:
        final_client_name = sanitized_answers.get("name") or sanitized_answers.get("full_name")
    if not final_client_name:
        final_client_name = "Клиент веб-формы"

    final_contact = (client_contact or "").strip()
    if not final_contact:
        for contact_key in ("contact", "phone", "email"):
            candidate = sanitized_answers.get(contact_key)
            if candidate:
                final_contact = candidate
                break

    final_username = (username or "").strip() or "web_form"

    now_local = dt.now()
    created_at = now_local.isoformat()
    created_date = now_local.strftime("%Y-%m-%d")
    created_time = now_local.strftime("%H:%M:%S")
    now_utc = dt.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    with get_db() as conn:
        conn.execute("BEGIN")
        try:
            user_id = _allocate_web_user_id(conn)

            ticket_id = None
            while True:
                candidate = f"web-{uuid4().hex[:8]}"
                exists = exec_with_retry(
                    lambda: conn.execute(
                        "SELECT 1 FROM tickets WHERE ticket_id = ?",
                        (candidate,),
                    )
                ).fetchone()
                if not exists:
                    ticket_id = candidate
                    break

            exec_with_retry(
                lambda: conn.execute(
                    "INSERT INTO tickets (user_id, group_msg_id, status, ticket_id, channel_id)"
                    " VALUES (?, NULL, 'pending', ?, ?)",
                    (user_id, ticket_id, channel_id),
                )
            )

            problem_text = combined_message or message_body
            _insert_ticket_message(
                conn,
                group_msg_id=None,
                user_id=user_id,
                business=field_values.get("business"),
                location_type=field_values.get("location_type"),
                city=field_values.get("city"),
                location_name=field_values.get("location_name"),
                problem=problem_text,
                created_at=created_at,
                username=final_username,
                category=field_values.get("category"),
                ticket_id=ticket_id,
                created_date=created_date,
                created_time=created_time,
                client_name=final_client_name,
                client_status=field_values.get("client_status"),
                updated_at=created_at,
                updated_by="web_form",
                channel_id=channel_id,
            )

            _add_chat_history_entry(
                conn,
                ticket_id=ticket_id,
                sender="user",
                message=combined_message or message_body,
                timestamp=now_utc,
                message_type="text",
                attachment=None,
                channel_id=channel_id,
                user_id=user_id,
            )

            answers_json = json.dumps(sanitized_answers, ensure_ascii=False)
            token = None
            for _ in range(6):
                candidate = secrets.token_urlsafe(16)
                try:
                    exec_with_retry(
                        lambda: conn.execute(
                            f"INSERT INTO {WEB_FORM_SESSIONS_TABLE} (token, ticket_id, channel_id, user_id, answers_json, client_name, client_contact, username, created_at, last_active_at)"
                            " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (
                                candidate,
                                ticket_id,
                                channel_id,
                                user_id,
                                answers_json,
                                final_client_name,
                                final_contact,
                                final_username,
                                created_at,
                                created_at,
                            ),
                        )
                    )
                    token = candidate
                    break
                except sqlite3.IntegrityError:
                    continue

            if not token:
                raise sqlite3.IntegrityError("Не удалось создать уникальный токен веб-формы")

            conn.commit()
        except Exception:
            conn.rollback()
            raise

    return {
        "token": token,
        "ticket_id": ticket_id,
        "answers": sanitized_answers,
        "summary": summary,
        "client_name": final_client_name,
        "client_contact": final_contact,
    }


def _load_public_form_session(channel_id: int, token: str):
    with get_db() as conn:
        session_row = exec_with_retry(
            lambda: conn.execute(
                f"""
                SELECT s.token, s.ticket_id, s.channel_id, s.user_id, s.answers_json,
                       s.client_name, s.client_contact, s.username,
                       s.created_at, s.last_active_at,
                       t.status, t.resolved_at, t.resolved_by
                  FROM {WEB_FORM_SESSIONS_TABLE} s
                  JOIN tickets t ON t.ticket_id = s.ticket_id AND t.channel_id = s.channel_id
                 WHERE s.channel_id = ? AND s.token = ?
                """,
                (channel_id, token),
            )
        ).fetchone()

        if not session_row:
            return None

        history_rows = exec_with_retry(
            lambda: conn.execute(
                """
                SELECT sender, message, timestamp, message_type, attachment
                  FROM chat_history
                 WHERE ticket_id = ? AND (channel_id = ? OR channel_id IS NULL)
                 ORDER BY substr(timestamp,1,19) ASC, rowid ASC
                """,
                (session_row["ticket_id"], channel_id),
            )
        ).fetchall()

    return session_row, history_rows

def format_time_duration(minutes):
    """
    Форматирует время в минутах в читаемый формат (часы и минуты)
    
    Args:
        minutes (int): количество минут
        
    Returns:
        str: отформатированная строка времени
    """
    if minutes is None:
        return "0 ч 0 мин"
    
    try:
        minutes = int(minutes)
    except (ValueError, TypeError):
        return "0 ч 0 мин"
    
    hours = minutes // 60
    mins = minutes % 60
    
    if hours == 0:
        return f"{mins} мин"
    elif mins == 0:
        return f"{hours} ч"
    else:
        return f"{hours} ч {mins} мин" 
    
# === Авторизация ===
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        with get_users_db() as conn:
            user = conn.execute(
                """
                SELECT u.*, r.name AS role_name, r.permissions AS role_permissions
                FROM users u
                LEFT JOIN roles r ON r.id = u.role_id
                WHERE LOWER(u.username) = LOWER(?)
                """,
                (username,),
            ).fetchone()
        if user and _check_user_password(user, password):
            # унифицируем ключи сессии, чтобы их видели и страницы, и API
            session.clear()
            resolved_username = _row_value(user, "username", username)
            session["user"] = resolved_username  # твой текущий ключ, на нём завязан login_required
            session["role"] = _row_value(user, "role_name") or _row_value(user, "role")

            # универсальные ключи для API и шаблонов (tasks.html берёт user_email)
            session["logged_in"] = True
            session["user_id"] = _row_value(user, "id")
            session["role_id"] = _row_value(user, "role_id")
            session["username"] = resolved_username
            session["user_email"] = resolved_username  # если реальной почты нет — используем username

            return redirect(url_for("index"))
        else:
            return "Неверный логин или пароль", 401
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# === Декоратор для защиты маршрутов ===
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# === Публичные эндпоинты веб-формы ===

@app.route("/public/forms/<string:channel_id>")
def public_form(channel_id: str):
    resolved_id = _resolve_channel_id(channel_id)
    if resolved_id is None:
        abort(404)
    config = _prepare_public_form_config(resolved_id)
    if not config:
        abort(404)
    token = (request.args.get("dialog") or request.args.get("token") or "").strip()
    channel_payload = config.get("channel", {})
    channel_ref = (channel_payload.get("public_id") or str(resolved_id)).strip()
    return render_template(
        "public_form.html",
        channel=channel_payload,
        channel_id=resolved_id,
        channel_ref=channel_ref,
        initial_token=token,
    )


@app.route("/api/public/forms/<string:channel_id>/config", methods=["GET"])
def api_public_form_config(channel_id: str):
    resolved_id = _resolve_channel_id(channel_id)
    if resolved_id is None:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404
    config = _prepare_public_form_config(resolved_id)
    if not config:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404
    return jsonify(_serialize_public_form_config(config))


def _parse_public_answers(payload: dict) -> dict:
    raw_answers = payload.get("answers")
    if isinstance(raw_answers, dict):
        return raw_answers
    if isinstance(raw_answers, str) and raw_answers.strip():
        try:
            parsed = json.loads(raw_answers)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


@app.route("/api/public/forms/<string:channel_id>/sessions", methods=["POST"])
def api_public_form_create_session(channel_id: str):
    resolved_id = _resolve_channel_id(channel_id)
    if resolved_id is None:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404
    config = _prepare_public_form_config(resolved_id)
    if not config:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404

    payload = request.get_json(silent=True) or {}
    if not payload and request.form:
        payload = request.form.to_dict(flat=True)

    message_text = str(payload.get("message") or "").strip()
    if not message_text:
        return jsonify({"success": False, "error": "Опишите проблему"}), 400

    answers = _parse_public_answers(payload)
    client_name = payload.get("client_name")
    client_contact = payload.get("contact") or payload.get("client_contact")
    username = payload.get("username")

    try:
        result = _create_web_form_session(
            resolved_id,
            config,
            answers,
            message_text,
            client_name=client_name,
            client_contact=client_contact,
            username=username,
        )
    except Exception as exc:
        logging.exception("Failed to create web form session: %s", exc)
        return jsonify({"success": False, "error": "Не удалось создать обращение"}), 500

    response = {
        "success": True,
        "token": result["token"],
        "ticket_id": result["ticket_id"],
        "channel": config.get("channel", {}),
    }
    return jsonify(response), 201


@app.route(
    "/api/public/forms/<string:channel_id>/sessions/<string:token>",
    methods=["GET"],
)
def api_public_form_session(channel_id: str, token: str):
    resolved_id = _resolve_channel_id(channel_id)
    if resolved_id is None:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404
    config = _prepare_public_form_config(resolved_id)
    if not config:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404

    loaded = _load_public_form_session(resolved_id, token)
    if not loaded:
        return jsonify({"success": False, "error": "Диалог не найден"}), 404

    session_row, history_rows = loaded
    try:
        answers = json.loads(session_row["answers_json"] or "{}") if session_row["answers_json"] else {}
    except Exception:
        answers = {}

    messages: list[dict] = []
    for row in history_rows:
        entry = {
            "sender": row["sender"],
            "message": row["message"] or "",
            "timestamp": row["timestamp"],
            "message_type": row["message_type"] or "text",
        }
        if row["attachment"]:
            entry["attachment"] = row["attachment"]
        messages.append(entry)

    payload = {
        "success": True,
        "session": {
            "token": session_row["token"],
            "ticket_id": session_row["ticket_id"],
            "channel_id": session_row["channel_id"],
            "channel": config.get("channel", {}),
            "status": session_row["status"],
            "resolved_at": session_row["resolved_at"],
            "resolved_by": session_row["resolved_by"],
            "client_name": session_row["client_name"],
            "client_contact": session_row["client_contact"],
            "created_at": session_row["created_at"],
            "last_active_at": session_row["last_active_at"],
            "answers": answers if isinstance(answers, dict) else {},
            "messages": messages,
        },
    }
    return jsonify(payload)


@app.route(
    "/api/public/forms/<string:channel_id>/sessions/<string:token>/messages",
    methods=["POST"],
)
def api_public_form_send_message(channel_id: str, token: str):
    resolved_id = _resolve_channel_id(channel_id)
    if resolved_id is None:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404
    config = _prepare_public_form_config(resolved_id)
    if not config:
        return jsonify({"success": False, "error": "Канал не найден или отключён"}), 404

    payload = request.get_json(silent=True) or {}
    if not payload and request.form:
        payload = request.form.to_dict(flat=True)

    text = str(payload.get("message") or "").strip()
    if not text:
        return jsonify({"success": False, "error": "Сообщение не может быть пустым"}), 400

    with get_db() as conn:
        conn.execute("BEGIN")
        try:
            session_row = exec_with_retry(
                lambda: conn.execute(
                    f"""
                    SELECT token, ticket_id, channel_id, user_id, client_name,
                           client_contact, username, answers_json
                      FROM {WEB_FORM_SESSIONS_TABLE}
                     WHERE channel_id = ? AND token = ?
                    """,
                    (resolved_id, token),
                )
            ).fetchone()

            if not session_row:
                conn.rollback()
                return jsonify({"success": False, "error": "Диалог не найден"}), 404

            user_id = session_row["user_id"]
            ticket_id = session_row["ticket_id"]
            username = session_row["username"] or "web_form"
            client_name = session_row["client_name"] or "Клиент веб-формы"

            stored_answers: dict[str, str] = {}
            try:
                parsed_answers = json.loads(session_row["answers_json"] or "{}") if session_row["answers_json"] else {}
                if isinstance(parsed_answers, dict):
                    stored_answers = parsed_answers
            except Exception:
                stored_answers = {}

            field_values: dict[str, str] = {}
            for q_id, question in (config.get("question_lookup") or {}).items():
                if not isinstance(question, dict):
                    continue
                if question.get("type") == "preset":
                    preset = question.get("preset") or {}
                    field = str((preset.get("field") if isinstance(preset, dict) else "") or "").strip()
                    if field:
                        value = stored_answers.get(q_id)
                        if isinstance(value, str) and value.strip():
                            field_values[field] = value.strip()

            for key in ("business", "location_type", "city", "location_name", "category", "client_status", "client_name"):
                value = stored_answers.get(key)
                if isinstance(value, str) and value.strip():
                    field_values.setdefault(key, value.strip())

            now_local = dt.now()
            created_at = now_local.isoformat()
            created_date = now_local.strftime("%Y-%m-%d")
            created_time = now_local.strftime("%H:%M:%S")
            now_utc = dt.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

            _insert_ticket_message(
                conn,
                group_msg_id=None,
                user_id=user_id,
                business=field_values.get("business"),
                location_type=field_values.get("location_type"),
                city=field_values.get("city"),
                location_name=field_values.get("location_name"),
                problem=text,
                created_at=created_at,
                username=username,
                category=field_values.get("category"),
                ticket_id=ticket_id,
                created_date=created_date,
                created_time=created_time,
                client_name=client_name,
                client_status=field_values.get("client_status"),
                updated_at=created_at,
                updated_by="web_form",
                channel_id=resolved_id,
            )

            _add_chat_history_entry(
                conn,
                ticket_id=ticket_id,
                sender="user",
                message=text,
                timestamp=now_utc,
                message_type="text",
                attachment=None,
                channel_id=resolved_id,
                user_id=user_id,
            )

            exec_with_retry(
                lambda: conn.execute(
                    "UPDATE tickets SET status='pending', resolved_at=NULL, resolved_by=NULL"
                    " WHERE ticket_id=? AND channel_id=?",
                    (ticket_id, resolved_id),
                )
            )

            exec_with_retry(
                lambda: conn.execute(
                    f"UPDATE {WEB_FORM_SESSIONS_TABLE} SET last_active_at=? WHERE token=?",
                    (created_at, token),
                )
            )

            conn.commit()
        except Exception as exc:
            conn.rollback()
            logging.exception("Failed to append message to web form session: %s", exc)
            return jsonify({"success": False, "error": "Не удалось отправить сообщение"}), 500

    return jsonify({"success": True})

from flask import jsonify, session

def _is_authenticated():
    # считаем авторизованным, если присутствует ЛЮБОЙ из признаков
    return bool(
        session.get('user') or
        session.get('logged_in') or
        session.get('user_id') or
        session.get('username') or
        session.get('user_email')
    )

def login_required_api(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not _is_authenticated():
            return jsonify({'error': 'auth', 'message': 'unauthorized'}), 401
        return f(*args, **kwargs)
    return wrapper

# === Отправка сообщения через Telegram API ===
def send_telegram_message(chat_id, text, parse_mode='HTML', **extra):
    """
    Отправка текста, поддерживает reply_to_message_id и пр. через **extra.
    Возвращает (ok: bool, result_or_error: dict|str)
    """
    try:
        # аккуратный фикс суррогатов (эмодзи)
        if isinstance(text, str):
            text = text.encode('utf-16', 'surrogatepass').decode('utf-16')

        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {
            'chat_id': chat_id,
            'text': text,
            'parse_mode': parse_mode,
            'disable_web_page_preview': True
        }
        if extra:
            payload.update(extra)

        r = requests.post(url, json=payload, timeout=20)
        j = r.json()
        if j.get('ok'):
            return True, j['result']  # здесь есть message_id
        return False, j.get('description', 'Unknown Telegram error')
    except Exception as e:
        return False, str(e)

# --- Починка "ломаных" эмодзи из браузера (lone surrogates) ---
def _fix_surrogates(s: str) -> str:
    if not isinstance(s, str):
        return s
    try:
        # перекодируем через UTF-16 с surrogatepass, чтобы склеить пары
        return s.encode('utf-16', 'surrogatepass').decode('utf-16')
    except Exception:
        return s
# --- Лимит 3 и запуск нового «спана» ---
def reopen_ticket_if_needed(ticket_id: str) -> bool | str:
    """
    Если тикет закрыт — переоткрываем (pending), увеличиваем reopen_count и открываем новый «спан».
    Лимит на переоткрытия: 3. Возвращает True | "LIMIT_EXCEEDED" | False.
    """
    try:
        conn = get_db()
        cur = conn.cursor()
        row = cur.execute(
            "SELECT status, COALESCE(reopen_count,0) AS rc FROM tickets WHERE ticket_id = ?",
            (ticket_id,)
        ).fetchone()

        if row and (row["status"] == "resolved"):
            if row["rc"] >= 3:
                return "LIMIT_EXCEEDED"
            cur.execute("""
                UPDATE tickets
                   SET status='pending',
                       resolved_by=NULL,
                       resolved_at=NULL,
                       reopen_count=COALESCE(reopen_count,0)+1
                 WHERE ticket_id=?
            """, (ticket_id,))
            conn.commit()
            # новый спан
            open_new_span(ticket_id)
            return True
        return False
    except Exception as e:
        app.logger.error(f"reopen_ticket_if_needed error: {e}")
        return False
    finally:
        try:
            conn.close()
        except Exception:
            pass

# === МАРШРУТЫ ===
from datetime import datetime as dt, timedelta, timezone
@app.route("/")
@login_required
@page_access_required("dialogs")
def index():
    settings = {"categories": ["Консультация", "Другое"]}
    if os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            settings = json.load(f)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
    SELECT 
        m.ticket_id,
        m.user_id,
        m.username as username,
        m.client_name as client_name,
        m.business,
        m.city,
        m.location_name,
        m.problem,
        m.created_at,
        t.status,
        t.resolved_by,
        t.resolved_at,
        m.created_date,
        m.created_time,
        cs.status as client_status
    FROM messages m
    LEFT JOIN tickets t ON m.ticket_id = t.ticket_id
    LEFT JOIN client_statuses cs 
        ON cs.user_id = m.user_id
        AND cs.updated_at = (
            SELECT MAX(updated_at) 
            FROM client_statuses 
            WHERE user_id = m.user_id
 )
    ORDER BY m.created_at DESC
    """)
    tickets = cur.fetchall()
    conn.close()

    total = len(tickets)
    resolved = len([t for t in tickets if t['status'] == 'resolved'])
    pending = total - resolved

    return render_template(
        "index.html",
        tickets=tickets,
        total=total,
        resolved=resolved,
        pending=pending,
        settings=settings,
    )

# serve_media для обработки изображений:
@app.route("/media/<ticket_id>/<path:filename>")
@login_required
def serve_media(ticket_id, filename):
    try:
        # Нормализуем идентификаторы и отбрасываем подкаталоги
        safe_ticket = secure_filename(ticket_id)
        base_name = os.path.basename(filename)  # режем "attachments/<id>/" и любую вложенность
        safe_name = secure_filename(base_name)

        file_path = os.path.join(ATTACHMENTS_DIR, safe_ticket, safe_name)

        if not (file_path.startswith(os.path.join(ATTACHMENTS_DIR, safe_ticket)) and os.path.isfile(file_path)):
            return "Файл не найден", 404

        mime_type, _ = mimetypes.guess_type(file_path)
        return send_file(file_path, mimetype=mime_type or 'application/octet-stream', as_attachment=False)
    except Exception as e:
        print(f"Ошибка при обслуживании медиафайла: {e}")
        return "Ошибка загрузки файла", 500

@app.route("/object_passports/media/<path:filename>")
@login_required
def object_passport_media(filename):
    safe_path = os.path.normpath(filename).replace("\\", "/")
    if safe_path.startswith("../") or safe_path.startswith("/"):
        abort(404)
    return send_from_directory(OBJECT_PASSPORT_UPLOADS_DIR, safe_path)

@app.route("/avatar/<int:user_id>")
@login_required
def avatar(user_id):
    """
    Отдаёт аватар клиента.
    - По умолчанию: маленькая версия (кэш: <user_id>.jpg)
    - При ?full=1: максимальная версия (кэш: <user_id>_full.jpg)
    Если фото нет — прозрачный пиксель.
    """
    import os, requests
    from flask import send_file, request

    AVA_DIR = os.path.join(ATTACHMENTS_DIR, "avatars")
    os.makedirs(AVA_DIR, exist_ok=True)

    want_full = (request.args.get("full", "").lower() in ("1", "true", "yes"))
    cache_name = f"{user_id}_full.jpg" if want_full else f"{user_id}.jpg"
    cache_path = os.path.join(AVA_DIR, cache_name)

    # 1) Есть в кэше — отдаём
    if os.path.exists(cache_path):
        return send_file(cache_path, mimetype="image/jpeg")

    # 2) Тянем из Telegram
    try:
        resp = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUserProfilePhotos",
            params={"user_id": user_id, "limit": 1},
            timeout=10
        ).json()
        photos = (resp.get("result") or {}).get("photos") or []
        if not photos:
            # прозрачный пиксель
            from io import BytesIO
            from PIL import Image
            img = Image.new("RGBA", (1, 1), (0, 0, 0, 0))
            bio = BytesIO()
            img.save(bio, format="PNG")
            bio.seek(0)
            return send_file(bio, mimetype="image/png")

        # маленькая или максимальная версия
        size_idx = -1 if want_full else 0
        file_id = photos[0][size_idx]["file_id"]

        resp2 = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getFile",
            params={"file_id": file_id},
            timeout=10
        ).json()
        file_path = (resp2.get("result") or {}).get("file_path")
        if not file_path:
            raise RuntimeError("file_path not found")

        url = f"https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}"
        binresp = requests.get(url, timeout=10)
        binresp.raise_for_status()
        with open(cache_path, "wb") as f:
            f.write(binresp.content)

        return send_file(cache_path, mimetype="image/jpeg")
    except Exception:
        # прозрачный пиксель на любой сбой
        from io import BytesIO
        from PIL import Image
        img = Image.new("RGBA", (1, 1), (0, 0, 0, 0))
        bio = BytesIO()
        img.save(bio, format="PNG")
        bio.seek(0)
        return send_file(bio, mimetype="image/png")


# === Обновление имени клиента ===
@app.route("/update_client_name", methods=["POST"])
@login_required
def update_client_name():
    data = request.json
    user_id = data["user_id"]
    client_name = data["client_name"]
    
    try:
        conn = get_db()
        # Обновляем имя во всех записях messages для этого пользователя
        conn.execute(
            "UPDATE messages SET client_name = ? WHERE user_id = ?",
            (client_name, user_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# === Обновление категории обращения ===
@app.route("/update_ticket_category", methods=["POST"])
@login_required
def update_ticket_category():
    data = request.json
    ticket_id = data["ticket_id"]
    category = data["category"]
    
    try:
        conn = get_db()
        conn.execute(
            "UPDATE messages SET category = ? WHERE ticket_id = ?",
            (category, ticket_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# === временная проверка можно удалить ===
@app.route("/api/whoami")
def whoami():
    # покажет, что реально лежит в session во время XHR
    from flask import jsonify, session, request
    return jsonify({
        "path": request.path,
        "cookies_sent": bool(request.cookies),
        "session_keys": list(session.keys()),
        "session": {
            k: session.get(k) for k in [
                "user", "role",
                "logged_in", "user_id", "username", "user_email"
            ]
        }
    })

# === временная проверка можно удалить ===
@app.route("/api/ping_auth")
def ping_auth():
    # быстро проверим, видит ли сервер вообще куку
    from flask import jsonify, session
    return jsonify({"ok": True, "has_user": bool(session.get("user"))})

# === Получение статуса клиента для карточки ===
@app.route("/client/<int:user_id>/status_info")
@login_required
def get_client_status(user_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT status FROM client_statuses WHERE user_id = ?",
            (user_id,)
        )
        status_row = cur.fetchone()
        conn.close()
        
        if status_row:
            return jsonify({"status": status_row['status']})
        else:
            return jsonify({"status": None})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/client/<int:user_id>/phones", methods=["GET"])
@login_required
def get_client_phones(user_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, phone, label, source, is_active, created_at, created_by
        FROM client_phones WHERE user_id=?
        ORDER BY source DESC, created_at DESC
    """, (user_id,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/client/<int:user_id>/phones", methods=["POST"])
@login_required
def add_client_phone(user_id):
    data = request.get_json(force=True)
    phone = (data.get("phone") or "").strip()
    label = (data.get("label") or "").strip() or None
    if not phone:
        return jsonify({"success": False, "error": "phone required"}), 400
    # простая нормализация
    phone_norm = ''.join(ch for ch in phone if ch.isdigit() or ch == '+')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO client_phones(user_id, phone, label, source, is_active, created_at, created_by)
        VALUES (?,?,?,?,1,?,?)
    """, (user_id, phone_norm, label, 'manual', dt.now().isoformat(), session.get("user","panel")))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"success": True, "id": new_id})

@app.route("/client/<int:user_id>/phones/<int:phone_id>", methods=["PATCH"])
@login_required
def update_client_phone(user_id, phone_id):
    data = request.get_json(force=True)
    label = data.get("label")
    is_active = data.get("is_active")
    sets, params = [], []
    if label is not None:
        sets.append("label = ?"); params.append(label)
    if is_active is not None:
        sets.append("is_active = ?"); params.append(1 if is_active else 0)
    if not sets:
        return jsonify({"success": False, "error": "nothing to update"}), 400
    params.extend([user_id, phone_id])
    conn = get_db()
    conn.execute(f"UPDATE client_phones SET {', '.join(sets)} WHERE user_id=? AND id=?", params)
    conn.commit(); conn.close()
    return jsonify({"success": True})

# === Обновление клиентской информации ===
@app.route("/update_client", methods=["POST"])
@login_required
def update_client():
    data = request.json
    user_id = data["user_id"]
    client_name = data.get("client_name", "")
    username = data.get("username", "")
    
    try:
        conn = get_db()
        # Обновляем имя и username во всех записях messages для этого пользователя
        conn.execute(
            "UPDATE messages SET client_name = ?, username = ? WHERE user_id = ?",
            (client_name, username, user_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.post("/message_edit")
@login_required
def message_edit():
    try:
        user_id = int(request.form["user_id"])
        ticket_id = request.form["ticket_id"]
        tg_message_id = int(request.form["tg_message_id"])
        new_text = _fix_surrogates(request.form.get("text", ""))

        # 1) пытаемся отредактировать сообщение у клиента (если это было сообщение поддержки)
        ok = False
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/editMessageText"
            payload = {
                "chat_id": user_id,
                "message_id": tg_message_id,
                "text": new_text,
                "parse_mode": "HTML",
                "disable_web_page_preview": True
            }
            r = requests.post(url, json=payload, timeout=15).json()
            ok = bool(r.get("ok"))
        except Exception:
            ok = False

        # 2) помечаем в истории (в базе панели)
        conn = get_db()
        from datetime import datetime, timezone
        edited_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00","Z")
        conn.execute(
          "UPDATE chat_history SET message=?, edited_at=? WHERE ticket_id=? AND tg_message_id=?",
          (new_text, edited_at, ticket_id, tg_message_id)
        )
        conn.commit(); conn.close()

        return jsonify({"success": True, "edited_on_client": ok})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.post("/message_delete")
@login_required
def message_delete():
    try:
        user_id = int(request.form["user_id"])
        ticket_id = request.form["ticket_id"]
        tg_message_id = int(request.form["tg_message_id"])

        # узнаем, кто отправил (support/user)
        conn = get_db()
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT sender FROM chat_history WHERE ticket_id=? AND tg_message_id=?",
            (ticket_id, tg_message_id)
        ).fetchone()
        conn.close()

        deleted_on_client = False
        if row and row["sender"] == "support":
            try:
                url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/deleteMessage"
                payload = {"chat_id": user_id, "message_id": tg_message_id}
                r = requests.post(url, json=payload, timeout=15).json()
                deleted_on_client = bool(r.get("ok"))
            except Exception:
                pass

        # мягкое удаление в истории — ВСЕГДА
        conn = get_db()
        deleted_at = datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        conn.execute(
            "UPDATE chat_history SET deleted_at=? WHERE ticket_id=? AND tg_message_id=?",
            (deleted_at, ticket_id, tg_message_id)
        )
        conn.commit(); conn.close()

        return jsonify({"success": True, "deleted_on_client": deleted_on_client})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/client_statuses")
@login_required
def get_client_statuses():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT user_id, status FROM client_statuses")
        statuses = {row['user_id']: row['status'] for row in cur.fetchall()}
        conn.close()
        return jsonify(statuses)
    except Exception as e:
        print(f"❌ Ошибка в /client_statuses: {e}")
        return jsonify({})

@app.route("/client/<int:user_id>/status", methods=["POST"])
@login_required
def set_client_status(user_id):
    data = request.json
    status = data.get("client_status")  # Изменено с "status" на "client_status"
    try: 
        conn = get_db()
        conn.execute(
            "INSERT OR REPLACE INTO client_statuses (user_id, status, updated_at) VALUES (?, ?, ?)",
            (user_id, status, datetime.datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
        
@app.route("/tickets_list")
@login_required
def tickets_list():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                m.ticket_id,
                m.user_id,
                m.username as username,
                COALESCE(
                    (SELECT client_name
                     FROM messages
                     WHERE user_id = m.user_id
                       AND client_name IS NOT NULL AND client_name != ''
                     ORDER BY created_at DESC LIMIT 1),
                m.client_name) AS client_name,
                m.username,
                m.business,
                m.city,
                m.location_name,
                m.problem,
                m.created_at,
                t.status AS raw_status,
                t.resolved_by,
                t.resolved_at,
                m.created_date,
                m.created_time,
                t.channel_id,
                (SELECT channel_name FROM channels c WHERE c.id = t.channel_id) AS channel_name,
                -- связанная задача (если создана из этого диалога)
                (SELECT tl.task_id FROM task_links tl WHERE tl.ticket_id = m.ticket_id LIMIT 1) AS linked_task_id,

                -- кто последний писал (с учётом канала)
                (SELECT sender
                 FROM chat_history h
                 WHERE h.ticket_id = m.ticket_id
                   AND (h.channel_id = t.channel_id OR h.channel_id IS NULL OR t.channel_id IS NULL)
                 ORDER BY h.rowid DESC
                 LIMIT 1
                ) AS last_sender,

                -- количество сообщений клиента (для подсчёта непрочитанных)
                (
                  SELECT COUNT(*)
                  FROM chat_history h
                  WHERE h.ticket_id = m.ticket_id
                    AND (h.channel_id = t.channel_id OR h.channel_id IS NULL OR t.channel_id IS NULL)
                    AND LOWER(COALESCE(h.sender, '')) = 'user'
                ) AS user_msg_count,

                -- время последнего сообщения клиента (может пригодиться на фронтенде)
                (
                  SELECT h.timestamp
                  FROM chat_history h
                  WHERE h.ticket_id = m.ticket_id
                    AND (h.channel_id = t.channel_id OR h.channel_id IS NULL OR t.channel_id IS NULL)
                    AND LOWER(COALESCE(h.sender, '')) = 'user'
                  ORDER BY h.rowid DESC
                  LIMIT 1
                ) AS last_user_message_at,

                -- был ли ответ поддержки (с учётом канала)
                EXISTS (
                  SELECT 1 FROM chat_history h2
                  WHERE h2.ticket_id = m.ticket_id
                    AND (h2.channel_id = t.channel_id OR h2.channel_id IS NULL OR t.channel_id IS NULL)
                    AND h2.sender != 'user'   -- 👈 ключевой момент: любой не user считается ответом панели
                  LIMIT 1
                ) AS has_support_reply

            FROM messages m
            JOIN tickets t ON t.ticket_id = m.ticket_id
            WHERE 1=1
            GROUP BY m.ticket_id
            ORDER BY m.created_at DESC
        """)
        tickets = cur.fetchall()

        result = []
        for t in tickets:
            # человеческий статус
            if not t['has_support_reply']:
                status_human = "Новая"
            else:
                # Если последнее сообщение от клиента
                if t['last_sender'] and t['last_sender'].lower() == 'user':
                    status_human = "Ожидает реакции"
                # Если последнее сообщение от кого-либо из панели (оператора, администратора и т.д.)
                elif t['last_sender']:
                    status_human = "Ожидает клиента"
                # Если нет информации о последнем отправителе
                else:
                    status_human = "Неизвестно"

            responsible_info = _resolve_ticket_responsible(cur, t['ticket_id'], t['channel_id'])

            row = dict(t)
            row['avatar_url'] = url_for('avatar', user_id=t['user_id'])
            row['status'] = status_human
            row['responsible'] = (responsible_info['responsible'] or '').strip()
            row['manual_responsible'] = (responsible_info['manual'] or '').strip()
            row['auto_responsible'] = (responsible_info['auto'] or '').strip()
            row['responsible_source'] = responsible_info['source'] or ''
            row['responsible_assigned_at'] = _coerce_to_iso(responsible_info['assigned_at'])
            row['responsible_assigned_by'] = (responsible_info['assigned_by'] or '').strip() if responsible_info['assigned_by'] else ''
            result.append(row)

        conn.close()
        return jsonify(result)
    except Exception as e:
        print(f"❌ Ошибка в /tickets_list: {e}")
        return jsonify([]), 500

def _coerce_to_iso(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        try:
            # timestamps in seconds vs milliseconds
            seconds = float(value)
            if seconds > 1e12:
                seconds = seconds / 1000.0
            dt_value = dt.fromtimestamp(seconds, tz=timezone.utc)
            return dt_value.isoformat().replace("+00:00", "Z")
        except (OSError, OverflowError, ValueError):
            return None

    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        if raw.isdigit():
            return _coerce_to_iso(int(raw))
        try:
            candidate = raw.replace("Z", "+00:00") if raw.endswith("Z") else raw
            dt_value = dt.fromisoformat(candidate)
            if dt_value.tzinfo is None:
                dt_value = dt_value.replace(tzinfo=timezone.utc)
            else:
                dt_value = dt_value.astimezone(timezone.utc)
            return dt_value.isoformat().replace("+00:00", "Z")
        except ValueError:
            try:
                return _coerce_to_iso(float(raw))
            except ValueError:
                return None

    return None


@app.route("/tickets/<ticket_id>")
@login_required
def get_ticket(ticket_id):
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT status, resolved_by, resolved_at, channel_id FROM tickets WHERE ticket_id = ?",
        (ticket_id,),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({})

    ticket = dict(row)
    ticket["resolved_at"] = _coerce_to_iso(ticket.get("resolved_at"))

    # Latest metadata about the client/location
    cur.execute(
        """
        SELECT business, location_name, client_name, username, channel_id
        FROM messages
        WHERE ticket_id = ?
        ORDER BY rowid DESC
        LIMIT 1
        """,
        (ticket_id,),
    )
    meta_row = cur.fetchone()
    if meta_row:
        meta_row = dict(meta_row)
        for key in ("business", "location_name", "client_name", "username", "channel_id"):
            value = meta_row.get(key)
            if value and not ticket.get(key):
                ticket[key] = value

    # Determine creation time from earliest message / ticket span
    created_at_raw = None
    cur.execute(
        """
        SELECT created_at, created_date, created_time
        FROM messages
        WHERE ticket_id = ?
        ORDER BY rowid ASC
        LIMIT 1
        """,
        (ticket_id,),
    )
    created_row = cur.fetchone()
    if created_row:
        created_row = dict(created_row)
        created_at_raw = created_row.get("created_at")
        if not created_at_raw:
            date_part = (created_row.get("created_date") or "").strip()
            time_part = (created_row.get("created_time") or "").strip()
            if date_part and time_part:
                created_at_raw = f"{date_part} {time_part}"

    cur.execute(
        """
        SELECT
            MIN(timestamp) AS first_message_at,
            MIN(CASE WHEN LOWER(COALESCE(sender,'')) = 'user' THEN timestamp END) AS first_user_message_at,
            MIN(CASE WHEN sender IS NOT NULL AND sender != '' AND LOWER(sender) != 'user' THEN timestamp END)
                AS first_support_message_at
        FROM chat_history
        WHERE ticket_id = ?
        """,
        (ticket_id,),
    )
    timeline_row = cur.fetchone()
    timeline = dict(timeline_row) if timeline_row else {}

    if not created_at_raw:
        created_at_raw = timeline.get("first_user_message_at") or timeline.get("first_message_at")

    if not created_at_raw:
        cur.execute(
            """
            SELECT started_at
            FROM ticket_spans
            WHERE ticket_id = ?
            ORDER BY span_no ASC
            LIMIT 1
            """,
            (ticket_id,),
        )
        span_row = cur.fetchone()
        if span_row:
            span_row = dict(span_row)
            created_at_raw = span_row.get("started_at")

    ticket["created_at"] = _coerce_to_iso(created_at_raw)
    ticket["first_user_message_at"] = _coerce_to_iso(timeline.get("first_user_message_at"))
    ticket["first_support_reply_at"] = _coerce_to_iso(timeline.get("first_support_message_at"))

    responsible_info = _resolve_ticket_responsible(cur, ticket_id, ticket.get("channel_id"))
    ticket["responsible"] = (responsible_info["responsible"] or "").strip()
    ticket["manual_responsible"] = (responsible_info["manual"] or "").strip()
    ticket["auto_responsible"] = (responsible_info["auto"] or "").strip()
    ticket["responsible_source"] = responsible_info["source"] or ""
    ticket["responsible_assigned_at"] = _coerce_to_iso(responsible_info["assigned_at"])
    ticket["responsible_assigned_by"] = responsible_info["assigned_by"] or ""

    conn.close()
    return jsonify(ticket)

@app.route("/tickets/<ticket_id>/category")
@login_required
def get_ticket_category(ticket_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT category FROM messages WHERE ticket_id = ? LIMIT 1", (ticket_id,))
    row = cur.fetchone()
    conn.close()
    
    if row and row['category']:
        return jsonify({"category": row['category']})
    return jsonify({"category": ""})

@app.route("/clients")
@login_required
@page_access_required("clients")
def clients_list():
    # фильтр по блэклисту: '1' (только в блэклисте), '0' (только не в блэклисте), '' или None (все)
    bl_filter = (request.args.get("blacklist") or "").strip()
    status_filter = (request.args.get("client_status") or "").strip()

    conn = get_db()
    cur = conn.cursor()

    # Базовый список клиентов
    cur.execute("""
        SELECT
            m.user_id,
            m.username,
            (SELECT client_name FROM messages
             WHERE user_id = m.user_id AND client_name IS NOT NULL AND client_name != ''
             ORDER BY created_at DESC LIMIT 1) as client_name,
            COUNT(*) as ticket_count,
            MIN(m.created_at) as first_contact,
            MAX(m.created_at) as last_contact
        FROM messages m
        GROUP BY m.user_id
        ORDER BY last_contact DESC
    """)
    clients = cur.fetchall()

    # Подтянем статусы блэклиста разом в словарь
    cur.execute("SELECT user_id, is_blacklisted, unblock_requested FROM client_blacklist")
    blmap = {r["user_id"]: (int(r["is_blacklisted"]), int(r["unblock_requested"])) for r in cur.fetchall()}

    # Подтянем статусы клиентов
    cur.execute("SELECT user_id, status FROM client_statuses")
    status_map = {}
    for row in cur.fetchall():
        status_value = (row["status"] or "").strip()
        if status_value:
            status_map[row["user_id"]] = status_value

    clients_with_time = []
    for client in clients:
        user_id = client['user_id']

        # Расчёт времени (как у вас было)
        time_query = """
            SELECT
                SUM(
                    CASE 
                        WHEN t.status = 'resolved' AND t.resolved_at IS NOT NULL 
                             AND ch.first_response_time IS NOT NULL
                        THEN CAST(
                            (julianday(t.resolved_at)
                             - julianday(replace(substr(ch.first_response_time,1,19),'T',' ')))
                             * 24 * 60
                             AS INTEGER
                        )
                        ELSE 0
                    END
                ) as total_minutes
            FROM messages m
            JOIN tickets t ON m.ticket_id = t.ticket_id
            LEFT JOIN (
                SELECT ticket_id, MIN(timestamp) as first_response_time
                FROM chat_history 
                WHERE sender = 'support'
                GROUP BY ticket_id
            ) ch ON t.ticket_id = ch.ticket_id
            WHERE m.user_id = ?
        """
        cur.execute(time_query, (user_id,))
        time_result = cur.fetchone()
        total_minutes = time_result['total_minutes'] or 0

        client_dict = dict(client)
        client_dict['total_minutes'] = total_minutes
        client_dict['formatted_time'] = format_time_duration(total_minutes)
        client_dict['client_status'] = status_map.get(user_id, "")

        # статусы блэклиста
        is_bl, unb = blmap.get(user_id, (0, 0))
        client_dict["blacklist"] = is_bl
        client_dict["unblock_requested"] = unb

        clients_with_time.append(client_dict)

    conn.close()

    # применим фильтр
    if bl_filter in ("0", "1"):
        clients_with_time = [c for c in clients_with_time if str(c["blacklist"]) == bl_filter]
    if status_filter:
        clients_with_time = [c for c in clients_with_time if (c.get("client_status") or "") == status_filter]

    return render_template(
        "clients.html",
        clients=clients_with_time,
        blacklist_filter=bl_filter,
        status_filter=status_filter
    )


def _clean_kb_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _serialize_kb_row(row: sqlite3.Row) -> dict[str, str | int]:
    return {
        "id": row["id"],
        "title": _clean_kb_value(row["title"]),
        "department": _clean_kb_value(row["department"]),
        "article_type": _clean_kb_value(row["article_type"]),
        "status": _clean_kb_value(row["status"]),
        "author": _clean_kb_value(row["author"]),
        "direction": _clean_kb_value(row["direction"]),
        "direction_subtype": _clean_kb_value(row["direction_subtype"]),
    }


def _fetch_knowledge_articles() -> list[dict[str, str | int]]:
    ensure_knowledge_base_schema()
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT id, title, department, article_type, status, author, direction, direction_subtype
            FROM knowledge_articles
            ORDER BY
                COALESCE(direction, '') COLLATE NOCASE,
                COALESCE(direction_subtype, '') COLLATE NOCASE,
                COALESCE(title, '') COLLATE NOCASE
            """
        ).fetchall()
    return [_serialize_kb_row(row) for row in rows]


def _serialize_kb_attachment(row: sqlite3.Row) -> dict[str, str | int]:
    stored_path = row["stored_path"] or ""
    download_url = url_for("knowledge_base_download_attachment", file_id=row["id"])
    inline_url = url_for("knowledge_base_view_attachment", file_id=row["id"])
    return {
        "id": row["id"],
        "name": row["original_name"] or os.path.basename(stored_path),
        "size": row["file_size"] or 0,
        "mime_type": row["mime_type"] or "application/octet-stream",
        "uploaded_at": row["uploaded_at"],
        "download_url": download_url,
        "inline_url": inline_url,
    }


def _kb_attachment_abs_path(stored_path: str) -> str:
    safe_relative = os.path.normpath(stored_path or "").replace("\\", "/")
    return os.path.join(KNOWLEDGE_BASE_ATTACHMENTS_DIR, safe_relative)


def _kb_attachment_rel_path(abs_path: str) -> str:
    relative = os.path.relpath(abs_path, KNOWLEDGE_BASE_ATTACHMENTS_DIR)
    return relative.replace("\\", "/")


def _fetch_kb_article(article_id: int) -> dict | None:
    ensure_knowledge_base_schema()
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT id, title, department, article_type, status, author, direction,
                   direction_subtype, summary, content, attachments, created_at, updated_at
            FROM knowledge_articles
            WHERE id = ?
            """,
            (article_id,),
        ).fetchone()
        if not row:
            return None
        attachments = conn.execute(
            """
            SELECT id, stored_path, original_name, mime_type, file_size, uploaded_at
            FROM knowledge_article_files
            WHERE article_id = ?
            ORDER BY uploaded_at DESC, id DESC
            """,
            (article_id,),
        ).fetchall()
    article = _serialize_kb_row(row)
    article.update(
        {
            "summary": _clean_kb_value(row["summary"]),
            "content": row["content"] or "",
            "attachments": json.loads(row["attachments"]) if row["attachments"] else [],
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
            "files": [_serialize_kb_attachment(item) for item in attachments],
        }
    )
    return article


def _now_iso() -> str:
    return dt.utcnow().replace(microsecond=0, tzinfo=timezone.utc).isoformat()


def _assign_attachments_to_article(conn: sqlite3.Connection, attachment_ids: list[int], article_id: int) -> list[dict]:
    if not attachment_ids:
        return []
    placeholders = ",".join(["?"] * len(attachment_ids))
    rows = conn.execute(
        f"""
        SELECT id, article_id, draft_token, stored_path, original_name, mime_type,
               file_size, uploaded_at
        FROM knowledge_article_files
        WHERE id IN ({placeholders})
        """,
        attachment_ids,
    ).fetchall()
    assigned = []
    for row in rows:
        if row["article_id"] and row["article_id"] != article_id:
            continue
        stored_path = row["stored_path"] or ""
        abs_path = _kb_attachment_abs_path(stored_path)
        if not os.path.isfile(abs_path):
            continue
        base_name = os.path.basename(abs_path)
        target_dir = os.path.join(KNOWLEDGE_BASE_ATTACHMENTS_DIR, f"article_{article_id}")
        os.makedirs(target_dir, exist_ok=True)
        target_path = os.path.join(target_dir, base_name)
        if os.path.abspath(os.path.dirname(abs_path)) != os.path.abspath(target_dir):
            if os.path.exists(target_path):
                stem, ext = os.path.splitext(base_name)
                target_path = os.path.join(
                    target_dir, f"{stem}_{uuid4().hex[:8]}{ext}"
                )
            shutil.move(abs_path, target_path)
            new_rel = _kb_attachment_rel_path(target_path)
        else:
            new_rel = _kb_attachment_rel_path(abs_path)

        conn.execute(
            """
            UPDATE knowledge_article_files
            SET article_id = ?, draft_token = NULL, stored_path = ?
            WHERE id = ?
            """,
            (article_id, new_rel, row["id"]),
        )
        assigned.append(row["id"])
    if assigned:
        conn.commit()
    return assigned


def _cleanup_orphan_attachments(conn: sqlite3.Connection, article_id: int, keep_ids: list[int]) -> None:
    keep_set = set(keep_ids or [])
    rows = conn.execute(
        """
        SELECT id, stored_path
        FROM knowledge_article_files
        WHERE article_id = ?
        """,
        (article_id,),
    ).fetchall()
    for row in rows:
        if row["id"] in keep_set:
            continue
        stored_path = row["stored_path"] or ""
        abs_path = _kb_attachment_abs_path(stored_path)
        conn.execute(
            "DELETE FROM knowledge_article_files WHERE id = ?",
            (row["id"],),
        )
        try:
            if os.path.isfile(abs_path):
                os.remove(abs_path)
        except Exception:
            pass
    conn.commit()


@app.route("/knowledge_base")
@app.route("/knowledge-base")
@login_required
@page_access_required("knowledge_base")
def knowledge_base_page():
    articles = _fetch_knowledge_articles()
    return render_template("knowledge_base.html", articles=articles)


@app.route("/knowledge_base/new")
@login_required
def knowledge_base_new_article():
    draft_token = uuid4().hex
    article = {
        "id": None,
        "title": "",
        "department": "",
        "article_type": "",
        "status": "",
        "author": "",
        "direction": "",
        "direction_subtype": "",
        "summary": "",
        "content": "",
        "files": [],
    }
    return render_template(
        "knowledge_base_article.html",
        article=article,
        draft_token=draft_token,
    )


@app.route("/knowledge_base/<int:article_id>")
@login_required
def knowledge_base_article_detail(article_id: int):
    article = _fetch_kb_article(article_id)
    if not article:
        abort(404)
    return render_template(
        "knowledge_base_article.html",
        article=article,
        draft_token=uuid4().hex,
    )


@app.route("/api/knowledge_base/articles", methods=["GET"])
@app.route("/api/knowledge-base/articles", methods=["GET"])
@login_required_api
def api_knowledge_base_articles():
    articles = _fetch_knowledge_articles()
    return jsonify({"items": articles, "count": len(articles)})


def _normalize_article_payload(payload: dict) -> tuple[dict, list[int]]:
    data = {
        "title": (payload.get("title") or "").strip(),
        "department": (payload.get("department") or "").strip(),
        "article_type": (payload.get("article_type") or "").strip(),
        "status": (payload.get("status") or "").strip(),
        "author": (payload.get("author") or "").strip(),
        "direction": (payload.get("direction") or "").strip(),
        "direction_subtype": (payload.get("direction_subtype") or "").strip(),
        "summary": (payload.get("summary") or "").strip(),
        "content": payload.get("content") or "",
    }
    attachment_ids = payload.get("attachments") or []
    if not isinstance(attachment_ids, list):
        attachment_ids = []
    clean_ids = []
    for value in attachment_ids:
        try:
            clean_ids.append(int(value))
        except (TypeError, ValueError):
            continue
    return data, clean_ids


@app.route("/api/knowledge_base/articles", methods=["POST"])
@app.route("/api/knowledge-base/articles", methods=["POST"])
@login_required_api
def api_create_knowledge_article():
    payload = request.json or {}
    data, attachment_ids = _normalize_article_payload(payload)
    if not data["title"]:
        return jsonify({"success": False, "error": "Укажите название статьи"}), 400

    now_iso = _now_iso()
    with get_db() as conn:
        cursor = conn.execute(
            """
            INSERT INTO knowledge_articles (
                title, department, article_type, status, author, direction,
                direction_subtype, summary, content, attachments, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["title"],
                data["department"],
                data["article_type"],
                data["status"],
                data["author"],
                data["direction"],
                data["direction_subtype"],
                data["summary"],
                data["content"],
                json.dumps(attachment_ids),
                now_iso,
                now_iso,
            ),
        )
        article_id = cursor.lastrowid
        if attachment_ids:
            assigned = _assign_attachments_to_article(conn, attachment_ids, article_id)
            data["attachments"] = assigned
        else:
            data["attachments"] = []
        conn.execute(
            "UPDATE knowledge_articles SET attachments = ? WHERE id = ?",
            (json.dumps(data["attachments"]), article_id),
        )
        conn.commit()

    article = _fetch_kb_article(article_id)
    return jsonify({"success": True, "item": article})


@app.route("/api/knowledge_base/articles/<int:article_id>", methods=["PUT"])
@app.route("/api/knowledge-base/articles/<int:article_id>", methods=["PUT"])
@login_required_api
def api_update_knowledge_article(article_id: int):
    article = _fetch_kb_article(article_id)
    if not article:
        return jsonify({"success": False, "error": "Статья не найдена"}), 404

    payload = request.json or {}
    data, attachment_ids = _normalize_article_payload(payload)
    if not data["title"]:
        return jsonify({"success": False, "error": "Укажите название статьи"}), 400

    now_iso = _now_iso()
    with get_db() as conn:
        conn.execute(
            """
            UPDATE knowledge_articles
            SET title = ?, department = ?, article_type = ?, status = ?, author = ?,
                direction = ?, direction_subtype = ?, summary = ?, content = ?,
                attachments = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                data["title"],
                data["department"],
                data["article_type"],
                data["status"],
                data["author"],
                data["direction"],
                data["direction_subtype"],
                data["summary"],
                data["content"],
                json.dumps(attachment_ids),
                now_iso,
                article_id,
            ),
        )
        if attachment_ids:
            _assign_attachments_to_article(conn, attachment_ids, article_id)
        _cleanup_orphan_attachments(conn, article_id, attachment_ids)
        conn.commit()

    article = _fetch_kb_article(article_id)
    return jsonify({"success": True, "item": article})


@app.route("/api/knowledge_base/articles/<int:article_id>", methods=["GET"])
@app.route("/api/knowledge-base/articles/<int:article_id>", methods=["GET"])
@login_required_api
def api_get_knowledge_article(article_id: int):
    article = _fetch_kb_article(article_id)
    if not article:
        return jsonify({"success": False, "error": "Статья не найдена"}), 404
    return jsonify({"success": True, "item": article})


def _save_kb_upload(file_storage, *, article_id: int | None, draft_token: str | None) -> tuple[int, dict] | tuple[None, str]:
    filename = secure_filename(file_storage.filename or "")
    if not filename:
        return None, "Недопустимое имя файла"

    if article_id:
        target_dir = os.path.join(
            KNOWLEDGE_BASE_ATTACHMENTS_DIR, f"article_{article_id}"
        )
    else:
        if not draft_token:
            draft_token = uuid4().hex
        target_dir = os.path.join(
            KNOWLEDGE_BASE_ATTACHMENTS_DIR, f"draft_{draft_token}"
        )
    os.makedirs(target_dir, exist_ok=True)

    unique_name = f"{int(time.time())}_{uuid4().hex}_{filename}"
    abs_path = os.path.join(target_dir, unique_name)
    file_storage.save(abs_path)
    file_size = os.path.getsize(abs_path)
    stored_path = _kb_attachment_rel_path(abs_path)

    with get_db() as conn:
        cursor = conn.execute(
            """
            INSERT INTO knowledge_article_files (
                article_id, draft_token, stored_path, original_name, mime_type, file_size
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                article_id,
                draft_token if not article_id else None,
                stored_path,
                file_storage.filename,
                file_storage.mimetype,
                file_size,
            ),
        )
        file_id = cursor.lastrowid
        conn.commit()

    metadata = {
        "id": file_id,
        "name": file_storage.filename,
        "size": file_size,
        "mime_type": file_storage.mimetype,
        "download_url": url_for("knowledge_base_download_attachment", file_id=file_id),
        "inline_url": url_for("knowledge_base_view_attachment", file_id=file_id),
    }
    return file_id, metadata


@app.route("/api/knowledge_base/uploads", methods=["POST"])
@app.route("/api/knowledge-base/uploads", methods=["POST"])
@login_required_api
def api_upload_knowledge_file():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "Файл не найден в запросе"}), 400
    file_storage = request.files["file"]
    if not file_storage or not (file_storage.filename or "").strip():
        return jsonify({"success": False, "error": "Выберите файл"}), 400

    article_id = request.form.get("article_id")
    draft_token = request.form.get("draft_token")
    try:
        article_id_int = int(article_id) if article_id else None
    except (TypeError, ValueError):
        article_id_int = None

    if article_id_int:
        existing = _fetch_kb_article(article_id_int)
        if not existing:
            return jsonify({"success": False, "error": "Статья не найдена"}), 404

    file_id, metadata_or_error = _save_kb_upload(
        file_storage,
        article_id=article_id_int,
        draft_token=draft_token,
    )
    if file_id is None:
        return jsonify({"success": False, "error": metadata_or_error}), 400
    return jsonify({"success": True, "file": metadata_or_error})


@app.route("/api/knowledge_base/uploads/<int:file_id>", methods=["DELETE"])
@app.route("/api/knowledge-base/uploads/<int:file_id>", methods=["DELETE"])
@login_required_api
def api_delete_knowledge_file(file_id: int):
    with get_db() as conn:
        row = conn.execute(
            "SELECT stored_path FROM knowledge_article_files WHERE id = ?",
            (file_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Файл не найден"}), 404
        conn.execute(
            "DELETE FROM knowledge_article_files WHERE id = ?",
            (file_id,),
        )
        conn.commit()
    stored_path = row["stored_path"] or ""
    abs_path = _kb_attachment_abs_path(stored_path)
    try:
        if os.path.isfile(abs_path):
            os.remove(abs_path)
    except Exception:
        pass
    return jsonify({"success": True})


@app.route("/knowledge_base/attachments/<int:file_id>")
@login_required
def knowledge_base_download_attachment(file_id: int):
    with get_db() as conn:
        row = conn.execute(
            "SELECT stored_path, original_name FROM knowledge_article_files WHERE id = ?",
            (file_id,),
        ).fetchone()
        if not row:
            abort(404)
    stored_path = row["stored_path"] or ""
    abs_path = _kb_attachment_abs_path(stored_path)
    if not os.path.isfile(abs_path):
        abort(404)
    directory = os.path.dirname(abs_path)
    filename = os.path.basename(abs_path)
    download_name = row["original_name"] or filename
    return send_from_directory(
        directory,
        filename,
        as_attachment=True,
        download_name=download_name,
    )


@app.route("/knowledge_base/attachments/<int:file_id>/view")
@login_required
def knowledge_base_view_attachment(file_id: int):
    with get_db() as conn:
        row = conn.execute(
            "SELECT stored_path, original_name, mime_type FROM knowledge_article_files WHERE id = ?",
            (file_id,),
        ).fetchone()
        if not row:
            abort(404)
    stored_path = row["stored_path"] or ""
    abs_path = _kb_attachment_abs_path(stored_path)
    if not os.path.isfile(abs_path):
        abort(404)
    directory = os.path.dirname(abs_path)
    filename = os.path.basename(abs_path)
    mimetype = row["mime_type"] or mimetypes.guess_type(filename)[0] or "application/octet-stream"
    return send_from_directory(
        directory,
        filename,
        mimetype=mimetype,
        as_attachment=False,
    )

# === Карточка клиента ===
@app.route("/client/<int:user_id>")
@login_required
def client_profile(user_id):
    conn = get_db()
    cur = conn.cursor()

    # Основная информация
    cur.execute("""
        SELECT user_id, username, client_name FROM messages WHERE user_id = ? LIMIT 1
    """, (user_id,))
    info = cur.fetchone()

    # Все заявки
    cur.execute("""
        SELECT 
            m.ticket_id,
            m.business,
            m.city,
            m.location_name,
            m.problem,
            m.created_at,
            t.status,
            t.resolved_by,
            CASE 
                WHEN t.resolved_at IS NOT NULL THEN datetime(t.resolved_at)
                ELSE NULL 
            END AS resolved_at,
            m.created_date,
            m.created_time,
            m.category,
            m.client_status
        FROM messages m
        LEFT JOIN tickets t ON m.ticket_id = t.ticket_id
        WHERE m.user_id = ?
        ORDER BY m.created_at DESC
    """, (user_id,))
    tickets = cur.fetchall()  # ✅ tickets — всегда определён

    # Добавляем время активности
    ticket_list = []
    for t in tickets:
        row = dict(t)
        if row['status'] == 'resolved' and row['resolved_at'] and row['created_date'] and row['created_time']:
            try:
                start = dt.fromisoformat(f"{row['created_date']} {row['created_time']}")
                end = dt.fromisoformat(row['resolved_at'])
                row['duration_minutes'] = int((end - start).total_seconds() // 60)
            except Exception as e:
                print(f"Ошибка расчёта времени: {e}")
                row['duration_minutes'] = None
        else:
            row['duration_minutes'] = None
        ticket_list.append(row)

    # Статистика
    cur.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN t.status = 'resolved' THEN 1 ELSE 0 END) as resolved,
            SUM(CASE WHEN t.status != 'resolved' THEN 1 ELSE 0 END) as pending
        FROM messages m
        LEFT JOIN tickets t ON m.ticket_id = t.ticket_id
        WHERE m.user_id = ?
    """, (user_id,))
    stats = cur.fetchone()

    # Оценки (если есть таблица feedbacks)
    try:
        cur.execute("""
            SELECT rating, timestamp FROM feedbacks WHERE user_id = ? ORDER BY timestamp DESC
        """, (user_id,))
        feedbacks = cur.fetchall()
    except:
        feedbacks = []

    # Получаем статус клиента из базы данных
    cur.execute("SELECT status FROM client_statuses WHERE user_id = ?", (user_id,))
    status_row = cur.fetchone()
    client_status = status_row['status'] if status_row else None

    # История username
    cur.execute(
        "SELECT username, seen_at FROM client_usernames WHERE user_id=? ORDER BY seen_at DESC",
        (user_id,),
    )
    username_history = [dict(r) for r in cur.fetchall()]

    # Телефоны
    cur.execute("""
        SELECT id, phone, label, source, is_active, created_at, created_by
        FROM client_phones
        WHERE user_id=? ORDER BY source DESC, created_at DESC
    """, (user_id,))
    phones_all = [dict(r) for r in cur.fetchall()]

    phones_telegram = [p for p in phones_all if p['source'] == 'telegram' and p['is_active']]
    phones_manual   = [p for p in phones_all if p['source'] == 'manual' and p['is_active']]

    # Данные блэклиста клиента
    cur.execute(
        """
        SELECT user_id, is_blacklisted, reason, added_at, added_by,
               unblock_requested, unblock_requested_at
        FROM client_blacklist
        WHERE user_id=?
        """,
        (user_id,),
    )
    client_blacklist_row = cur.fetchone()
    client_blacklist: dict[str, Any] | None = None
    if client_blacklist_row:
        client_blacklist = dict(client_blacklist_row)
        added_at_raw = client_blacklist.get("added_at")
        added_at_display = None
        if added_at_raw:
            try:
                added_at_display = dt.fromisoformat(str(added_at_raw).replace("Z", "")).strftime("%d.%m.%Y %H:%M")
            except Exception:
                added_at_display = added_at_raw
        client_blacklist["added_at_display"] = added_at_display

    cur.execute(
        """
        SELECT id, reason, created_at, status, decided_at, decided_by, decision_comment, channel_id
        FROM client_unblock_requests
        WHERE user_id=?
        ORDER BY created_at DESC, id DESC
        """,
        (user_id,),
    )
    unblock_requests_rows = cur.fetchall()
    unblock_requests: list[dict[str, Any]] = []
    for row in unblock_requests_rows:
        item = dict(row)
        created_at_raw = item.get("created_at")
        decided_at_raw = item.get("decided_at")
        if created_at_raw:
            try:
                item["created_at_display"] = dt.fromisoformat(str(created_at_raw).replace("Z", "")).strftime("%d.%m.%Y %H:%M")
            except Exception:
                item["created_at_display"] = created_at_raw
        else:
            item["created_at_display"] = None
        if decided_at_raw:
            try:
                item["decided_at_display"] = dt.fromisoformat(str(decided_at_raw).replace("Z", "")).strftime("%d.%m.%Y %H:%M")
            except Exception:
                item["decided_at_display"] = decided_at_raw
        else:
            item["decided_at_display"] = None
        unblock_requests.append(item)

    conn.close()

    settings = {"categories": ["Консультация", "Другое"]}
    if os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            settings = json.load(f)

    return render_template(
        "client_profile.html",
        client=dict(info),
        tickets=ticket_list,
        stats=dict(stats),
        feedbacks=[dict(f) for f in feedbacks],
        datetime=dt,
        settings=settings,
        client_status=client_status,
        username_history=username_history,
        phones_telegram=phones_telegram,
        phones_manual=phones_manual,
        client_blacklist=client_blacklist,
        unblock_requests=unblock_requests,
    )
       
# === API для дашборда с фильтрами ===
@app.route("/api/dashboard/data", methods=["POST"])
@login_required
def api_dashboard_data():
    try:
        data = request.get_json()
        start_date = data.get('startDate')
        end_date = data.get('endDate')
        restaurants = data.get('restaurants', [])
        
        # Если restaurants пустая строка, преобразуем в пустой список
        if restaurants == '':
            restaurants = []
        
        conn = get_db()
        cur = conn.cursor()
        
        # Базовый запрос для статистики
        query = """
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN t.status = 'new' THEN 1 ELSE 0 END) as new,
                SUM(CASE WHEN t.status = 'in_progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN t.status = 'resolved' THEN 1 ELSE 0 END) as resolved
            FROM tickets t
            JOIN messages m ON t.ticket_id = m.ticket_id
            WHERE 1=1
        """
        
        params = []
        
        # Фильтр по дате
        if start_date and end_date:
            query += " AND DATE(m.created_at) BETWEEN ? AND ?"
            params.extend([start_date, end_date])
                
        # Фильтр по ресторанам
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                query += f" AND m.location_name IN ({placeholders})"
                params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                query += " AND m.location_name = ?"
                params.append(restaurants)
                
        # Выполняем запрос для текущего периода
        cur.execute(query, params)
        current_result = cur.fetchone()
        
        # Безопасное преобразование результата в словарь
        current_stats = {}
        if current_result:
            current_stats = {
                'total': current_result['total'] or 0,
                'new': current_result['new'] or 0,
                'in_progress': current_result['in_progress'] or 0,
                'resolved': current_result['resolved'] or 0
            }
                
        # Запрос для предыдущего периода (месяц назад)
        prev_query = query
        prev_params = params.copy()
                
        if start_date and end_date:
            # Вычисляем предыдущий период такой же длительности
            start_dt = dt.strptime(start_date, '%Y-%m-%d')
            end_dt = dt.strptime(end_date, '%Y-%m-%d')
            period_days = (end_dt - start_dt).days
                    
            prev_start = (start_dt - timedelta(days=period_days + 1)).strftime('%Y-%m-%d')
            prev_end = (start_dt - timedelta(days=1)).strftime('%Y-%m-%d')
                    
            # Заменяем даты в параметрах
            if ' AND DATE(m.created_at) BETWEEN ? AND ?' in prev_query:
                date_index = prev_params.index(start_date)
                prev_params[date_index] = prev_start
                prev_params[date_index + 1] = prev_end
            
        cur.execute(prev_query, prev_params)
        previous_result = cur.fetchone()
        
        # Безопасное преобразование результата в словарь
        previous_stats = {}
        if previous_result:
            previous_stats = {
                'total': previous_result['total'] or 0,
                'new': previous_result['new'] or 0,
                'in_progress': previous_result['in_progress'] or 0,
                'resolved': previous_result['resolved'] or 0
            }

        # ВРЕМЕННАЯ СТАТИСТИКА - восстановленная логика
        time_query = """
            SELECT 
                SUM(
                    CASE 
                        WHEN t.status = 'resolved' AND t.resolved_at IS NOT NULL 
                        AND m.created_date IS NOT NULL AND m.created_time IS NOT NULL
                        THEN (
                            (julianday(t.resolved_at) - julianday(m.created_date || ' ' || m.created_time)) 
                            * 24 * 60  -- время в минутах
                        )
                        ELSE 0
                    END
                ) as total_minutes,
                COUNT(
                    CASE 
                        WHEN t.status = 'resolved' AND t.resolved_at IS NOT NULL 
                        THEN 1
                    END
                ) as resolved_count
            FROM tickets t
            JOIN messages m ON t.ticket_id = m.ticket_id
            WHERE 1=1
        """
        
        time_params = []
        
        # Фильтр по дате для временной статистики
        if start_date and end_date:
            time_query += " AND DATE(m.created_at) BETWEEN ? AND ?"
            time_params.extend([start_date, end_date])
            
        # Фильтр по ресторанам для временной статистики
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                time_query += f" AND m.location_name IN ({placeholders})"
                time_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                time_query += " AND m.location_name = ?"
                time_params.append(restaurants)
        
        # Выполняем запрос времени
        cur.execute(time_query, time_params)
        time_result = cur.fetchone()
        
        total_minutes = time_result['total_minutes'] or 0 if time_result else 0
        resolved_count = time_result['resolved_count'] or 0 if time_result else 0
        
        # Форматируем время для отображения
        avg_minutes = total_minutes / resolved_count if resolved_count > 0 else 0
        
        time_stats = {
            'total_minutes': total_minutes,
            'total_hours': total_minutes / 60,
            'resolved_count': resolved_count,
            'avg_minutes': avg_minutes,
            'formatted_total': format_time_duration(total_minutes),
            'formatted_avg': format_time_duration(avg_minutes) if resolved_count > 0 else "0 мин"
        }
        
        # СТАТИСТИКА ПО СОТРУДНИКАМ - восстановленная логика
        staff_time_query = """
            SELECT 
                t.resolved_by as staff_name,
                SUM(
                    CASE 
                        WHEN t.status = 'resolved' AND t.resolved_at IS NOT NULL 
                        AND m.created_date IS NOT NULL AND m.created_time IS NOT NULL
                        THEN (
                            (julianday(t.resolved_at) - julianday(m.created_date || ' ' || m.created_time)) 
                            * 24 * 60  -- время в минутах
                        )
                        ELSE 0
                    END
                ) as total_minutes,
                COUNT(
                    CASE 
                        WHEN t.status = 'resolved' AND t.resolved_at IS NOT NULL 
                        THEN 1
                    END
                ) as resolved_count,
                AVG(
                    CASE 
                        WHEN ch.first_response_time IS NOT NULL 
                        AND m.created_date IS NOT NULL AND m.created_time IS NOT NULL
                        THEN (
                            (julianday(ch.first_response_time) - julianday(m.created_date || ' ' || m.created_time)) 
                            * 24 * 60  -- время первого ответа в минутах
                        )
                    END
                ) as avg_first_response_minutes
            FROM tickets t
            JOIN messages m ON t.ticket_id = m.ticket_id
            LEFT JOIN (
                SELECT ticket_id, MIN(timestamp) as first_response_time
                FROM chat_history 
                WHERE sender = 'support'
                GROUP BY ticket_id
            ) ch ON t.ticket_id = ch.ticket_id
            WHERE t.resolved_by IS NOT NULL AND t.resolved_by != ''
        """
        
        staff_time_params = []
        
        if start_date and end_date:
            staff_time_query += " AND DATE(m.created_at) BETWEEN ? AND ?"
            staff_time_params.extend([start_date, end_date])
        
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                staff_time_query += f" AND m.location_name IN ({placeholders})"
                staff_time_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                staff_time_query += " AND m.location_name = ?"
                staff_time_params.append(restaurants)
        
        staff_time_query += " GROUP BY t.resolved_by"
        
        cur.execute(staff_time_query, staff_time_params)
        staff_time_results = cur.fetchall()
        
        staff_stats = []
        for row in staff_time_results:
            if row and row['staff_name'] and row['resolved_count'] > 0:
                staff_minutes = row['total_minutes'] or 0
                staff_resolved = row['resolved_count'] or 0
                staff_avg = staff_minutes / staff_resolved if staff_resolved > 0 else 0
                staff_avg_response = row['avg_first_response_minutes'] or 0
        
                staff_stats.append({
                    'name': row['staff_name'],
                    'total_minutes': staff_minutes,
                    'resolved_count': staff_resolved,
                    'avg_minutes': staff_avg,
                    'avg_response_minutes': staff_avg_response,
                    'formatted_total': format_time_duration(staff_minutes),
                    'formatted_avg': format_time_duration(staff_avg) if staff_resolved > 0 else "0 мин",
                    'formatted_avg_response': format_time_duration(staff_avg_response)
                })
                
        # Данные для графиков (остальная часть кода остается без изменений)
        charts_data = {}
                
        # Статусы заявок
        status_query = """
            SELECT t.status, COUNT(*) as count
            FROM tickets t
            JOIN messages m ON t.ticket_id = m.ticket_id
            WHERE 1=1
        """
                
        status_params = []
                
        if start_date and end_date:
            status_query += " AND DATE(m.created_at) BETWEEN ? AND ?"
            status_params.extend([start_date, end_date])
                
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                status_query += f" AND m.location_name IN ({placeholders})"
                status_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                status_query += " AND m.location_name = ?"
                status_params.append(restaurants)
                
        status_query += " GROUP BY t.status"
                
        cur.execute(status_query, status_params)
        status_rows = cur.fetchall()
        status_data = {row['status']: row['count'] for row in status_rows} if status_rows else {}
        charts_data['status'] = status_data
                
        # Категории обращений
        category_query = """
            SELECT category, COUNT(*) as count
            FROM messages
            WHERE 1=1
        """

        category_params = []
        if start_date and end_date:
            category_query += " AND DATE(created_at) BETWEEN ? AND ?"
            category_params.extend([start_date, end_date])
            
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                category_query += f" AND location_name IN ({placeholders})"
                category_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                category_query += " AND location_name = ?"
                category_params.append(restaurants)
            
        category_query += " GROUP BY category"
                
        cur.execute(category_query, category_params)
        category_rows = cur.fetchall()
        category_data = {row['category'] or 'без категории': row['count'] for row in category_rows} if category_rows else {}
        charts_data['category'] = category_data
                
        # По бизнесу
        business_query = """
            SELECT business, COUNT(*) as count
            FROM messages
            WHERE 1=1
        """
                
        business_params = []
                
        if start_date and end_date:
            business_query += " AND DATE(created_at) BETWEEN ? AND ?"
            business_params.extend([start_date, end_date])
                
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                business_query += f" AND location_name IN ({placeholders})"
                business_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                business_query += " AND location_name = ?"
                business_params.append(restaurants)
                
        business_query += " GROUP BY business"
                
        cur.execute(business_query, business_params)
        business_rows = cur.fetchall()
        business_data = {row['business'] or 'не указан': row['count'] for row in business_rows} if business_rows else {}
        charts_data['business'] = business_data

        # По сетям (Корпоративная сеть vs Партнёры-франчайзи)
        network_query = """
            SELECT 
                CASE 
                    WHEN m.location_type = 'Корпоративная сеть' AND m.business LIKE '%БлинБери%' THEN 'Корпоративная сеть (БлинБери)'
                    WHEN m.location_type = 'Корпоративная сеть' AND m.business LIKE '%СушиВесла%' THEN 'Корпоративная сеть (СушиВесла)'
                    WHEN m.location_type = 'Партнёры-франчайзи' AND m.business LIKE '%БлинБери%' THEN 'Партнёры-франчайзи (БлинБери)'
                    WHEN m.location_type = 'Партнёры-франчайзи' AND m.business LIKE '%СушиВесла%' THEN 'Партнёры-франчайзи (СушиВесла)'
                    WHEN m.location_type = 'Корпоративная сеть' THEN 'Корпоративная сеть (Другое)'
                    WHEN m.location_type = 'Партнёры-франчайзи' THEN 'Партнёры-франчайзи (Другое)'
                    ELSE 'Другое'
                END as network_type,
                COUNT(*) as count
            FROM messages m
            WHERE 1=1
        """

        network_params = []

        if start_date and end_date:
            network_query += " AND DATE(m.created_at) BETWEEN ? AND ?"
            network_params.extend([start_date, end_date])

        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                network_query += f" AND m.location_name IN ({placeholders})"
                network_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                network_query += " AND m.location_name = ?"
                network_params.append(restaurants)

        network_query += " GROUP BY network_type"

        cur.execute(network_query, network_params)
        network_rows = cur.fetchall()
        network_data = {row['network_type']: row['count'] for row in network_rows} if network_rows else {}
        charts_data['network'] = network_data

        # По городам
        city_query = """
            SELECT city, COUNT(*) as count
            FROM messages
            WHERE 1=1
        """
                
        city_params = []
                
        if start_date and end_date:
            city_query += " AND DATE(created_at) BETWEEN ? AND ?"
            city_params.extend([start_date, end_date])
                
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                city_query += f" AND location_name IN ({placeholders})"
                city_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                city_query += " AND location_name = ?"
                city_params.append(restaurants)
                
        city_query += " GROUP BY city"
                
        cur.execute(city_query, city_params)
        city_rows = cur.fetchall()
        city_data = {row['city'] or 'не указан': row['count'] for row in city_rows} if city_rows else {}
        charts_data['city'] = city_data
                
        # По ресторанам (топ-10)
        restaurant_query = """
            SELECT location_name, COUNT(*) as count
            FROM messages
            WHERE 1=1
        """
                
        restaurant_params = []
                
        if start_date and end_date:
            restaurant_query += " AND DATE(created_at) BETWEEN ? AND ?"
            restaurant_params.extend([start_date, end_date])
                
        if restaurants:
            if isinstance(restaurants, list) and restaurants:
                placeholders = ','.join(['?'] * len(restaurants))
                restaurant_query += f" AND location_name IN ({placeholders})"
                restaurant_params.extend(restaurants)
            elif isinstance(restaurants, str) and restaurants:
                restaurant_query += " AND location_name = ?"
                restaurant_params.append(restaurants)
                
        restaurant_query += " GROUP BY location_name ORDER BY count DESC LIMIT 10"
                
        cur.execute(restaurant_query, restaurant_params)
        restaurant_rows = cur.fetchall()
        restaurant_data = {row['location_name'] or 'не указан': row['count'] for row in restaurant_rows} if restaurant_rows else {}
        charts_data['restaurant'] = restaurant_data
                
        conn.close()
                
        return jsonify({
            'stats': {
                'current': current_stats,
                'previous': previous_stats
            },
            'charts': charts_data,
            'time_stats': time_stats,
            'staff_time_stats': staff_stats
        })
        
    except Exception as e:
        print(f"Ошибка в API дашборда: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# === API: уведомления (колокольчик) ===
@app.get('/api/notifications')
@login_required_api
def api_notifications():
    user = session.get('user_email') or session.get('user') or session.get('username') or 'all'
    since = request.args.get('since')
    params = [user,]
    where_since = ''
    if since:
        try:
            since_id = int(since)
            where_since = 'AND id > ?'
            params.append(since_id)
        except:
            pass
    conn = get_db()
    rows = conn.execute(f"""
        SELECT id, text, url, created_at
          FROM notifications
         WHERE (user = ? OR user = 'all') AND is_read = 0 {where_since}
         ORDER BY id DESC
         LIMIT 100
    """, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.post('/api/notifications/mark_read')
@login_required_api
def api_notifications_mark():
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get('ids') or []
    if not ids:
        return jsonify({'ok': True})
    try:
        conn = get_db()
        q = 'UPDATE notifications SET is_read=1 WHERE id IN ({})'.format(','.join(['?']*len(ids)))
        conn.execute(q, ids)
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# === Аналитика (заявки и клиенты) ===

def _prepare_ticket_analytics(conn):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT business, location_type, city, location_name, category, status, COUNT(*) as cnt
        FROM messages JOIN tickets USING(ticket_id)
        GROUP BY business, location_type, city, location_name, category, status
        """
    )
    rows = cur.fetchall()
    total_tickets = sum(row["cnt"] for row in rows)

    cur.execute("SELECT DISTINCT business FROM messages WHERE business IS NOT NULL")
    businesses = [row["business"] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT location_type FROM messages WHERE location_type IS NOT NULL")
    location_types = [row["location_type"] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT city FROM messages WHERE city IS NOT NULL")
    cities = [row["city"] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT location_name FROM messages WHERE location_name IS NOT NULL")
    locations = [row["location_name"] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT category FROM messages WHERE category IS NOT NULL")
    categories = [row["category"] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT status FROM tickets WHERE status IS NOT NULL")
    statuses = [row["status"] for row in cur.fetchall()]

    filters = {
        "businesses": businesses,
        "location_types": location_types,
        "cities": cities,
        "locations": locations,
        "categories": categories,
        "statuses": statuses,
    }
    return rows, total_tickets, filters


def _prepare_client_analytics(conn):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            m.user_id,
            COALESCE(
                (
                    SELECT client_name
                    FROM messages
                    WHERE user_id = m.user_id
                      AND client_name IS NOT NULL
                      AND client_name != ''
                    ORDER BY created_at DESC
                    LIMIT 1
                ),
                m.client_name
            ) AS client_name,
            m.username,
            m.business,
            m.location_type,
            m.city,
            m.location_name,
            m.category,
            t.status,
            COUNT(*) AS tickets
        FROM messages m
        JOIN tickets t ON m.ticket_id = t.ticket_id
        GROUP BY
            m.user_id,
            client_name,
            m.username,
            m.business,
            m.location_type,
            m.city,
            m.location_name,
            m.category,
            t.status
        ORDER BY tickets DESC
        """
    )
    rows = cur.fetchall()

    expanded_rows = []
    for row in rows:
        categories_raw = row["category"] or ""
        parts = [part.strip() for part in categories_raw.split(",") if part.strip()]
        if not parts:
            parts = ["без категории"]
        for part in parts:
            new_row = dict(row)
            new_row["category"] = part
            expanded_rows.append(new_row)

    total_tickets = sum(r["tickets"] for r in expanded_rows)

    cur.execute(
        """
        SELECT DISTINCT
            business,
            location_type,
            city,
            location_name,
            category
        FROM messages
        WHERE business IS NOT NULL
        """
    )
    filter_rows = cur.fetchall()
    businesses = sorted({row["business"] for row in filter_rows if row["business"]})
    location_types = sorted({row["location_type"] for row in filter_rows if row["location_type"]})
    cities = sorted({row["city"] for row in filter_rows if row["city"]})
    locations = sorted({row["location_name"] for row in filter_rows if row["location_name"]})
    categories = sorted({row["category"] for row in filter_rows if row["category"]})

    cur.execute("SELECT DISTINCT status FROM tickets WHERE status IS NOT NULL")
    statuses = [row["status"] for row in cur.fetchall()]

    filters = {
        "businesses": businesses,
        "location_types": location_types,
        "cities": cities,
        "locations": locations,
        "categories": categories,
        "statuses": statuses,
    }
    return expanded_rows, total_tickets, filters


@app.route("/analytics/clients")
@login_required
@page_access_required("analytics")
def analytics_clients():
    tab = request.args.get("tab") or "clients"
    return redirect(url_for("analytics", tab=tab))


@app.route("/analytics")
@login_required
@page_access_required("analytics")
def analytics():
    conn = get_db()
    try:
        ticket_rows, ticket_total, ticket_filters = _prepare_ticket_analytics(conn)
        client_rows, client_total, client_filters = _prepare_client_analytics(conn)
    finally:
        conn.close()

    selected_tab = request.args.get("tab") or "tickets"

    return render_template(
        "analytics.html",
        ticket_stats=ticket_rows,
        ticket_total=ticket_total,
        ticket_filters=ticket_filters,
        client_stats=client_rows,
        client_total=client_total,
        client_filters=client_filters,
        active_tab=selected_tab,
    )


# === Детали клиента для аналитики ===
@app.route("/analytics/clients/<int:user_id>/details")
@login_required
@page_access_required("analytics")
def client_analytics_details(user_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Основная информация о клиенте
        cur.execute("""
            SELECT username, client_name, MAX(created_at) as last_contact
            FROM messages 
            WHERE user_id = ?
            GROUP BY user_id
        """, (user_id,))
        client_info = cur.fetchone()
        
        # Общее количество заявок
        cur.execute("SELECT COUNT(*) as total_tickets FROM messages WHERE user_id = ?", (user_id,))
        total_tickets = cur.fetchone()['total_tickets']
        
        # Самая частая категория
        cur.execute("""
            SELECT category, COUNT(*) as cnt 
            FROM messages 
            WHERE user_id = ? AND category IS NOT NULL
            GROUP BY category 
            ORDER BY cnt DESC 
            LIMIT 1
        """, (user_id,))
        category_row = cur.fetchone()
        most_common_category = category_row['category'] if category_row else None
        
        # Самая частая локация
        cur.execute("""
            SELECT location_name, COUNT(*) as cnt 
            FROM messages 
            WHERE user_id = ? AND location_name IS NOT NULL
            GROUP BY location_name 
            ORDER BY cnt DESC 
            LIMIT 1
        """, (user_id,))
        location_row = cur.fetchone()
        most_common_location = location_row['location_name'] if location_row else None
        
        conn.close()
        
        return jsonify({
            'username': client_info['username'] if client_info else None,
            'client_name': client_info['client_name'] if client_info else None,
            'last_contact': client_info['last_contact'] if client_info else None,
            'total_tickets': total_tickets,
            'most_common_category': most_common_category,
            'most_common_location': most_common_location
        })
        
    except Exception as e:
        print(f"Ошибка получения деталей клиента: {e}")
        return jsonify({"error": str(e)}), 500

# === Экспорт аналитики клиентов ===
@app.route("/analytics/clients/export", methods=["POST"])
@login_required
def export_analytics_clients():
    try:
        data = request.json
        format_type = data.get('format', 'xlsx')
        filters = data.get('filters', {})
        export_filtered = data.get('exportFiltered', False)
        
        conn = get_db()
        cur = conn.cursor()
        
        # Базовый запрос для аналитики клиентов
        query = """
            SELECT 
                m.business,
                m.location_type,
                m.city,
                m.location_name,
                m.category,
                t.status,
                COUNT(DISTINCT m.user_id) as cnt
            FROM messages m
            JOIN tickets t ON m.ticket_id = t.ticket_id
            WHERE 1=1
        """
        
        # Применяем фильтры
        where_conditions = []
        params = []
        
        if export_filtered and filters:
            # Фильтры по бизнесу
            if filters.get('business') and filters['business']:
                placeholders = ','.join(['?'] * len(filters['business']))
                where_conditions.append(f"m.business IN ({placeholders})")
                params.extend(filters['business'])
            
            # Фильтры по типу локации
            if filters.get('location_type') and filters['location_type']:
                placeholders = ','.join(['?'] * len(filters['location_type']))
                where_conditions.append(f"m.location_type IN ({placeholders})")
                params.extend(filters['location_type'])
            
            # Фильтры по городу
            if filters.get('city') and filters['city']:
                placeholders = ','.join(['?'] * len(filters['city']))
                where_conditions.append(f"m.city IN ({placeholders})")
                params.extend(filters['city'])
            
            # Фильтры по локации
            if filters.get('location') and filters['location']:
                placeholders = ','.join(['?'] * len(filters['location']))
                where_conditions.append(f"m.location_name IN ({placeholders})")
                params.extend(filters['location'])
            
            # Фильтры по категории
            if filters.get('category') and filters['category']:
                placeholders = ','.join(['?'] * len(filters['category']))
                where_conditions.append(f"m.category IN ({placeholders})")
                params.extend(filters['category'])
            
            # Фильтры по статусу
            if filters.get('status') and filters['status']:
                placeholders = ','.join(['?'] * len(filters['status']))
                where_conditions.append(f"t.status IN ({placeholders})")
                params.extend(filters['status'])
            
            # Фильтр по количеству заявок
            if filters.get('ticketCount'):
                if filters['ticketCount'] == 'custom' and filters.get('customTicketValue'):
                    operator = filters.get('customTicketOperator', '>=')
                    where_conditions.append(f"COUNT(DISTINCT m.user_id) {operator} ?")
                    params.append(int(filters['customTicketValue']))
                else:
                    min_count = int(filters['ticketCount'])
                    where_conditions.append("COUNT(DISTINCT m.user_id) >= ?")
                    params.append(min_count)
        
        if where_conditions:
            query += " AND " + " AND ".join(where_conditions)
        
        query += " GROUP BY m.business, m.location_type, m.city, m.location_name, m.category, t.status"
        
        cur.execute(query, params)
        rows = cur.fetchall()
        
        conn.close()
        
        # Формируем данные для экспорта
        export_data = []
        headers = ['Бизнес', 'Тип локации', 'Город', 'Локация', 'Категория', 'Статус', 'Количество клиентов']
        
        for row in rows:
            export_data.append([
                row['business'] or '',
                row['location_type'] or '',
                row['city'] or '',
                row['location_name'] or '',
                row['category'] or '',
                row['status'] or '',
                row['cnt']
            ])
        
        # Экспорт в выбранном формате
        if format_type == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow(headers)
            
            # Данные
            writer.writerows(export_data)
            
            response = Response(output.getvalue(), mimetype='text/csv')
            response.headers['Content-Disposition'] = 'attachment; filename=clients_analytics_export.csv'
            
        else:  # xlsx используя openpyxl напрямую
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from io import BytesIO
            
            output = BytesIO()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Аналитика клиентов"
            
            # Заголовки с жирным шрифтом
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            # Данные
            for row_idx, row_data in enumerate(export_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Авто-ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output)
            output.seek(0)
            
            response = Response(output.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response.headers['Content-Disposition'] = 'attachment; filename=clients_analytics_export.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Ошибка при экспорте аналитики клиентов: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/analytics/export", methods=["POST"])
@login_required
def export_analytics():
    try:
        data = request.json
        format_type = data.get("format", "xlsx")
        filters = data.get("filters", {})
        export_filtered = data.get("exportFiltered", False)
        
        conn = get_db()
        cur = conn.cursor()
        
        # Базовый запрос
        query = """
            SELECT 
                business, 
                location_type, 
                city, 
                location_name, 
                category, 
                status, 
                COUNT(*) as cnt
            FROM messages 
            JOIN tickets USING(ticket_id)
        """
        
        # Применяем фильтры если нужно
        where_conditions = []
        params = []
        
        if export_filtered and filters:
            if filters.get('business') and filters['business']:
                where_conditions.append("business IN ({})".format(','.join(['?'] * len(filters['business']))))
                params.extend(filters['business'])
            
            if filters.get('location_type') and filters['location_type']:
                where_conditions.append("location_type IN ({})".format(','.join(['?'] * len(filters['location_type']))))
                params.extend(filters['location_type'])
            
            if filters.get('city') and filters['city']:
                where_conditions.append("city IN ({})".format(','.join(['?'] * len(filters['city']))))
                params.extend(filters['city'])
            
            if filters.get('location') and filters['location']:
                where_conditions.append("location_name IN ({})".format(','.join(['?'] * len(filters['location']))))
                params.extend(filters['location'])
            
            if filters.get('category') and filters['category']:
                where_conditions.append("category IN ({})".format(','.join(['?'] * len(filters['category']))))
                params.extend(filters['category'])
            
            if filters.get('status') and filters['status']:
                where_conditions.append("status IN ({})".format(','.join(['?'] * len(filters['status']))))
                params.extend(filters['status'])
        
        if where_conditions:
            query += " WHERE " + " AND ".join(where_conditions)
        
        query += " GROUP BY business, location_type, city, location_name, category, status"
        
        cur.execute(query, params)
        rows = cur.fetchall()
        
        # Формируем данные для экспорта
        export_data = []
        headers = ['Бизнес', 'Тип локации', 'Город', 'Локация', 'Категория', 'Статус', 'Количество']
        
        for row in rows:
            export_data.append([
                row['business'] or '',
                row['location_type'] or '',
                row['city'] or '',
                row['location_name'] or '',
                row['category'] or '',
                row['status'] or '',
                row['cnt']
            ])
        
        conn.close()
        
        # Экспорт в выбранном формате
        if format_type == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow(headers)
            
            # Данные
            writer.writerows(export_data)
            
            response = Response(output.getvalue(), mimetype='text/csv')
            response.headers['Content-Disposition'] = 'attachment; filename=analytics_export.csv'
            
        else:  # xlsx используя openpyxl напрямую
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from io import BytesIO
            
            output = BytesIO()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Аналитика"
            
            # Заголовки с жирным шрифтом
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            # Данные
            for row_idx, row_data in enumerate(export_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Авто-ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output)
            output.seek(0)
            
            response = Response(output.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response.headers['Content-Disposition'] = 'attachment; filename=analytics_export.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Ошибка при экспорте аналитики: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/users_list")
@login_required
@page_access_required("user_management")
def users_page():
    return render_template("users_list.html")

@app.route("/dashboard")
@login_required
@page_access_required("dashboard")
def dashboard():
    conn = get_db()
    cur = conn.cursor()

    # Получаем список всех ресторанов для фильтра
    cur.execute("SELECT DISTINCT location_name FROM messages WHERE location_name IS NOT NULL ORDER BY location_name")
    restaurants = [row['location_name'] for row in cur.fetchall()]

    # Остальной код остается без изменений...
    cur.execute("SELECT status, COUNT(*) AS cnt FROM tickets GROUP BY status")
    status_rows = cur.fetchall()
    status_data = {row['status'] if row['status'] is not None else 'неизвестно': row['cnt'] for row in status_rows}

    cur.execute("SELECT category, COUNT(*) AS cnt FROM messages GROUP BY category")
    category_rows = cur.fetchall()
    category_data = {row['category'] if row['category'] is not None else 'без категории': row['cnt'] for row in category_rows}

    cur.execute("SELECT business, COUNT(*) AS cnt FROM messages GROUP BY business")
    business_rows = cur.fetchall()
    business_data = {row['business'] if row['business'] is not None else 'не указан': row['cnt'] for row in business_rows}

    cur.execute("SELECT city, COUNT(*) AS cnt FROM messages GROUP BY city")
    city_rows = cur.fetchall()
    city_data = {row['city'] if row['city'] is not None else 'не указан': row['cnt'] for row in city_rows}

    conn.close()
    
    return render_template("dashboard.html", 
                         status_data=status_data, 
                         category_data=category_data,
                         business_data=business_data,
                         city_data=city_data,
                         restaurants=restaurants)  # Добавляем restaurants

# --- helpers for tasks people/notifications ---
def _get_people(conn, task_id: int, role: str) -> list[str]:
    rows = conn.execute(
        "SELECT identity FROM task_people WHERE task_id=? AND role=? ORDER BY identity",
        (task_id, role)
    ).fetchall()
    return [r["identity"] for r in rows]

def _add_people(conn, task_id: int, role: str, identities: list[str]) -> None:
    conn.execute("DELETE FROM task_people WHERE task_id=? AND role=?", (task_id, role))
    for ident in (identities or []):
        if ident:
            conn.execute(
                "INSERT OR IGNORE INTO task_people(task_id, role, identity) VALUES(?,?,?)",
                (task_id, role, ident.strip())
            )

def _notify_many(users: list[str], text: str, url: str | None = None) -> None:
    if not users:
        return
    with get_db() as nconn:
        for u in users:
            if u:
                nconn.execute(
                    "INSERT INTO notifications(user, text, url) VALUES(?,?,?)",
                    (u, text, url)
                )
        nconn.commit()

# === КАНАЛЫ (боты) ===
import requests

@app.route('/tasks')
@login_required
@page_access_required("tasks")
def tasks_page():
    import sqlite3
    users = []
    # 1) пытаемся достать пользователей из основной БД (если там есть таблица users)
    try:
        with get_db() as conn:
            conn.row_factory = sqlite3.Row
            users = conn.execute("SELECT id, username, role FROM users ORDER BY username").fetchall()
    except Exception:
        users = []

    # 2) если список пуст — пробуем users-ХРАНИЛИЩЕ (get_users_db)
    if not users:
        try:
            with get_users_db() as uconn:
                uconn.row_factory = sqlite3.Row
                users = uconn.execute("SELECT id, username, role FROM users ORDER BY username").fetchall()
        except Exception:
            users = []

    return render_template('tasks.html', users=users)

@app.route("/api/blacklist/check", methods=["GET"])
def api_blacklist_check():
    """
    Внешние интеграции (бот) могут дергать этот эндпоинт перед обработкой сообщения.
    Возвращает:
      - is_blacklisted: 0/1
      - unblock_requested: 0/1
      - message: текст для клиента (если блэклист)
    """
    user_id = (request.args.get("user_id") or "").strip()
    if not user_id:
        return jsonify({"ok": False, "error": "user_id required"}), 400

    row = None
    try:
        with get_db() as conn:
            row = conn.execute("""
                SELECT is_blacklisted, unblock_requested
                FROM client_blacklist WHERE user_id=?
            """, (user_id,)).fetchone()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    is_bl = int(row["is_blacklisted"]) if row else 0
    unb_req = int(row["unblock_requested"]) if row else 0
    msg = "Ваш аккаунт временно ограничен. Если вы считаете это ошибкой, нажмите «Запросить разблокировку»." if is_bl else ""
    return jsonify({"ok": True, "is_blacklisted": is_bl, "unblock_requested": unb_req, "message": msg})

@app.route("/api/blacklist/add", methods=["POST"])
@login_required
def api_blacklist_add():
    data = request.get_json(force=True) if request.is_json else request.form
    user_id = (data.get("user_id") or "").strip()
    reason = (data.get("reason") or "").strip()
    operator = session.get("username") or "operator"

    if not user_id:
        return jsonify({"ok": False, "error": "user_id required"}), 400

    now_iso = dt.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    try:
        with get_db() as conn:
            row_before = conn.execute(
                "SELECT unblock_requested FROM client_blacklist WHERE user_id=?",
                (user_id,),
            ).fetchone()
            had_pending = bool(row_before and row_before["unblock_requested"])
            conn.execute("""
                INSERT INTO client_blacklist (user_id, is_blacklisted, reason, added_at, added_by, unblock_requested, unblock_requested_at)
                VALUES (?, 1, ?, ?, ?, 0, NULL)
                ON CONFLICT(user_id) DO UPDATE SET
                    is_blacklisted=excluded.is_blacklisted,
                    reason=excluded.reason,
                    added_at=excluded.added_at,
                    added_by=excluded.added_by,
                    unblock_requested=0,
                    unblock_requested_at=NULL
            """, (user_id, reason, now_iso, operator))
            if had_pending:
                pending_row = conn.execute(
                    """
                    SELECT id FROM client_unblock_requests
                    WHERE user_id=? AND status='pending'
                    ORDER BY id DESC
                    LIMIT 1
                    """,
                    (user_id,),
                ).fetchone()
                if pending_row:
                    decision_comment = reason or "Оставлен в блокировке"
                    conn.execute(
                        """
                        UPDATE client_unblock_requests
                        SET status='rejected', decided_at=?, decided_by=?, decision_comment=?
                        WHERE id=?
                        """,
                        (
                            now_iso,
                            operator,
                            decision_comment,
                            pending_row["id"] if "id" in pending_row.keys() else pending_row[0],
                        ),
                    )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/blacklist/remove", methods=["POST"])
@login_required
def api_blacklist_remove():
    data = request.get_json(force=True) if request.is_json else request.form
    user_id = (data.get("user_id") or "").strip()
    operator = session.get("username") or "operator"

    if not user_id:
        return jsonify({"ok": False, "error": "user_id required"}), 400

    now_iso = dt.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    try:
        with get_db() as conn:
            row_before = conn.execute(
                "SELECT unblock_requested FROM client_blacklist WHERE user_id=?",
                (user_id,),
            ).fetchone()
            had_pending = bool(row_before and row_before["unblock_requested"])
            conn.execute("""
                INSERT INTO client_blacklist (user_id, is_blacklisted, reason, added_at, added_by, unblock_requested, unblock_requested_at)
                VALUES (?, 0, '', NULL, NULL, 0, NULL)
                ON CONFLICT(user_id) DO UPDATE SET
                    is_blacklisted=0,
                    reason='',
                    added_at=NULL,
                    added_by=NULL,
                    unblock_requested=0,
                    unblock_requested_at=NULL
            """, (user_id,))
            if had_pending:
                pending_row = conn.execute(
                    """
                    SELECT id FROM client_unblock_requests
                    WHERE user_id=? AND status='pending'
                    ORDER BY id DESC
                    LIMIT 1
                    """,
                    (user_id,),
                ).fetchone()
                if pending_row:
                    conn.execute(
                        """
                        UPDATE client_unblock_requests
                        SET status='approved', decided_at=?, decided_by=?, decision_comment='Разблокирован оператором'
                        WHERE id=?
                        """,
                        (
                            now_iso,
                            operator,
                            pending_row["id"] if "id" in pending_row.keys() else pending_row[0],
                        ),
                    )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/blacklist/reject-request", methods=["POST"])
@login_required
def api_blacklist_reject_request():
    data = request.get_json(force=True) if request.is_json else request.form
    user_id = (data.get("user_id") or "").strip()
    comment = (data.get("comment") or "").strip()
    operator = session.get("username") or "operator"

    if not user_id:
        return jsonify({"ok": False, "error": "user_id required"}), 400

    now_iso = dt.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    try:
        with get_db() as conn:
            pending_row = conn.execute(
                """
                SELECT id FROM client_unblock_requests
                WHERE user_id=? AND status='pending'
                ORDER BY id DESC
                LIMIT 1
                """,
                (user_id,),
            ).fetchone()
            if not pending_row:
                return jsonify({"ok": False, "error": "Нет активного запроса на разблокировку"}), 400

            decision_comment = comment or "Оставлен в блокировке"
            conn.execute(
                """
                UPDATE client_unblock_requests
                SET status='rejected', decided_at=?, decided_by=?, decision_comment=?
                WHERE id=?
                """,
                (
                    now_iso,
                    operator,
                    decision_comment,
                    pending_row["id"] if "id" in pending_row.keys() else pending_row[0],
                ),
            )
            conn.execute(
                """
                INSERT INTO client_blacklist (user_id, is_blacklisted, unblock_requested, unblock_requested_at)
                VALUES (?, 1, 0, NULL)
                ON CONFLICT(user_id) DO UPDATE SET
                    unblock_requested=0,
                    unblock_requested_at=NULL
                """,
                (user_id,),
            )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/blacklist/request-unblock", methods=["POST"])
def api_blacklist_request_unblock():
    """
    Вызывается, когда клиент нажимает «Запросить разблокировку».
    НЕ требует login_required, т.к. может прилетать извне (например, из бота).
    """
    data = request.get_json(force=True) if request.is_json else request.form
    user_id = (data.get("user_id") or "").strip()
    reason = (data.get("reason") or "").strip()
    channel_id_raw = data.get("channel_id")
    channel_id_val = None
    if channel_id_raw not in (None, ""):
        try:
            channel_id_val = int(channel_id_raw)
        except (TypeError, ValueError):
            channel_id_val = None

    if not user_id:
        return jsonify({"ok": False, "error": "user_id required"}), 400

    from datetime import datetime, timezone
    now_iso = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00","Z")
    try:
        created_new = False
        with get_db() as conn:
            conn.execute("""
                INSERT INTO client_blacklist (user_id, is_blacklisted, unblock_requested, unblock_requested_at)
                VALUES (?, 1, 1, ?)
                ON CONFLICT(user_id) DO UPDATE SET
                    unblock_requested=1,
                    unblock_requested_at=excluded.unblock_requested_at
            """, (user_id, now_iso))
            existing = conn.execute(
                """
                SELECT id FROM client_unblock_requests
                WHERE user_id=? AND status='pending'
                ORDER BY id DESC
                LIMIT 1
                """,
                (user_id,),
            ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE client_unblock_requests
                    SET reason=?, created_at=?, channel_id=?, status='pending',
                        decided_at=NULL, decided_by=NULL, decision_comment=NULL
                    WHERE id=?
                    """,
                    (
                        reason,
                        now_iso,
                        channel_id_val,
                        existing["id"] if "id" in existing.keys() else existing[0],
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO client_unblock_requests (user_id, channel_id, reason, created_at, status)
                    VALUES (?, ?, ?, ?, 'pending')
                    """,
                    (user_id, channel_id_val, reason, now_iso),
                )
                created_new = True
        if created_new:
            create_unblock_request_task(user_id, reason)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/tasks', methods=['GET'])
@login_required_api
def api_tasks_list():
    q = request.args
    cond, args = [], []

    # --- фильтры (как раньше) ---
    if q.get('num'):
        val = q['num']
        if '_' in val:
            src, num = val.split('_',1)
            cond.append("source = ? AND seq = ?")
            args += [src, int(num)]
        else:
            cond.append("seq = ?")
            args.append(int(val))
    if q.get('title'):
        cond.append("title LIKE ?"); args.append('%'+q['title']+'%')
    if q.get('assignee'):
        cond.append("assignee LIKE ?"); args.append('%'+q['assignee']+'%')
    if q.get('tag'):
        cond.append("tag LIKE ?"); args.append('%'+q['tag']+'%')
    if q.get('status'):
        cond.append("status = ?"); args.append(q['status'])
    # периоды
    if q.get('created_from'):
        cond.append("created_at >= ?"); args.append(q['created_from'] + " 00:00:00")
    if q.get('created_to'):
        cond.append("created_at <= ?"); args.append(q['created_to'] + " 23:59:59")
    if q.get('due_from'):
        cond.append("(due_at IS NOT NULL AND due_at >= ?)"); args.append(q['due_from'] + " 00:00:00")
    if q.get('due_to'):
        cond.append("(due_at IS NOT NULL AND due_at <= ?)"); args.append(q['due_to'] + " 23:59:59")
    if q.get('closed_from'):
        cond.append("(closed_at IS NOT NULL AND closed_at >= ?)"); args.append(q['closed_from'] + " 00:00:00")
    if q.get('closed_to'):
        cond.append("(closed_at IS NOT NULL AND closed_at <= ?)"); args.append(q['closed_to'] + " 23:59:59")

    where = ("WHERE " + " AND ".join(cond)) if cond else ""

    # --- сортировка (белый список) ---
    sort_by = (q.get('sort_by') or 'last_activity_at')
    sort_dir = (q.get('sort_dir') or 'desc').lower()
    allowed_cols = {'seq','source','title','assignee','due_at','last_activity_at','created_at','closed_at','tag','status'}
    if sort_by not in allowed_cols: sort_by = 'last_activity_at'
    if sort_dir not in {'asc','desc'}: sort_dir = 'desc'
    order_sql = f"ORDER BY {sort_by} {sort_dir.upper()}"

    # --- пагинация ---
    try:
        page = max(1, int(q.get('page') or 1))
    except: page = 1
    try:
        page_size = min(200, max(5, int(q.get('page_size') or 20)))
    except: page_size = 20
    offset = (page - 1) * page_size

    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        total = conn.execute(f"SELECT COUNT(*) AS c FROM tasks {where}", args).fetchone()['c']
        rows = conn.execute(f"""
            SELECT * FROM tasks {where}
            {order_sql}
            LIMIT ? OFFSET ?
        """, (*args, page_size, offset)).fetchall()

        items = [{
            'id': r['id'],
            'display_no': (f"{r['source']}_{r['seq']}" if r['source'] else str(r['seq'])),
            'title': r['title'],
            'assignee': r['assignee'],
            'due_at': r['due_at'],
            'last_activity_at': r['last_activity_at'],
            'created_at': r['created_at'],
            'closed_at': r['closed_at'],
            'tag': r['tag'],
            'status': r['status'],
        } for r in rows]

        return jsonify({
            'items': items,
            'total': total,
            'page': page,
            'page_size': page_size
        })

@app.route('/api/tasks/export', methods=['GET'])
@login_required_api
def api_tasks_export():
    # Берём всё без пагинации — повторяем условия из api_tasks_list
    q = request.args
    cond, args = [], []
    def add_like(field,key):
        if q.get(key):
            cond.append(f"{field} LIKE ?"); args.append('%'+q[key]+'%')
    if q.get('num'):
        val = q['num']
        if '_' in val:
            src, num = val.split('_',1)
            cond.append("source = ? AND seq = ?"); args += [src, int(num)]
        else:
            cond.append("seq = ?"); args.append(int(val))
    add_like('title','title')
    add_like('assignee','assignee')
    add_like('tag','tag')
    if q.get('status'):
        cond.append("status = ?"); args.append(q['status'])
    if q.get('created_from'):
        cond.append("created_at >= ?"); args.append(q['created_from'] + " 00:00:00")
    if q.get('created_to'):
        cond.append("created_at <= ?"); args.append(q['created_to'] + " 23:59:59")
    if q.get('due_from'):
        cond.append("(due_at IS NOT NULL AND due_at >= ?)"); args.append(q['due_from'] + " 00:00:00")
    if q.get('due_to'):
        cond.append("(due_at IS NOT NULL AND due_at <= ?)"); args.append(q['due_to'] + " 23:59:59")
    if q.get('closed_from'):
        cond.append("(closed_at IS NOT NULL AND closed_at >= ?)"); args.append(q['closed_from'] + " 00:00:00")
    if q.get('closed_to'):
        cond.append("(closed_at IS NOT NULL AND closed_at <= ?)"); args.append(q['closed_to'] + " 23:59:59")
    where = ("WHERE " + " AND ".join(cond)) if cond else ""

    sort_by = (q.get('sort_by') or 'last_activity_at')
    sort_dir = (q.get('sort_dir') or 'desc').lower()
    allowed_cols = {'seq','source','title','assignee','due_at','last_activity_at','created_at','closed_at','tag','status'}
    if sort_by not in allowed_cols: sort_by = 'last_activity_at'
    if sort_dir not in {'asc','desc'}: sort_dir = 'desc'
    order_sql = f"ORDER BY {sort_by} {sort_dir.upper()}"

    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(f"SELECT * FROM tasks {where} {order_sql}").fetchall()

    from io import StringIO
    import csv
    si = StringIO()
    cw = csv.writer(si, delimiter=';')
    cw.writerow(['№','Название','Исполнитель','Крайний срок','Последняя активность','Создано','Закрыто','Тег','Статус'])
    for r in rows:
        disp = (f"{r['source']}_{r['seq']}" if r['source'] else str(r['seq']))
        cw.writerow([disp, r['title'] or '', r['assignee'] or '', r['due_at'] or '', r['last_activity_at'] or '', r['created_at'] or '', r['closed_at'] or '', r['tag'] or '', r['status'] or ''])
    out = si.getvalue().encode('utf-8-sig')
    return Response(out, headers={
        'Content-Disposition': 'attachment; filename="tasks_export.csv"',
        'Content-Type': 'text/csv; charset=utf-8'
    })

@app.route('/api/tasks', methods=['POST'])
@login_required_api
def api_tasks_save():
    f = request.form
    files = request.files.getlist('files')

    id_ = (f.get('id') or '').strip()
    title = (f.get('title') or '').strip()
    body_html = f.get('body_html') or ''
    # авто-значения, если не передали
    current_user = session.get('user_email') or session.get('username') or ''
    creator = (f.get('creator') or '').strip() or current_user
    assignee = (f.get('assignee') or '').strip() or current_user

    co = [s.strip() for s in (f.get('co') or '').split(',') if s.strip()]
    watchers = [s.strip() for s in (f.get('watchers') or '').split(',') if s.strip()]
    tag = (f.get('tag') or '').strip()
    due_at = (f.get('due_at') or '').strip() or None
    status = (f.get('status') or 'Новая').strip() or 'Новая'

    # файлы до 250 МБ
    attachments = []
    if files:
        base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "attachments"))
        os.makedirs(base, exist_ok=True)
        for fl in files:
            if not fl.filename: continue
            size = getattr(fl, 'content_length', None)
            if size and size > 250 * 1024 * 1024:
                return jsonify({'error':'Файл превышает 250 МБ'}), 400
            safe = secure_filename(fl.filename)
            path = os.path.join(base, safe)
            fl.save(path)
            attachments.append(safe)
        if attachments:
            body_html += "<div class='mt-2'><b>Файлы:</b> " + ", ".join(attachments) + "</div>"

    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        if id_:
            # UPDATE
            conn.execute("""
                UPDATE tasks SET title=?, body_html=?, creator=?, assignee=?, tag=?, status=?, due_at=?,
                    last_activity_at = datetime('now'),
                    closed_at = CASE WHEN ?='Завершена' THEN COALESCE(closed_at, datetime('now')) ELSE NULL END
                WHERE id=?
            """, (title, body_html, creator, assignee, tag, status, due_at, status, id_))
            _add_people(conn, id_, 'co', co)
            _add_people(conn, id_, 'watcher', watchers)
            conn.execute("INSERT INTO task_history(task_id,text) VALUES(?,?)", (id_, f"Задача изменена ({status})"))
            _touch_activity(conn, id_)
            targets = set([assignee] + co + watchers); targets.discard('')
            _notify_many(list(targets), f"Обновление задачи «{title or 'без названия'}»", url=url_for('tasks_page'))
            return jsonify({'ok': True, 'id': id_})
        else:
            # INSERT
            seq = _next_task_seq(conn)
            source = None
            conn.execute("""
                INSERT INTO tasks(seq, source, title, body_html, creator, assignee, tag, status, due_at)
                VALUES(?,?,?,?,?,?,?,?,?)
            """, (seq, source, title, body_html, creator, assignee, tag, status, due_at))
            task_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
            _add_people(conn, task_id, 'co', co)
            _add_people(conn, task_id, 'watcher', watchers)
            conn.execute("INSERT INTO task_history(task_id,text) VALUES(?,?)", (task_id, "Задача создана"))
            _touch_activity(conn, task_id)
            targets = set([assignee] + co + watchers); targets.discard('')
            _notify_many(list(targets), f"Новая задача «{title or 'без названия'}»", url=url_for('tasks_page'))
            return jsonify({'ok': True, 'id': task_id})

@app.route('/api/tasks/<int:task_id>', methods=['GET'])
@login_required_api
def api_tasks_get(task_id):
    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        t = conn.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
        if not t:
            return jsonify({'error': 'not found'}), 404

        comments = conn.execute("""
            SELECT author, html, created_at
            FROM task_comments
            WHERE task_id=?
            ORDER BY created_at ASC
        """, (task_id,)).fetchall()

        history = conn.execute("""
            SELECT at, text
            FROM task_history
            WHERE task_id=?
            ORDER BY at DESC
        """, (task_id,)).fetchall()

        return jsonify({
            'id': t['id'],
            'display_no': _display_no(t['source'], t['seq']),
            'title': t['title'],
            'body_html': t['body_html'],
            'creator': t['creator'],
            'assignee': t['assignee'],
            'tag': t['tag'],
            'status': t['status'],
            'due_at': t['due_at'],
            'created_at': t['created_at'],
            'closed_at': t['closed_at'],
            'last_activity_at': t['last_activity_at'],
            'co': _get_people(conn, task_id, 'co'),
            'watchers': _get_people(conn, task_id, 'watcher'),
            'comments': [{'author': c['author'], 'html': c['html'], 'created_at': c['created_at']} for c in comments],
            'history':  [{'at': h['at'], 'text': h['text']} for h in history],
        })

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@login_required_api
def api_tasks_delete(task_id):
    with get_db() as conn:
        conn.execute("DELETE FROM tasks WHERE id=?", (task_id,))
    return jsonify({'ok': True})

@app.route('/api/tasks/from_dialog', methods=['POST'])
@login_required_api
def api_task_from_dialog():
    data = request.get_json(force=True, silent=True) or {}
    problem = (data.get('problem') or '').strip()
    location = (data.get('location') or '').strip()
    title = (data.get('title') or '').strip() or (problem[:60] if problem else "Задача из диалога")

    current_user = session.get('user_email') or session.get('username') or session.get('user') or ''
    creator = current_user
    assignee = current_user

    body_html = ""
    if problem:
        body_html += f"<p><b>Проблема:</b> {problem}</p>"
    if location:
        body_html += f"<p><b>Локация:</b> {location}</p>"

    with get_db() as conn:
        seq = _next_task_seq(conn)
        conn.execute("""
            INSERT INTO tasks(seq, source, title, body_html, creator, assignee, status)
            VALUES(?,?,?,?,?,?, 'Новая')
        """, (seq, 'DL', title, body_html, creator, assignee))
        task_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

        # связь задача ↔ диалог (если передан ticket_id)
        if data.get('ticket_id'):
            conn.execute("INSERT OR IGNORE INTO task_links(task_id, ticket_id) VALUES(?,?)",
                         (task_id, str(data['ticket_id'])))

        conn.execute("INSERT INTO task_history(task_id,text) VALUES(?,?)", (task_id, "Создано из диалога"))
        _touch_activity(conn, task_id)

    # уведомим участника
    try:
        _notify_many([assignee] if assignee else [], f"Новая задача «{title}» (из диалога)", url=url_for('tasks_page'))
    except:
        pass

    return jsonify({'ok': True, 'id': task_id})

@app.route('/api/tasks/<int:task_id>/comments', methods=['POST'])
@login_required_api
def api_task_add_comment(task_id):
    author = session.get('user_email') or session.get('user') or session.get('username') or 'user'
    html = (request.form.get('html') or '').strip()
    if not html:
        return jsonify({'ok': False, 'error': 'empty'}), 400
    with get_db() as conn:
        conn.execute("INSERT INTO task_comments(task_id, author, html) VALUES(?,?,?)", (task_id, author, html))
        row = conn.execute("SELECT id, author, html, created_at FROM task_comments WHERE rowid = last_insert_rowid()").fetchone()
        conn.execute("INSERT INTO task_history(task_id,text) VALUES(?,?)", (task_id, f"Комментарий от {author}"))
        _touch_activity(conn, task_id)
    return jsonify({'ok': True, 'item': dict(row)})

@app.route('/api/tickets/<ticket_id>/active', methods=['POST'])
@login_required_api
def api_ticket_set_active(ticket_id):
    user = session.get('user_email') or session.get('user') or session.get('username') or ''
    if not user:
        return jsonify({'ok': False}), 400
    with get_db() as conn:
        conn.execute("""
            INSERT INTO ticket_active(ticket_id, user, last_seen) VALUES(?,?,datetime('now'))
            ON CONFLICT(ticket_id) DO UPDATE SET user=excluded.user, last_seen=datetime('now')
        """, (ticket_id, user))
    return jsonify({'ok': True})

@app.route('/api/admin/users', methods=['GET'])
@login_required_api
def api_admin_users():
    try:
        with get_users_db() as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT username FROM users ORDER BY username").fetchall()
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

    users = [row['username'] for row in rows if row['username']]
    return jsonify({'users': users})

@app.route('/api/tickets/<ticket_id>/invite', methods=['POST'])
@login_required_api
def api_ticket_invite(ticket_id):
    ticket_id = str(ticket_id or '').strip()
    if not ticket_id:
        return jsonify({'success': False, 'error': 'Некорректный идентификатор заявки'}), 400

    data = request.get_json(force=True, silent=True) or {}
    invitee = (data.get('invitee') or '').strip()
    if not invitee:
        return jsonify({'success': False, 'error': 'Не указан пользователь'}), 400

    try:
        with get_users_db() as conn:
            conn.row_factory = sqlite3.Row
            user_row = conn.execute("SELECT username FROM users WHERE username = ?", (invitee,)).fetchone()
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500

    if not user_row:
        return jsonify({'success': False, 'error': 'Пользователь не найден'}), 404

    with get_db() as conn:
        ticket_row = conn.execute("SELECT 1 FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket_row:
        return jsonify({'success': False, 'error': 'Заявка не найдена'}), 404

    inviter = session.get('user_email') or session.get('username') or session.get('user') or 'оператор'
    text = f"{inviter} приглашает вас в диалог №{ticket_id}"
    url = url_for('index', ticket_id=ticket_id)

    try:
        _notify_many([invitee], text, url)
    except Exception as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500

    return jsonify({'success': True})


@app.route('/api/tickets/<ticket_id>/responsible', methods=['POST'])
@login_required_api
def api_ticket_responsible(ticket_id):
    ticket_id = str(ticket_id or '').strip()
    if not ticket_id:
        return jsonify({'success': False, 'error': 'Некорректный идентификатор заявки'}), 400

    data = request.get_json(force=True, silent=True) or {}
    responsible = (data.get('responsible') or '').strip()
    current_user = session.get('user_email') or session.get('username') or session.get('user') or ''

    with get_db() as conn:
        cur = conn.cursor()
        ticket_row = cur.execute("SELECT channel_id FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
        if not ticket_row:
            return jsonify({'success': False, 'error': 'Заявка не найдена'}), 404

        if responsible:
            cur.execute(
                """
                INSERT INTO ticket_responsibles(ticket_id, responsible, assigned_at, assigned_by)
                VALUES(?, ?, datetime('now'), ?)
                ON CONFLICT(ticket_id) DO UPDATE SET
                    responsible = excluded.responsible,
                    assigned_at = datetime('now'),
                    assigned_by = excluded.assigned_by
                """,
                (ticket_id, responsible, current_user),
            )
        else:
            cur.execute("DELETE FROM ticket_responsibles WHERE ticket_id = ?", (ticket_id,))

        info = _resolve_ticket_responsible(cur, ticket_id, ticket_row['channel_id'])

    return jsonify({
        'success': True,
        'responsible': (info['responsible'] or '').strip(),
        'manual': (info['manual'] or '').strip(),
        'auto': (info['auto'] or '').strip(),
        'source': info['source'] or '',
        'responsible_assigned_at': _coerce_to_iso(info['assigned_at']),
        'responsible_assigned_by': info['assigned_by'] or '',
    })

@app.route('/api/notifications/unread_count')
@login_required_api
def api_notify_count():
    user = session.get('user_email') or session.get('user') or session.get('username') or 'all'
    with get_db() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM notifications WHERE (user=? OR user='all') AND is_read=0",
            (user,)
        ).fetchone()
        return jsonify({'count': row['c'] if row else 0})
    with get_db() as conn:
        row = conn.execute("SELECT COUNT(*) AS c FROM notifications WHERE (user=? OR user='all') AND is_read=0", (user,)).fetchone()
        return jsonify({'count': row['c'] if row else 0})

@app.route('/api/notifications')
@login_required_api
def api_notify_list():
    user = session.get('user_email') or session.get('user') or session.get('username') or 'all'
    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, text, url, created_at
            FROM notifications
            WHERE user=? OR user='all'
            ORDER BY created_at DESC
            LIMIT 100
        """, (user,)).fetchall()
        items = [{'id': r['id'], 'text': r['text'], 'url': r['url'], 'created_at': r['created_at']} for r in rows]
        return jsonify({'items': items})
    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, text, url, created_at FROM notifications
            WHERE user=? OR user='all'
            ORDER BY created_at DESC
            LIMIT 100
        """, (user,)).fetchall()
        items = [{'id':r['id'],'text':r['text'],'url':r['url'],'created_at':r['created_at']} for r in rows]
        return jsonify({'items': items})

@app.route('/api/notifications/<int:nid>/read', methods=['POST'])
@login_required_api
def api_notify_read(nid):
    user = session.get('user_email') or session.get('user') or session.get('username') or 'all'
    with get_db() as conn:
        conn.execute(
            "UPDATE notifications SET is_read=1 WHERE id=? AND (user=? OR user='all')",
            (nid, user)
        )
    return jsonify({'ok': True})
    with get_db() as conn:
        conn.execute("UPDATE notifications SET is_read=1 WHERE id=? AND (user=? OR user='all')", (nid, user))
    return jsonify({'ok': True})

@app.route("/channels", methods=["GET"])
@login_required
@page_access_required("channels")
def channels_page():
    conn = get_db()
    rows = conn.execute("SELECT * FROM channels ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("channels.html", channels=rows)

@app.route("/api/channels", methods=["GET"])
@login_required
def api_channels_list():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM channels ORDER BY id DESC").fetchall()
        result = []
        for row in rows:
            data = dict(row)
            refreshed = _refresh_bot_identity_if_needed(conn, data)
            result.append(refreshed)
        return jsonify(result)
    finally:
        conn.close()

@app.route("/api/channels", methods=["POST"])
@login_required
def api_channels_create():
    data = request.get_json(force=True)
    token = (data.get("token") or "").strip()
    channel_name = (data.get("channel_name") or "").strip()
    raw_questions_cfg = data.get("questions_cfg") or {}
    max_questions = int(data.get("max_questions") or 0)
    is_active = 1 if data.get("is_active", True) else 0
    question_template_id = (data.get("question_template_id") or "").strip()
    rating_template_id = (data.get("rating_template_id") or "").strip()

    if not token or not channel_name:
        return jsonify({"success": False, "error": "token и channel_name обязательны"}), 400

    try:
        bot_display_name, bot_username = _fetch_bot_identity(token)
    except Exception as e:
        return jsonify({"success": False, "error": f"Ошибка проверки токена: {e}"}), 400

    bot_settings = _load_sanitized_bot_settings_payload()
    question_templates = bot_settings.get("question_templates") or []
    question_ids = {tpl.get("id") for tpl in question_templates if isinstance(tpl, dict) and tpl.get("id")}
    default_question_template_id = (
        bot_settings.get("active_template_id")
        if isinstance(bot_settings.get("active_template_id"), str)
        else None
    ) or (next(iter(question_ids)) if question_ids else None)

    rating_templates = bot_settings.get("rating_templates") or []
    rating_ids = {tpl.get("id") for tpl in rating_templates if isinstance(tpl, dict) and tpl.get("id")}
    default_rating_template_id = (
        bot_settings.get("active_rating_template_id")
        if isinstance(bot_settings.get("active_rating_template_id"), str)
        else None
    ) or (next(iter(rating_ids)) if rating_ids else None)

    if isinstance(raw_questions_cfg, str):
        try:
            questions_cfg = json.loads(raw_questions_cfg)
        except Exception:
            questions_cfg = {}
    elif isinstance(raw_questions_cfg, dict):
        questions_cfg = dict(raw_questions_cfg)
    else:
        questions_cfg = {}

    q_from_cfg = (questions_cfg.get("question_template_id")
        or questions_cfg.get("questionTemplateId")
        or "").strip()
    r_from_cfg = (questions_cfg.get("rating_template_id")
        or questions_cfg.get("ratingTemplateId")
        or "").strip()
    if q_from_cfg:
        question_template_id = q_from_cfg
    if r_from_cfg:
        rating_template_id = r_from_cfg

    if question_template_id not in question_ids:
        question_template_id = default_question_template_id
    if rating_template_id not in rating_ids:
        rating_template_id = default_rating_template_id

    if question_template_id:
        questions_cfg["question_template_id"] = question_template_id
    if rating_template_id:
        questions_cfg["rating_template_id"] = rating_template_id

    conn = get_db()
    cur = conn.cursor()
    public_id = _generate_unique_channel_public_id(cur)
    cur.execute(
        """
        INSERT INTO channels(token, bot_name, bot_username, channel_name, questions_cfg, max_questions, is_active, question_template_id, rating_template_id, public_id)
        VALUES (:token, :bot_name, :bot_username, :channel_name, :questions_cfg, :max_questions, :is_active, :question_template_id, :rating_template_id, :public_id)
        """,
        {
            "token": token,
            "bot_name": bot_display_name,
            "bot_username": bot_username,
            "channel_name": channel_name,
            "questions_cfg": json.dumps(questions_cfg, ensure_ascii=False),
            "max_questions": max_questions,
            "is_active": is_active,
            "question_template_id": question_template_id,
            "rating_template_id": rating_template_id,
            "public_id": public_id,
        },
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "public_id": public_id})

@app.route("/api/channels/<int:channel_id>", methods=["PATCH"])
@login_required
def api_channels_update(channel_id):
    data = request.get_json(force=True)
    if "question_template_id" in data:
        value = data.get("question_template_id")
        data["question_template_id"] = value.strip() if isinstance(value, str) else value
    if "rating_template_id" in data:
        value = data.get("rating_template_id")
        data["rating_template_id"] = value.strip() if isinstance(value, str) else value
    fields, params = [], {"id": channel_id}
    for k in ("channel_name", "questions_cfg", "max_questions", "is_active", "question_template_id", "rating_template_id"):
        if k in data:
            fields.append(f"{k} = :{k}")
            params[k] = json.dumps(data[k], ensure_ascii=False) if k == "questions_cfg" else data[k]
    if not fields:
        return jsonify({"success": False, "error": "Нет полей для обновления"}), 400

    conn = get_db()
    conn.execute(f"UPDATE channels SET {', '.join(fields)} WHERE id = :id", params)
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/channels/<int:channel_id>", methods=["DELETE"])
@login_required
def api_channels_delete(channel_id):
    conn = get_db()
    # Проверка привязок
    exists = conn.execute("SELECT 1 FROM tickets WHERE channel_id = ?", (channel_id,)).fetchone()
    if exists:
        conn.close()
        return jsonify({"success": False, "error": "Есть связанные заявки — удаление запрещено"}), 400
    conn.execute("DELETE FROM channels WHERE id = ?", (channel_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/settings")
@login_required
@page_access_required("settings")
def settings_page():
    settings = load_settings()
    locations_payload = load_locations()
    location_tree = locations_payload.get("tree", {}) if isinstance(locations_payload, dict) else {}
    city_names = sorted({
        city
        for brand in (location_tree or {}).values()
        for partner_types in (brand or {}).values()
        for city in (partner_types or {}).keys()
    })
    contract_usage = {}
    try:
        with get_passport_db() as conn:
            rows = conn.execute(
                """
                SELECT TRIM(COALESCE(network_contract_number, '')) AS contract_number,
                       COUNT(*) AS total
                FROM object_passports
                WHERE TRIM(COALESCE(network_contract_number, '')) != ''
                GROUP BY TRIM(COALESCE(network_contract_number, ''))
                """
            ).fetchall()
            contract_usage = {
                row["contract_number"]: row["total"]
                for row in rows
                if row["contract_number"]
            }
    except Exception:
        contract_usage = {}

    status_usage = {}
    status_conn = None
    try:
        status_conn = get_db()
        cur = status_conn.cursor()
        cur.execute(
            """
            SELECT status, COUNT(*) as total
            FROM client_statuses
            WHERE status IS NOT NULL AND TRIM(status) != ''
            GROUP BY status
            """
        )
        status_usage = {
            row["status"]: row["total"]
            for row in cur.fetchall()
            if row["status"]
        }
    except Exception:
        status_usage = {}
    finally:
        if status_conn is not None:
            status_conn.close()

    bot_question_presets = build_location_presets(
        location_tree, base_definitions=DEFAULT_BOT_PRESET_DEFINITIONS
    )

    return render_template(
        "settings.html",
        settings=settings,
        locations=locations_payload,
        cities=city_names,
        parameter_types=PARAMETER_TYPES,
        it_connection_categories=get_it_connection_categories(settings),
        it_connection_category_fields=DEFAULT_IT_CONNECTION_CATEGORY_FIELDS,
        contract_usage=contract_usage,
        status_usage=status_usage,
        bot_question_presets=bot_question_presets,
    )

@app.route("/settings", methods=["POST"])
@login_required
def update_settings():
    try:
        data = request.json or {}
        settings = load_settings()
        settings_modified = False

        if "auto_close_hours" in data:
            try:
                settings["auto_close_hours"] = int(data["auto_close_hours"])
            except (TypeError, ValueError):
                settings["auto_close_hours"] = data["auto_close_hours"]
            settings_modified = True

        if "auto_close_config" in data:
            auto_close_config = sanitize_auto_close_config(
                data.get("auto_close_config"), fallback_hours=settings.get("auto_close_hours", 24)
            )
            templates = auto_close_config.get("templates", [])
            active_id = auto_close_config.get("active_template_id")
            active_template = next((tpl for tpl in templates if tpl.get("id") == active_id), templates[0] if templates else None)
            if active_template:
                settings["auto_close_hours"] = active_template.get("hours", settings.get("auto_close_hours", 24))
            settings["auto_close_config"] = auto_close_config
            settings_modified = True

        if "categories" in data:
            settings["categories"] = [cat for cat in data.get("categories", []) if str(cat).strip()]
            settings_modified = True

        if "client_statuses" in data:
            settings["client_statuses"] = [status for status in data.get("client_statuses", []) if str(status).strip()]
            settings_modified = True

        if "network_profiles" in data:
            profiles = []
            for item in data.get("network_profiles") or []:
                if not isinstance(item, dict):
                    continue
                provider = (item.get("provider") or "").strip()
                contract_number = (item.get("contract_number") or "").strip()
                support_phone = (item.get("support_phone") or "").strip()
                legal_entity = (item.get("legal_entity") or "").strip()
                raw_restaurant_ids = item.get("restaurant_ids")
                restaurant_ids = []
                if isinstance(raw_restaurant_ids, (list, tuple)):
                    for entry in raw_restaurant_ids:
                        for chunk in re.split(r"[,;\n]", str(entry or "")):
                            value = chunk.strip()
                            if value and value not in restaurant_ids:
                                restaurant_ids.append(value)
                elif isinstance(raw_restaurant_ids, str):
                    for entry in re.split(r"[,;\n]", raw_restaurant_ids):
                        value = entry.strip()
                        if value and value not in restaurant_ids:
                            restaurant_ids.append(value)
                restaurant_id = (item.get("restaurant_id") or "").strip()
                if restaurant_id and restaurant_id not in restaurant_ids:
                    restaurant_ids.insert(0, restaurant_id)
                primary_restaurant_id = restaurant_ids[0] if restaurant_ids else restaurant_id
                if not any([provider, contract_number, support_phone, legal_entity, primary_restaurant_id]):
                    continue
                profile_payload = {
                    "provider": provider,
                    "contract_number": contract_number,
                    "support_phone": support_phone,
                    "legal_entity": legal_entity,
                    "restaurant_id": primary_restaurant_id,
                }
                if restaurant_ids:
                    profile_payload["restaurant_ids"] = restaurant_ids
                profiles.append(profile_payload)
            settings["network_profiles"] = profiles
            settings_modified = True

        if "bot_settings" in data:
            locations_payload = load_locations()
            location_tree = (
                locations_payload.get("tree", {})
                if isinstance(locations_payload, dict)
                else {}
            )
            definitions = build_location_presets(
                location_tree, base_definitions=DEFAULT_BOT_PRESET_DEFINITIONS
            )
            settings["bot_settings"] = sanitize_bot_settings(
                data.get("bot_settings"), definitions=definitions
            )
            settings_modified = True

        if any(
            key in data
            for key in (
                "dialog_category_templates",
                "dialog_question_templates",
                "dialog_completion_templates",
                "dialog_time_metrics",
            )
        ):
            dialog_config = settings.get("dialog_config")
            if not isinstance(dialog_config, dict):
                dialog_config = {}

            if "dialog_category_templates" in data:
                dialog_config["category_templates"] = _sanitize_category_templates(
                    data.get("dialog_category_templates"),
                    fallback=settings.get("categories"),
                )

            if "dialog_question_templates" in data:
                dialog_config["question_templates"] = _sanitize_question_templates(
                    data.get("dialog_question_templates")
                )

            if "dialog_completion_templates" in data:
                dialog_config["completion_templates"] = _sanitize_completion_templates(
                    data.get("dialog_completion_templates")
                )

            if "dialog_time_metrics" in data:
                dialog_config["time_metrics"] = _sanitize_time_metrics(
                    data.get("dialog_time_metrics")
                )

            settings["dialog_config"] = dialog_config
            settings_modified = True

        if settings_modified:
            settings = ensure_dialog_config(settings)
            save_settings(settings)

        if "locations" in data:
            save_locations(data["locations"])

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

def _parameter_usage_counts():
    usage: dict[str, dict] = {key: {} for key in PARAMETER_TYPES.keys()}
    it_usage = usage.setdefault(
        "it_connection",
        {
            "equipment_type": {},
            "equipment_vendor": {},
            "equipment_model": {},
            "equipment_status": {},
        },
    )
    try:
        with get_passport_db() as conn:
            for slug, column in PARAMETER_USAGE_MAPPING.items():
                try:
                    rows = conn.execute(
                        f"""
                        SELECT {column} AS value, COUNT(*) AS total
                        FROM object_passports
                        WHERE TRIM(COALESCE({column}, '')) != ''
                        GROUP BY {column}
                        """
                    ).fetchall()
                except sqlite3.OperationalError:
                    continue
                if slug == "it_connection":
                    bucket = it_usage.setdefault("equipment_type", {})
                else:
                    bucket = usage.setdefault(slug, {})
                for row in rows:
                    value = (row["value"] or "").strip()
                    if value:
                        bucket[value] = row["total"]

            equipment_columns = [
                ("equipment_type", "equipment_type"),
                ("equipment_vendor", "vendor"),
                ("equipment_model", "model"),
                ("equipment_status", "status"),
            ]
            for category, column in equipment_columns:
                try:
                    rows = conn.execute(
                        f"""
                        SELECT TRIM(COALESCE({column}, '')) AS value, COUNT(*) AS total
                        FROM object_passport_equipment
                        WHERE TRIM(COALESCE({column}, '')) != ''
                        GROUP BY TRIM(COALESCE({column}, ''))
                        """
                    ).fetchall()
                except sqlite3.OperationalError:
                    continue
                bucket = it_usage.setdefault(category, {})
                for row in rows:
                    value = (row["value"] or "").strip()
                    if value:
                        bucket[value] = bucket.get(value, 0) + row["total"]
    except Exception:
        return usage
    return usage


def _fetch_parameters_grouped(conn, *, include_deleted: bool = False):
    ensure_remote_access_defaults(conn)
    query = (
        "SELECT id, param_type, value, state, is_deleted, deleted_at, extra_json "
        "FROM settings_parameters "
    )
    if not include_deleted:
        query += "WHERE is_deleted = 0 "
    query += "ORDER BY param_type, value COLLATE NOCASE"
    try:
        rows = conn.execute(query).fetchall()
    except sqlite3.OperationalError as exc:
        if _is_missing_table_error(exc, "settings_parameters"):
            ensure_settings_parameters_schema()
            return {key: [] for key in PARAMETER_TYPES.keys()}
        raise
    usage = _parameter_usage_counts()
    it_usage = usage.get("it_connection") if isinstance(usage.get("it_connection"), dict) else {}
    grouped = {key: [] for key in PARAMETER_TYPES.keys()}
    it_connection_categories = get_it_connection_categories()
    for row in rows:
        slug = row["param_type"]
        if slug not in grouped:
            continue
        value = row["value"]
        normalized = (value or "").strip()
        extra_payload = {}
        raw_extra = row["extra_json"]
        if raw_extra:
            try:
                extra_payload = json.loads(raw_extra)
            except Exception:
                extra_payload = {}
        if slug == "it_connection":
            usage_count = 0
        else:
            bucket = usage.get(slug, {})
            usage_count = bucket.get(normalized, 0) if isinstance(bucket, dict) else 0

        entry = {
            "id": row["id"],
            "value": value,
            "state": row["state"] or "Активен",
            "is_deleted": bool(row["is_deleted"]),
            "deleted_at": row["deleted_at"],
            "usage_count": usage_count,
            "extra": extra_payload,
        }
        if slug == "it_connection":
            equipment_type = (extra_payload.get("equipment_type") or "").strip()
            equipment_vendor = (extra_payload.get("equipment_vendor") or "").strip()
            equipment_model = (extra_payload.get("equipment_model") or "").strip()
            equipment_status = (extra_payload.get("equipment_status") or "").strip()
            if not any((equipment_type, equipment_vendor, equipment_model, equipment_status)):
                parts = [part.strip() for part in (value or "").split("/") if part.strip()]
                if parts:
                    equipment_type = parts[0]
                if len(parts) >= 2:
                    equipment_vendor = parts[1]
                if len(parts) >= 3:
                    equipment_model = parts[2]
            if extra_payload is None:
                extra_payload = {}
            category_raw = (extra_payload.get("category") or "").strip()
            category_label = (extra_payload.get("category_label") or "").strip()
            if not category_raw:
                if equipment_type:
                    category_raw = "equipment_type"
                elif equipment_vendor:
                    category_raw = "equipment_vendor"
                elif equipment_model:
                    category_raw = "equipment_model"
                elif equipment_status:
                    category_raw = "equipment_status"
                else:
                    category_raw = "equipment_type"
            if category_raw not in it_connection_categories:
                fallback_label = category_label or category_raw
                it_connection_categories.setdefault(category_raw, fallback_label)
            category_field = DEFAULT_IT_CONNECTION_CATEGORY_FIELDS.get(category_raw)
            value_by_category = {
                "equipment_type": equipment_type,
                "equipment_vendor": equipment_vendor,
                "equipment_model": equipment_model,
                "equipment_status": equipment_status,
            }
            if category_field:
                effective_value = (value_by_category.get(category_field) or value or "").strip()
            else:
                effective_value = (value or "").strip()
            effective_label = (
                category_label
                or it_connection_categories.get(category_raw)
                or category_raw
            )
            extra_payload.update(
                {
                    "category": category_raw,
                    "category_label": effective_label,
                    "equipment_type": equipment_type,
                    "equipment_vendor": equipment_vendor,
                    "equipment_model": equipment_model,
                    "equipment_status": equipment_status,
                }
            )
            category_usage = 0
            if isinstance(it_usage, dict):
                bucket = it_usage.get(category_raw)
                if isinstance(bucket, dict):
                    category_usage = bucket.get(effective_value, 0)
            entry.update(
                {
                    "value": effective_value,
                    "category": category_raw,
                    "category_label": effective_label,
                    "equipment_type": equipment_type,
                    "equipment_vendor": equipment_vendor,
                    "equipment_model": equipment_model,
                    "equipment_status": equipment_status,
                    "extra": extra_payload,
                    "usage_count": category_usage,
                }
            )
        elif slug == "iiko_server":
            server_name = (extra_payload.get("server_name") or "").strip()
            if extra_payload is None:
                extra_payload = {}
            extra_payload["server_name"] = server_name
            entry.update(
                {
                    "server_name": server_name,
                    "extra": extra_payload,
                }
            )
        elif slug == "partner_contact":
            normalized_extra = normalize_partner_contact_extra(extra_payload, value)
            contact_type = normalized_extra.get("contact_type", "partner")
            phones = normalized_extra.get("phones", [])
            emails = normalized_extra.get("emails", [])
            served_id = normalized_extra.get("served_legal_entity_id")
            served_name = normalized_extra.get("served_legal_entity_name") or ""
            internal_name = normalized_extra.get("internal_name", "")
            served_entities = normalized_extra.get("served_legal_entities", [])
            served_ids = normalized_extra.get("served_legal_entity_ids", [])
            served_names = normalized_extra.get("served_legal_entity_names", [])
            entry.update(
                {
                    "contact_type": contact_type,
                    "phones": phones,
                    "emails": emails,
                    "served_legal_entity_id": served_id,
                    "served_legal_entity_name": served_name,
                    "internal_name": internal_name,
                    "served_legal_entities": served_entities,
                    "served_legal_entity_ids": served_ids,
                    "served_legal_entity_names": served_names,
                    "extra": normalized_extra,
                }
            )
        grouped[slug].append(entry)
    return grouped

def ensure_remote_access_defaults(conn) -> None:
    """Ensure that the default remote access entries exist."""
    try:
        rows = conn.execute(
            "SELECT value FROM settings_parameters WHERE param_type = ?",
            ("remote_access",),
        ).fetchall()
    except sqlite3.Error:
        return

    existing = {
        (row["value"] or "").strip().lower()
        for row in rows
        if row and (row["value"] or "").strip()
    }

    for value in REMOTE_ACCESS_DEFAULTS:
        normalized = (value or "").strip()
        if not normalized:
            continue
        if normalized.lower() in existing:
            continue
        conn.execute(
            """
            INSERT INTO settings_parameters(param_type, value, state, is_deleted, extra_json)
            VALUES (?, ?, ?, 0, NULL)
            """,
            ("remote_access", normalized, PARAMETER_ALLOWED_STATES[0]),
        )
        existing.add(normalized.lower())

@app.route("/api/settings/parameters", methods=["GET"])
@login_required
def api_get_parameters():
    ensure_settings_parameters_schema()
    conn = get_db()
    try:
        data = _fetch_parameters_grouped(conn, include_deleted=True)
        return jsonify(data)
    finally:
        conn.close()

@app.route("/api/settings/it-connection-categories", methods=["POST"])
@login_required
def api_create_it_connection_category():
    payload = request.json or {}
    label = (payload.get("label") or "").strip()
    requested_key = (payload.get("key") or "").strip()
    if not label:
        return jsonify({"success": False, "error": "Название категории обязательно"}), 400

    settings = load_settings()
    custom_categories = get_custom_it_connection_categories(settings)
    categories = get_it_connection_categories(settings)
    existing_keys = set(categories.keys())
    normalized_label = label.lower()
    for key, existing_label in categories.items():
        if existing_label and existing_label.strip().lower() == normalized_label:
            return (
                jsonify({"success": False, "error": "Такая категория уже существует"}),
                409,
            )

    if requested_key:
        category_key = requested_key
        if category_key in existing_keys:
            return (
                jsonify({"success": False, "error": "Идентификатор категории уже используется"}),
                409,
            )
    else:
        category_key = slugify_it_connection_category(label, existing_keys)

    custom_categories[category_key] = label
    settings = save_it_connection_categories(custom_categories, settings=settings)
    categories_updated = get_it_connection_categories(settings)

    return jsonify(
        {
            "success": True,
            "data": {
                "key": category_key,
                "label": label,
                "categories": categories_updated,
            },
        }
    )


@app.route("/api/settings/parameters", methods=["POST"])
@login_required
def api_create_parameter():
    ensure_settings_parameters_schema()
    payload = request.json or {}
    param_type = (payload.get("param_type") or "").strip()
    value = (payload.get("value") or "").strip()
    state = payload.get("state")

    if param_type not in PARAMETER_TYPES:
        return jsonify({"success": False, "error": "Неизвестный тип параметра"}), 400

    if not value:
        return jsonify({"success": False, "error": "Значение не может быть пустым"}), 400

    normalized_state = _normalize_parameter_state(param_type, state)
    extra_payload = {}
    if param_type == "it_connection":
        category = (payload.get("category") or "").strip()
        categories = get_it_connection_categories()
        if category not in categories:
            return (
                jsonify({"success": False, "error": "Неизвестная категория подключения"}),
                400,
            )
        sanitized = {
            "category": category,
            "category_label": categories.get(category, category),
            "equipment_type": (payload.get("equipment_type") or "").strip(),
            "equipment_vendor": (payload.get("equipment_vendor") or "").strip(),
            "equipment_model": (payload.get("equipment_model") or "").strip(),
            "equipment_status": (payload.get("equipment_status") or "").strip(),
        }
        category_field = DEFAULT_IT_CONNECTION_CATEGORY_FIELDS.get(category)
        if category_field:
            if not sanitized.get(category_field):
                sanitized[category_field] = value
        else:
            sanitized.update(
                {
                    "equipment_type": "",
                    "equipment_vendor": "",
                    "equipment_model": "",
                    "equipment_status": "",
                }
            )
        extra_payload = sanitized
    elif param_type == "iiko_server":
        server_name = (payload.get("server_name") or "").strip()
        if not server_name:
            return (
                jsonify({"success": False, "error": "Поле «Имя сервера» обязательно"}),
                400,
            )
        extra_payload = {"server_name": server_name}
    elif param_type == "legal_entity":
        extra_payload = {
            "inn": (payload.get("inn") or "").strip(),
            "manager_contacts": (payload.get("manager_contacts") or "").strip(),
        }
    elif param_type == "partner_contact":
        extra_payload = sanitize_partner_contact_extra(payload, value=value)

    extra_json = json.dumps(extra_payload, ensure_ascii=False) if extra_payload else None

    conn = get_db()
    try:
        try:
            cur = conn.execute(
                """
                INSERT INTO settings_parameters (param_type, value, state, is_deleted, extra_json)
                VALUES (?, ?, ?, 0, ?)
                """,
                (param_type, value, normalized_state, extra_json),
            )
            new_id = cur.lastrowid
            conn.commit()
        except sqlite3.IntegrityError:
            return (
                jsonify({"success": False, "error": "Такое значение уже существует"}),
                409,
            )

        data = _fetch_parameters_grouped(conn, include_deleted=True)
        return jsonify({"success": True, "id": new_id, "data": data})
    finally:
        conn.close()

@app.route("/api/settings/parameters/<int:param_id>", methods=["PATCH"])
@login_required
def api_update_parameter(param_id):
    ensure_settings_parameters_schema()
    payload = request.json or {}

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id, param_type, extra_json FROM settings_parameters WHERE id = ?",
            (param_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Параметр не найден"}), 404

        param_type = row["param_type"]
        updates = []
        params = []

        try:
            extra_payload = json.loads(row["extra_json"] or "{}")
        except Exception:
            extra_payload = {}
        if not isinstance(extra_payload, dict):
            extra_payload = {}

        new_value = None
        if "value" in payload:
            new_value = (payload.get("value") or "").strip()
            if not new_value:
                return jsonify({"success": False, "error": "Значение не может быть пустым"}), 400
            updates.append("value = ?")
            params.append(new_value)

        if "state" in payload:
            normalized_state = _normalize_parameter_state(param_type, payload.get("state"))
            updates.append("state = ?")
            params.append(normalized_state)

        if param_type == "legal_entity":
            inn_value = (extra_payload.get("inn") or "").strip()
            contacts_value = (extra_payload.get("manager_contacts") or "").strip()
            updated = False
            if "inn" in payload:
                inn_value = (payload.get("inn") or "").strip()
                updated = True
            if "manager_contacts" in payload:
                contacts_value = (payload.get("manager_contacts") or "").strip()
                updated = True
            if updated or not extra_payload:
                extra_payload["inn"] = inn_value
                extra_payload["manager_contacts"] = contacts_value
                updates.append("extra_json = ?")
                params.append(json.dumps(extra_payload, ensure_ascii=False))

        if param_type == "partner_contact":
            fields = {
                "contact_type",
                "phones",
                "emails",
                "contacts",
                "internal_name",
                "served_legal_entities",
                "served_legal_entity_ids",
                "served_legal_entity_names",
                "served_legal_entity_id",
                "served_legal_entity_name",
            }
            should_update_extra = bool(fields.intersection(payload.keys())) if isinstance(payload, dict) else False
            if new_value is not None or should_update_extra or not extra_payload:
                effective_value = new_value if new_value is not None else row["value"]
                sanitized = sanitize_partner_contact_extra(
                    payload if should_update_extra else {}, existing=extra_payload, value=effective_value
                )
                updates.append("extra_json = ?")
                params.append(json.dumps(sanitized, ensure_ascii=False))

        if param_type == "iiko_server":
            existing_name = (extra_payload.get("server_name") or "").strip()
            new_name = existing_name
            should_update_extra = False
            if "server_name" in payload:
                new_name = (payload.get("server_name") or "").strip()
                if not new_name:
                    return (
                        jsonify({"success": False, "error": "Поле «Имя сервера» обязательно"}),
                        400,
                    )
                should_update_extra = True
            if "value" in payload and not new_name:
                return (
                    jsonify({"success": False, "error": "Поле «Имя сервера» обязательно"}),
                    400,
                )
            if should_update_extra or (existing_name and new_name != existing_name) or (not existing_name and new_name):
                extra_payload["server_name"] = new_name
                updates.append("extra_json = ?")
                params.append(json.dumps(extra_payload, ensure_ascii=False))

        if param_type == "it_connection":
            categories = get_it_connection_categories()
            sanitized = {
                "category": (extra_payload.get("category") or "").strip(),
                "category_label": (extra_payload.get("category_label") or "").strip(),
                "equipment_type": (extra_payload.get("equipment_type") or "").strip(),
                "equipment_vendor": (extra_payload.get("equipment_vendor") or "").strip(),
                "equipment_model": (extra_payload.get("equipment_model") or "").strip(),
                "equipment_status": (extra_payload.get("equipment_status") or "").strip(),
            }
            current_category = sanitized["category"]
            if not current_category:
                if sanitized["equipment_type"]:
                    current_category = "equipment_type"
                elif sanitized["equipment_vendor"]:
                    current_category = "equipment_vendor"
                elif sanitized["equipment_model"]:
                    current_category = "equipment_model"
                elif sanitized["equipment_status"]:
                    current_category = "equipment_status"
                else:
                    current_category = "equipment_type"
            if current_category not in categories and sanitized.get("category_label"):
                categories = get_it_connection_categories(include={current_category: sanitized["category_label"]})
            sanitized["category"] = current_category

            if "category" in payload:
                new_category = (payload.get("category") or "").strip()
                if new_category not in categories:
                    return (
                        jsonify({"success": False, "error": "Неизвестная категория подключения"}),
                        400,
                    )
                current_category = new_category
                sanitized["category"] = new_category
                sanitized["category_label"] = categories.get(new_category, new_category)
            else:
                sanitized["category_label"] = sanitized.get("category_label") or categories.get(
                    current_category, current_category
                )

            for field in ("equipment_type", "equipment_vendor", "equipment_model", "equipment_status"):
                if field in payload:
                    sanitized[field] = (payload.get(field) or "").strip()

            category_field = DEFAULT_IT_CONNECTION_CATEGORY_FIELDS.get(current_category)
            if new_value is not None and category_field:
                sanitized[category_field] = new_value

            if not category_field:
                sanitized.update(
                    {
                        "equipment_type": "",
                        "equipment_vendor": "",
                        "equipment_model": "",
                        "equipment_status": "",
                    }
                )

            extra_payload = sanitized
            updates.append("extra_json = ?")
            params.append(json.dumps(extra_payload, ensure_ascii=False))

        if "is_deleted" in payload:
            is_deleted = 1 if bool(payload.get("is_deleted")) else 0
            updates.append("is_deleted = ?")
            params.append(is_deleted)
            if is_deleted:
                updates.append("deleted_at = datetime('now')")
            else:
                updates.append("deleted_at = NULL")

        if not updates:
            return jsonify({"success": False, "error": "Нет данных для обновления"}), 400

        query = f"UPDATE settings_parameters SET {', '.join(updates)} WHERE id = ?"
        params.append(param_id)

        try:
            conn.execute(query, params)
            conn.commit()
        except sqlite3.IntegrityError:
            return (
                jsonify({"success": False, "error": "Такое значение уже существует"}),
                409,
            )

        data = _fetch_parameters_grouped(conn, include_deleted=True)
        return jsonify({"success": True, "data": data})
    finally:
        conn.close()


@app.route("/api/settings/parameters/<int:param_id>", methods=["DELETE"])
@login_required
def api_delete_parameter(param_id):
    ensure_settings_parameters_schema()
    conn = get_db()
    try:
        cur = conn.execute(
            """
            UPDATE settings_parameters
            SET is_deleted = 1, deleted_at = datetime('now')
            WHERE id = ?
            """,
            (param_id,),
        )
        if cur.rowcount == 0:
            return jsonify({"success": False, "error": "Параметр не найден"}), 404
        conn.commit()
        data = _fetch_parameters_grouped(conn, include_deleted=True)
        return jsonify({"success": True, "data": data})
    finally:
        conn.close()

@app.route("/api/settings/it-equipment", methods=["GET"])
@login_required
def api_list_it_equipment():
    ensure_it_equipment_catalog_schema()
    conn = get_db()
    try:
        items = _fetch_it_equipment_catalog(conn)
        return jsonify({"success": True, "items": items})
    finally:
        conn.close()


@app.route("/api/settings/it-equipment", methods=["POST"])
@login_required
def api_create_it_equipment():
    ensure_it_equipment_catalog_schema()
    payload = request.json or {}
    equipment_type = (payload.get("equipment_type") or "").strip()
    equipment_vendor = (payload.get("equipment_vendor") or "").strip()
    equipment_model = (payload.get("equipment_model") or "").strip()
    photo_url = (payload.get("photo_url") or payload.get("photo") or "").strip()
    serial_number = (payload.get("serial_number") or "").strip()
    accessories = (payload.get("accessories") or payload.get("additional_equipment") or "").strip()

    if not equipment_type:
        return jsonify({"success": False, "error": "Поле «Тип оборудования» обязательно"}), 400
    if not equipment_vendor:
        return jsonify({"success": False, "error": "Поле «Производитель оборудования» обязательно"}), 400
    if not equipment_model:
        return jsonify({"success": False, "error": "Поле «Модель оборудования» обязательно"}), 400

    conn = get_db()
    try:
        cur = conn.execute(
            """
            INSERT INTO it_equipment_catalog (
                equipment_type, equipment_vendor, equipment_model, photo_url, serial_number, accessories, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
            """,
            (equipment_type, equipment_vendor, equipment_model, photo_url, serial_number, accessories),
        )
        new_id = cur.lastrowid
        row = conn.execute(
            """
            SELECT id, equipment_type, equipment_vendor, equipment_model, photo_url, serial_number, accessories
            FROM it_equipment_catalog
            WHERE id = ?
            """,
            (new_id,),
        ).fetchone()
        item = _serialize_it_equipment_row(row) if row else None
        items = _fetch_it_equipment_catalog(conn)
        return jsonify({"success": True, "item": item, "items": items, "id": new_id})
    finally:
        conn.close()


@app.route("/api/settings/it-equipment/<int:item_id>", methods=["PATCH"])
@login_required
def api_update_it_equipment(item_id):
    ensure_it_equipment_catalog_schema()
    payload = request.json or {}
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT id FROM it_equipment_catalog WHERE id = ?
            """,
            (item_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Оборудование не найдено"}), 404

        updates = []
        params = []

        if "equipment_type" in payload:
            equipment_type = (payload.get("equipment_type") or "").strip()
            if not equipment_type:
                return jsonify({"success": False, "error": "Поле «Тип оборудования» обязательно"}), 400
            updates.append("equipment_type = ?")
            params.append(equipment_type)

        if "equipment_vendor" in payload:
            equipment_vendor = (payload.get("equipment_vendor") or "").strip()
            if not equipment_vendor:
                return (
                    jsonify({"success": False, "error": "Поле «Производитель оборудования» обязательно"}),
                    400,
                )
            updates.append("equipment_vendor = ?")
            params.append(equipment_vendor)

        if "equipment_model" in payload:
            equipment_model = (payload.get("equipment_model") or "").strip()
            if not equipment_model:
                return jsonify({"success": False, "error": "Поле «Модель оборудования» обязательно"}), 400
            updates.append("equipment_model = ?")
            params.append(equipment_model)

        if "photo_url" in payload or "photo" in payload:
            photo_url = (payload.get("photo_url") or payload.get("photo") or "").strip()
            updates.append("photo_url = ?")
            params.append(photo_url)

        if "serial_number" in payload:
            serial_number = (payload.get("serial_number") or "").strip()
            updates.append("serial_number = ?")
            params.append(serial_number)

        if "accessories" in payload or "additional_equipment" in payload:
            accessories = (payload.get("accessories") or payload.get("additional_equipment") or "").strip()
            updates.append("accessories = ?")
            params.append(accessories)

        if not updates:
            return jsonify({"success": False, "error": "Нет данных для обновления"}), 400

        updates.append("updated_at = datetime('now')")
        query = f"UPDATE it_equipment_catalog SET {', '.join(updates)} WHERE id = ?"
        params.append(item_id)
        conn.execute(query, params)

        updated_row = conn.execute(
            """
            SELECT id, equipment_type, equipment_vendor, equipment_model, photo_url, serial_number, accessories
            FROM it_equipment_catalog
            WHERE id = ?
            """,
            (item_id,),
        ).fetchone()
        item = _serialize_it_equipment_row(updated_row) if updated_row else None
        items = _fetch_it_equipment_catalog(conn)
        return jsonify({"success": True, "item": item, "items": items})
    finally:
        conn.close()


@app.route("/api/settings/it-equipment/<int:item_id>", methods=["DELETE"])
@login_required
def api_delete_it_equipment(item_id):
    ensure_it_equipment_catalog_schema()
    conn = get_db()
    try:
        cur = conn.execute("DELETE FROM it_equipment_catalog WHERE id = ?", (item_id,))
        if cur.rowcount == 0:
            return jsonify({"success": False, "error": "Оборудование не найдено"}), 404
        items = _fetch_it_equipment_catalog(conn)
        return jsonify({"success": True, "items": items})
    finally:
        conn.close()

@app.route("/object_passports")
@login_required
@page_access_required("object_passports")
def object_passports_page():
    settings = load_settings()
    network_profiles = settings.get("network_profiles", []) if isinstance(settings, dict) else []
    contract_numbers = sorted(
        {
            (profile.get("contract_number") or "").strip()
            for profile in network_profiles
            if isinstance(profile, dict) and (profile.get("contract_number") or "").strip()
        },
        key=lambda value: value.lower(),
    )
    return render_template(
        "object_passports.html",
        parameter_values=_parameter_values_for_passports(),
        statuses=list(PASSPORT_STATUSES),
        cities=_city_options(),
        contract_numbers=contract_numbers,
    )


@app.route("/object_passports/new")
@login_required
def object_passport_new_page():
    return _render_passport_template(_blank_passport_detail(), True)


@app.route("/object_passports/<int:passport_id>")
@login_required
def object_passport_detail_page(passport_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT * FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not row:
            abort(404)
        detail = _serialize_passport_detail(conn, row)
    finally:
        conn.close()
    return _render_passport_template(detail, False)


@app.route("/api/object_passports", methods=["GET"])
@login_required_api
def api_object_passports_list():
    rows = _fetch_passport_rows(request.args)
    items = [_serialize_passport_row(row) for row in rows]
    return jsonify({"items": items})


@app.route("/api/object_passports", methods=["POST"])
@login_required_api
def api_object_passports_create():
    record, _, error = _prepare_passport_record(request.json or {})
    if error:
        return jsonify({"success": False, "error": error}), 400

    conn = get_passport_db()
    try:
        cur = conn.execute(
            """
            INSERT INTO object_passports (
                department, business, partner_type, country, legal_entity,
                city, location_address, status, start_date, end_date, suspension_date,
                resume_date, schedule_json,
                network, network_provider, network_restaurant_id, network_contract_number,
                network_legal_entity, network_support_phone, network_speed,
                network_tunnel,
                network_connection_params,
                it_connection_type, it_connection_id, it_connection_password,
                it_object_phone, it_manager_name, it_manager_phone,
                it_iiko_server,
                status_task_id
            )
            VALUES (
                :department, :business, :partner_type, :country, :legal_entity,
                :city, :location_address, :status, :start_date, :end_date, :suspension_date,
                :resume_date, :schedule_json,
                :network, :network_provider, :network_restaurant_id, :network_contract_number,
                :network_legal_entity, :network_support_phone, :network_speed,
                :network_tunnel,
                :network_connection_params,
                :it_connection_type, :it_connection_id, :it_connection_password,
                :it_object_phone, :it_manager_name, :it_manager_phone,
                :it_iiko_server,
                :status_task_id
            )
            """,
            record,
        )
        passport_id = cur.lastrowid
        _sync_status_period(
            conn,
            passport_id,
            record.get("status"),
            record.get("suspension_date"),
            record.get("resume_date"),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        detail = _serialize_passport_detail(conn, row)
    except sqlite3.IntegrityError:
        conn.rollback()
        return (
            jsonify({"success": False, "error": "Паспорт для этого департамента уже существует"}),
            409,
        )
    finally:
        conn.close()

    return jsonify({"success": True, "id": passport_id, "passport": detail})


@app.route("/api/object_passports/<int:passport_id>", methods=["PUT"])
@login_required_api
def api_object_passports_update(passport_id):
    record, _, error = _prepare_passport_record(request.json or {})
    if error:
        return jsonify({"success": False, "error": error}), 400

    conn = get_passport_db()
    try:
        existing = conn.execute(
            "SELECT id FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404

        conn.execute(
            """
            UPDATE object_passports
            SET department = :department,
                business = :business,
                partner_type = :partner_type,
                country = :country,
                legal_entity = :legal_entity,
                city = :city,
                location_address = :location_address,
                status = :status,
                start_date = :start_date,
                end_date = :end_date,
                schedule_json = :schedule_json,
                network = :network,
                network_provider = :network_provider,
                network_restaurant_id = :network_restaurant_id,
                network_contract_number = :network_contract_number,
                network_legal_entity = :network_legal_entity,
                network_support_phone = :network_support_phone,
                network_speed = :network_speed,
                network_tunnel = :network_tunnel,
                network_connection_params = :network_connection_params,
                it_connection_type = :it_connection_type,
                it_connection_id = :it_connection_id,
                it_connection_password = :it_connection_password,
                it_object_phone = :it_object_phone,
                it_manager_name = :it_manager_name,
                it_manager_phone = :it_manager_phone,
                it_iiko_server = :it_iiko_server,
                suspension_date = :suspension_date,
                resume_date = :resume_date,
                status_task_id = :status_task_id,
                updated_at = datetime('now')
            WHERE id = :id
            """,
            {**record, "id": passport_id},
        )
        _sync_status_period(
            conn,
            passport_id,
            record.get("status"),
            record.get("suspension_date"),
            record.get("resume_date"),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        detail = _serialize_passport_detail(conn, row)
    except sqlite3.IntegrityError:
        conn.rollback()
        return (
            jsonify({"success": False, "error": "Паспорт с таким департаментом уже существует"}),
            409,
        )
    finally:
        conn.close()

    return jsonify({"success": True, "passport": detail})


@app.route("/api/object_passports/<int:passport_id>", methods=["GET"])
@login_required_api
def api_object_passports_get(passport_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT * FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404
        detail = _serialize_passport_detail(conn, row)
    finally:
        conn.close()
    return jsonify({"success": True, "passport": detail})


@app.route("/api/object_passports/<int:passport_id>/cases", methods=["GET"])
@login_required_api
def api_object_passport_cases(passport_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT department FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404
        department = row["department"]
    finally:
        conn.close()

    total_minutes = _calculate_department_case_minutes(department)
    return jsonify(
        {
            "success": True,
            "items": _fetch_cases_for_department(department),
            "total_minutes": total_minutes,
            "total_display": format_time_duration(total_minutes),
        }
    )

@app.route("/api/object_passports/<int:passport_id>/tasks", methods=["GET"])
@login_required_api
def api_object_passport_tasks(passport_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT department FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404
        department = row["department"]
    finally:
        conn.close()

    search = request.args.get("search") if request.args else None
    limit = request.args.get("limit") if request.args else None
    items = _fetch_tasks_for_department(department, search=search, limit=limit)
    total_minutes = _calculate_department_task_minutes(department)
    return jsonify(
        {
            "success": True,
            "items": items,
            "total_minutes": total_minutes,
            "total_display": format_time_duration(total_minutes),
        }
    )


@app.route("/object_passports/export")
@login_required
def object_passports_export():
    try:
        import pandas as pd
    except ImportError as exc:
        logging.exception("Failed to import pandas for object passport export")
        abort(
            500,
            description=(
                "Не удалось сформировать Excel-файл: отсутствует библиотека pandas "
                "или её зависимости. Установите зависимости из requirements.txt"
            ),
        )

    rows = _fetch_passport_rows(request.args)
    dataset = []
    for row in rows:
        schedule = _load_schedule(row["schedule_json"])
        dataset.append(
            {
                "Бизнес": row["business"] or "",
                "Тип партнёра": row["partner_type"] or "",
                "Страна": row["country"] or "",
                "ЮЛ": row["legal_entity"] or "",
                "Город": row["city"] or "",
                "Департамент": row["department"] or "",
                "Статус": row["status"] or "",
                "Сеть": row["network"] or "",
                "Провайдер": row["network_provider"] or "",
                "ID ресторана": row["network_restaurant_id"] or "",
                "Номер договора": row["network_contract_number"] or "",
                "ЮЛ (сеть)": row["network_legal_entity"] or "",
                "Телефон ТП провайдера": row["network_support_phone"] or "",
                "Скорость подключения": row["network_speed"] or "",
                "Параметры подключения": row["network_connection_params"] or "",
                "Дата запуска": row["start_date"] or "",
                "Дата закрытия": row["end_date"] or "",
                "Дата приостановки": row["suspension_date"] or "",
                "Дата разморозки": row["resume_date"] or "",
                "ID задачи статуса": row["status_task_id"] or "",
                "Общее время работы": _format_total_time(row["start_date"], row["end_date"]),
                "Расписание": _format_schedule(schedule),
            }
        )

    df = pd.DataFrame(dataset)
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Паспорта")
    except ImportError:
        logging.exception("Failed to import Excel writer dependency")
        abort(
            500,
            description=(
                "Не удалось сформировать Excel-файл: отсутствует библиотека openpyxl. "
                "Установите зависимости из requirements.txt"
            ),
        )
    output.seek(0)
    filename = f"passports_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/object_passports/<int:passport_id>/photos", methods=["POST"])
@login_required_api
def api_object_passport_add_photo(passport_id):
    conn = get_passport_db()
    try:
        exists = conn.execute(
            "SELECT id FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not exists:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404
    finally:
        conn.close()

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "error": "Необходимо выбрать файл"}), 400

    category = (request.form.get("category") or "archive").strip().lower()
    if category not in ("title", "archive"):
        return jsonify({"success": False, "error": "Недопустимый тип фото"}), 400

    caption = (request.form.get("caption") or "").strip()
    filename = secure_filename(file.filename)
    if not filename:
        return jsonify({"success": False, "error": "Недопустимое имя файла"}), 400

    subdir = os.path.join(OBJECT_PASSPORT_UPLOADS_DIR, f"passport_{passport_id}")
    os.makedirs(subdir, exist_ok=True)
    unique_name = f"{int(time.time())}_{filename}"
    file_path = os.path.join(subdir, unique_name)
    file.save(file_path)
    relative_path = os.path.relpath(file_path, OBJECT_PASSPORT_UPLOADS_DIR).replace("\\", "/")

    with get_passport_db() as conn:
        cur = conn.execute(
            """
            INSERT INTO object_passport_photos (passport_id, category, caption, filename)
            VALUES (?, ?, ?, ?)
            """,
            (passport_id, category, caption, relative_path),
        )
        photo_id = cur.lastrowid
        if category == "title":
            _ensure_single_title_photo(conn, passport_id, photo_id)
        conn.commit()
        photos = _fetch_passport_photos(conn, passport_id)

    return jsonify({"success": True, "photos": photos})

@app.route("/api/object_passports/<int:passport_id>/network_files", methods=["POST"])
@login_required_api
def api_object_passport_add_network_file(passport_id):
    conn = get_passport_db()
    try:
        exists = conn.execute(
            "SELECT id FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not exists:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404
    finally:
        conn.close()

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "error": "Необходимо выбрать файл"}), 400

    original_name = file.filename
    safe_name = secure_filename(original_name)
    if not safe_name:
        return jsonify({"success": False, "error": "Недопустимое имя файла"}), 400

    subdir = os.path.join(OBJECT_PASSPORT_UPLOADS_DIR, f"passport_{passport_id}", "network")
    os.makedirs(subdir, exist_ok=True)
    unique_name = f"{int(time.time())}_{uuid4().hex}_{safe_name}"
    file_path = os.path.join(subdir, unique_name)
    file.save(file_path)
    file_size = os.path.getsize(file_path)
    relative_path = os.path.relpath(file_path, OBJECT_PASSPORT_UPLOADS_DIR).replace("\\", "/")

    with get_passport_db() as conn:
        conn.execute(
            """
            INSERT INTO object_passport_network_files (
                passport_id, original_name, filename, content_type, file_size
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (passport_id, original_name, relative_path, file.mimetype, file_size),
        )
        conn.commit()
        files = _fetch_network_files(conn, passport_id)

    return jsonify({"success": True, "files": files})


@app.route("/api/object_passports/network_files/<int:file_id>", methods=["DELETE"])
@login_required_api
def api_object_passport_delete_network_file(file_id):
    with get_passport_db() as conn:
        row = conn.execute(
            "SELECT passport_id, filename FROM object_passport_network_files WHERE id = ?",
            (file_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Файл не найден"}), 404

        conn.execute(
            "DELETE FROM object_passport_network_files WHERE id = ?",
            (file_id,),
        )
        conn.commit()
        files = _fetch_network_files(conn, row["passport_id"])

    file_path = os.path.join(OBJECT_PASSPORT_UPLOADS_DIR, row["filename"])
    try:
        if os.path.isfile(file_path):
            os.remove(file_path)
    except Exception:
        pass

    return jsonify({"success": True, "files": files})

@app.route("/api/object_passports/network_files/<int:file_id>/download")
@login_required
def api_object_passport_download_network_file(file_id):
    with get_passport_db() as conn:
        row = conn.execute(
            "SELECT filename, original_name FROM object_passport_network_files WHERE id = ?",
            (file_id,),
        ).fetchone()
        if not row:
            abort(404)

    safe_path = os.path.normpath(row["filename"]).replace("\\", "/")
    if safe_path.startswith("../") or safe_path.startswith("/"):
        abort(404)

    download_name = row["original_name"] or os.path.basename(safe_path)
    return send_from_directory(
        OBJECT_PASSPORT_UPLOADS_DIR,
        safe_path,
        as_attachment=True,
        download_name=download_name,
    )

@app.route("/api/object_passports/photos/<int:photo_id>", methods=["PATCH"])
@login_required_api
def api_object_passport_update_photo(photo_id):
    payload = request.json or {}
    updates = []
    params = []
    set_title = False

    if "caption" in payload:
        caption = (payload.get("caption") or "").strip()
        updates.append("caption = ?")
        params.append(caption)

    if "category" in payload:
        category = (payload.get("category") or "").strip().lower()
        if category not in ("title", "archive"):
            return jsonify({"success": False, "error": "Недопустимый тип фото"}), 400
        updates.append("category = ?")
        params.append(category)
        set_title = category == "title"

    if not updates:
        return jsonify({"success": False, "error": "Нет данных для обновления"}), 400

    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT passport_id FROM object_passport_photos WHERE id = ?",
            (photo_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Фото не найдено"}), 404

        params.append(photo_id)
        conn.execute(
            f"UPDATE object_passport_photos SET {', '.join(updates)}, updated_at = datetime('now') WHERE id = ?",
            params,
        )
        if set_title:
            _ensure_single_title_photo(conn, row["passport_id"], photo_id)
        conn.commit()
        photos = _fetch_passport_photos(conn, row["passport_id"])
    finally:
        conn.close()

    return jsonify({"success": True, "photos": photos})


@app.route("/api/object_passports/photos/<int:photo_id>", methods=["DELETE"])
@login_required_api
def api_object_passport_delete_photo(photo_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT passport_id, filename FROM object_passport_photos WHERE id = ?",
            (photo_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Фото не найдено"}), 404
        conn.execute("DELETE FROM object_passport_photos WHERE id = ?", (photo_id,))
        conn.commit()
        passport_id = row["passport_id"]
        filename = row["filename"]
        photos = _fetch_passport_photos(conn, passport_id)
    finally:
        conn.close()

    if filename:
        try:
            os.remove(os.path.join(OBJECT_PASSPORT_UPLOADS_DIR, filename))
        except FileNotFoundError:
            pass

    return jsonify({"success": True, "photos": photos})


@app.route("/api/object_passports/<int:passport_id>/equipment", methods=["POST"])
@login_required_api
def api_object_passport_add_equipment(passport_id):
    payload = request.json or {}

    def clean(field):
        return (payload.get(field) or "").strip() or None

    conn = get_passport_db()
    try:
        exists = conn.execute(
            "SELECT id FROM object_passports WHERE id = ?",
            (passport_id,),
        ).fetchone()
        if not exists:
            return jsonify({"success": False, "error": "Паспорт не найден"}), 404

        conn.execute(
            """
            INSERT INTO object_passport_equipment (
                passport_id,
                equipment_type,
                vendor,
                name,
                model,
                serial_number,
                status,
                ip_address,
                connection_type,
                connection_id,
                connection_password,
                description
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                passport_id,
                clean("equipment_type"),
                clean("vendor"),
                clean("name"),
                clean("model"),
                clean("serial_number"),
                clean("status"),
                clean("ip_address"),
                clean("connection_type"),
                clean("connection_id"),
                clean("connection_password"),
                clean("description"),
            ),
        )
        conn.commit()
        equipment = _fetch_passport_equipment(conn, passport_id)
    finally:
        conn.close()

    return jsonify({"success": True, "equipment": equipment})


@app.route("/api/object_passports/equipment/<int:equipment_id>", methods=["PATCH"])
@login_required_api
def api_object_passport_update_equipment(equipment_id):
    payload = request.json or {}
    updates = []
    params = []

    for field in [
        "equipment_type",
        "vendor",
        "name",
        "model",
        "serial_number",
        "status",
        "ip_address",
        "connection_type",
        "connection_id",
        "connection_password",
        "description",
    ]:
        if field in payload:
            value = (payload.get(field) or "").strip()
            updates.append(f"{field} = ?")
            params.append(value or None)

    if not updates:
        return jsonify({"success": False, "error": "Нет данных для обновления"}), 400

    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT passport_id FROM object_passport_equipment WHERE id = ?",
            (equipment_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Оборудование не найдено"}), 404

        params.append(equipment_id)
        conn.execute(
            f"UPDATE object_passport_equipment SET {', '.join(updates)}, updated_at = datetime('now') WHERE id = ?",
            params,
        )
        conn.commit()
        equipment = _fetch_passport_equipment(conn, row["passport_id"])
    finally:
        conn.close()

    return jsonify({"success": True, "equipment": equipment})


@app.route("/api/object_passports/equipment/<int:equipment_id>", methods=["DELETE"])
@login_required_api
def api_object_passport_delete_equipment(equipment_id):
    photos = []
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT passport_id FROM object_passport_equipment WHERE id = ?",
            (equipment_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Оборудование не найдено"}), 404
        passport_id = row["passport_id"]
        photos = conn.execute(
            "SELECT filename FROM object_passport_equipment_photos WHERE equipment_id = ?",
            (equipment_id,),
        ).fetchall()
        conn.execute("DELETE FROM object_passport_equipment WHERE id = ?", (equipment_id,))
        conn.commit()
    finally:
        conn.close()

    for photo in photos:
        filename = photo["filename"]
        if filename:
            try:
                os.remove(os.path.join(OBJECT_PASSPORT_UPLOADS_DIR, filename))
            except FileNotFoundError:
                pass

    with get_passport_db() as conn:
        equipment = _fetch_passport_equipment(conn, passport_id)

    return jsonify({"success": True, "equipment": equipment})


@app.route("/api/object_passports/equipment/<int:equipment_id>/photos", methods=["POST"])
@login_required_api
def api_object_passport_equipment_add_photo(equipment_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT passport_id FROM object_passport_equipment WHERE id = ?",
            (equipment_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Оборудование не найдено"}), 404
        passport_id = row["passport_id"]
    finally:
        conn.close()

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "error": "Необходимо выбрать файл"}), 400

    caption = (request.form.get("caption") or "").strip()
    filename = secure_filename(file.filename)
    if not filename:
        return jsonify({"success": False, "error": "Недопустимое имя файла"}), 400

    subdir = os.path.join(
        OBJECT_PASSPORT_UPLOADS_DIR,
        f"passport_{passport_id}",
        f"equipment_{equipment_id}",
    )
    os.makedirs(subdir, exist_ok=True)
    unique_name = f"{int(time.time())}_{filename}"
    file_path = os.path.join(subdir, unique_name)
    file.save(file_path)
    relative_path = os.path.relpath(file_path, OBJECT_PASSPORT_UPLOADS_DIR).replace("\\", "/")

    with get_passport_db() as conn:
        conn.execute(
            """
            INSERT INTO object_passport_equipment_photos (equipment_id, caption, filename)
            VALUES (?, ?, ?)
            """,
            (equipment_id, caption, relative_path),
        )
        conn.commit()
        equipment = _fetch_passport_equipment(conn, passport_id)

    return jsonify({"success": True, "equipment": equipment})


@app.route("/api/object_passports/equipment/photos/<int:photo_id>", methods=["PATCH"])
@login_required_api
def api_object_passport_equipment_update_photo(photo_id):
    payload = request.json or {}
    if "caption" not in payload:
        return jsonify({"success": False, "error": "Нет данных для обновления"}), 400

    caption = (payload.get("caption") or "").strip()

    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT equipment_id FROM object_passport_equipment_photos WHERE id = ?",
            (photo_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Фото не найдено"}), 404
        equipment_row = conn.execute(
            "SELECT passport_id FROM object_passport_equipment WHERE id = ?",
            (row["equipment_id"],),
        ).fetchone()
        if not equipment_row:
            return jsonify({"success": False, "error": "Оборудование не найдено"}), 404

        conn.execute(
            "UPDATE object_passport_equipment_photos SET caption = ? WHERE id = ?",
            (caption, photo_id),
        )
        conn.commit()
        equipment = _fetch_passport_equipment(conn, equipment_row["passport_id"])
    finally:
        conn.close()

    return jsonify({"success": True, "equipment": equipment})


@app.route("/api/object_passports/equipment/photos/<int:photo_id>", methods=["DELETE"])
@login_required_api
def api_object_passport_equipment_delete_photo(photo_id):
    conn = get_passport_db()
    try:
        row = conn.execute(
            "SELECT equipment_id, filename FROM object_passport_equipment_photos WHERE id = ?",
            (photo_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Фото не найдено"}), 404
        equipment_row = conn.execute(
            "SELECT passport_id FROM object_passport_equipment WHERE id = ?",
            (row["equipment_id"],),
        ).fetchone()
        passport_id = equipment_row["passport_id"] if equipment_row else None
        conn.execute("DELETE FROM object_passport_equipment_photos WHERE id = ?", (photo_id,))
        conn.commit()
    finally:
        conn.close()

    if row["filename"]:
        try:
            os.remove(os.path.join(OBJECT_PASSPORT_UPLOADS_DIR, row["filename"]))
        except FileNotFoundError:
            pass

    equipment = []
    if passport_id:
        with get_passport_db() as conn:
            equipment = _fetch_passport_equipment(conn, passport_id)

    return jsonify({"success": True, "equipment": equipment})

# === API для управления пользователями и ролями ===


def _resolve_role_assignment(conn, role_id: int | None, role_name: str | None):
    if role_id:
        row = conn.execute(
            "SELECT id, name FROM roles WHERE id = ?",
            (role_id,),
        ).fetchone()
        if not row:
            abort(400, description="Указанная роль не найдена")
        return row
    if role_name:
        row = conn.execute(
            "SELECT id, name FROM roles WHERE LOWER(name) = LOWER(?)",
            (role_name,),
        ).fetchone()
        if not row:
            abort(400, description="Указанная роль не найдена")
        return row
    row = conn.execute(
        "SELECT id, name FROM roles WHERE name = 'user'"
    ).fetchone()
    if not row:
        raise RuntimeError("Роль по умолчанию не найдена")
    return row


def _parse_user_phones(raw) -> list[dict]:
    if not raw:
        return []
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            return []
    if not isinstance(raw, list):
        return []
    result: list[dict] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        number = (item.get("value") or item.get("number") or "").strip()
        if not number:
            continue
        phone_type = (item.get("type") or item.get("label") or "").strip()
        result.append({"type": phone_type, "value": number})
    return result


def _normalize_phone_payload(raw) -> list[dict]:
    if not raw:
        return []
    result: list[dict] = []
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            raw = []
    if not isinstance(raw, list):
        return []
    for item in raw:
        if not isinstance(item, dict):
            continue
        number = str(item.get("value") or item.get("number") or "").strip()
        if not number:
            continue
        phone_type = str(item.get("type") or item.get("label") or "").strip()
        result.append({"type": phone_type, "value": number})
    return result


def _serialize_user_row(row: sqlite3.Row) -> dict:
    return {
        "id": _row_value(row, "id"),
        "username": _row_value(row, "username", ""),
        "full_name": _row_value(row, "full_name"),
        "role": _row_value(row, "role_name", ""),
        "role_id": _row_value(row, "role_id"),
        "photo": _row_value(row, "photo"),
        "registration_date": _row_value(row, "registration_date"),
        "birth_date": _row_value(row, "birth_date"),
        "email": _row_value(row, "email"),
        "department": _row_value(row, "department"),
        "phones": _parse_user_phones(_row_value(row, "phones")),
        "is_blocked": bool(_row_value(row, "is_blocked", 0)),
    }


def _fetch_user_summary(conn, user_id: int):
    return conn.execute(
        """
        SELECT
            u.id,
            u.username,
            u.full_name,
            u.role_id,
            COALESCE(r.name, u.role) AS role_name,
            u.photo,
            u.registration_date,
            u.birth_date,
            u.email,
            u.department,
            u.phones,
            u.is_blocked
        FROM users u
        LEFT JOIN roles r ON r.id = u.role_id
        WHERE u.id = ?
        """,
        (user_id,),
    ).fetchone()


@app.route("/api/users/photo-upload", methods=["POST"])
@login_required_api
def api_users_upload_photo():
    file_storage = request.files.get("photo")
    if not file_storage or not (file_storage.filename or "").strip():
        return jsonify({"success": False, "error": "Выберите файл с изображением."}), 400

    filename = secure_filename(file_storage.filename or "")
    if not filename:
        return jsonify({"success": False, "error": "Недопустимое имя файла."}), 400

    extension = os.path.splitext(filename)[1].lower()
    if extension not in ALLOWED_USER_PHOTO_EXTENSIONS:
        return jsonify({"success": False, "error": "Поддерживаются изображения PNG, JPG, GIF или WebP."}), 400

    mimetype = (file_storage.mimetype or "").lower()
    if mimetype and not mimetype.startswith("image/"):
        return jsonify({"success": False, "error": "Загрузите файл изображения."}), 400

    try:
        file_storage.stream.seek(0, os.SEEK_END)
        file_size = file_storage.stream.tell()
        file_storage.stream.seek(0)
    except (OSError, AttributeError, io.UnsupportedOperation):
        file_bytes = file_storage.read()
        file_size = len(file_bytes)
        file_storage.stream = io.BytesIO(file_bytes)
        file_storage.stream.seek(0)

    if file_size > MAX_USER_PHOTO_SIZE:
        return jsonify({"success": False, "error": "Размер файла не должен превышать 5 МБ."}), 400
    if file_size == 0:
        return jsonify({"success": False, "error": "Файл не может быть пустым."}), 400

    os.makedirs(USER_PHOTOS_DIR, exist_ok=True)
    unique_name = f"{int(time.time())}_{uuid4().hex}{extension}"
    save_path = os.path.join(USER_PHOTOS_DIR, unique_name)
    file_storage.save(save_path)

    file_url = url_for("static", filename=f"user_photos/{unique_name}")
    return jsonify({"success": True, "url": file_url, "filename": unique_name})


@app.route("/users")
@login_required
def get_users():
    if not (has_page_access("user_management") or has_page_access("settings")):
        abort(403)
    try:
        with get_users_db() as conn:
            rows = conn.execute(
                """
                SELECT
                    u.id,
                    u.username,
                    u.full_name,
                    u.role_id,
                    COALESCE(r.name, u.role) AS role_name,
                    u.photo,
                    u.registration_date,
                    u.birth_date,
                    u.email,
                    u.department,
                    u.phones,
                    u.is_blocked
                FROM users u
                LEFT JOIN roles r ON r.id = u.role_id
                ORDER BY LOWER(u.username)
                """
            ).fetchall()
        return jsonify([_serialize_user_row(row) for row in rows])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/users", methods=["POST"])
@login_required
def add_user():
    if not has_field_edit_permission("user.create"):
        abort(403)
    data = request.json or {}
    username = (data.get("username") or "").strip()
    full_name = (data.get("full_name") or "").strip() or None
    password = (data.get("password") or "").strip()
    role_id = data.get("role_id")
    role_name = (data.get("role") or "").strip() or None
    photo = (data.get("photo") or "").strip() or None
    birth_date = (data.get("birth_date") or "").strip() or None
    email = (data.get("email") or "").strip() or None
    department = (data.get("department") or "").strip() or None
    phones_payload = _normalize_phone_payload(data.get("phones"))
    registration_date = datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    phones_json = json.dumps(phones_payload, ensure_ascii=False) if phones_payload else None

    if not username or not password:
        return jsonify({"success": False, "error": "Имя пользователя и пароль не могут быть пустыми"}), 400

    try:
        with get_users_db() as conn:
            existing = conn.execute(
                "SELECT id FROM users WHERE LOWER(username) = LOWER(?)",
                (username,),
            ).fetchone()
            if existing:
                return jsonify({"success": False, "error": "Пользователь уже существует"}), 400

            role_row = _resolve_role_assignment(conn, role_id, role_name)
            cursor = conn.execute(
                """
                INSERT INTO users (
                    username,
                    full_name,
                    role_id,
                    role,
                    photo,
                    registration_date,
                    birth_date,
                    email,
                    department,
                    phones,
                    is_blocked
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    username,
                    full_name,
                    role_row["id"],
                    role_row["name"],
                    photo,
                    registration_date,
                    birth_date or None,
                    email or None,
                    department or None,
                    phones_json,
                    0,
                ),
            )
            new_user_id = cursor.lastrowid
            _set_user_password(conn, new_user_id, password)
            conn.commit()
            summary = _fetch_user_summary(conn, new_user_id)

        sync_org_structure_user_department(new_user_id, None, summary.get("department") if summary else None)
        return jsonify({"success": True, "user": _serialize_user_row(summary)})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/users/<int:user_id>", methods=["PUT", "PATCH"])
@login_required
def update_user(user_id):
    data = request.json or {}
    original_department: str | None = None
    with get_users_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Пользователь не найден"}), 404

        original_department = (row["department"] or "").strip() or None
        updates_made = False
        can_edit_profile = _can_edit_user_profile(user_id)

        if "username" in data:
            new_username = (data.get("username") or "").strip()
            if not new_username:
                return jsonify({"success": False, "error": "Имя пользователя не может быть пустым"}), 400
            if not has_field_edit_permission("user.username"):
                abort(403)
            if new_username.lower() != (row["username"] or "").lower():
                existing = conn.execute(
                    "SELECT id FROM users WHERE LOWER(username) = LOWER(?) AND id <> ?",
                    (new_username, user_id),
                ).fetchone()
                if existing:
                    return jsonify({"success": False, "error": "Пользователь с таким именем уже существует"}), 400
                conn.execute(
                    "UPDATE users SET username = ? WHERE id = ?",
                    (new_username, user_id),
                )
                updates_made = True

        if "role_id" in data or "role" in data:
            if not has_field_edit_permission("user.role"):
                abort(403)
            role_row = _resolve_role_assignment(
                conn,
                data.get("role_id"),
                (data.get("role") or "").strip() or None,
            )
            conn.execute(
                "UPDATE users SET role_id = ?, role = ? WHERE id = ?",
                (role_row["id"], role_row["name"], user_id),
            )
            updates_made = True

        if "password" in data:
            new_password = (data.get("password") or "").strip()
            if not new_password:
                return jsonify({"success": False, "error": "Пароль не может быть пустым"}), 400
            current_id = _current_user_id()
            if user_id != current_id and not has_field_edit_permission("user.password"):
                abort(403)
            _set_user_password(conn, user_id, new_password)
            updates_made = True

        if "is_blocked" in data:
            if user_id == _current_user_id():
                return jsonify({"success": False, "error": "Нельзя изменить блокировку своей учётной записи"}), 400
            if not has_field_edit_permission("user.block"):
                abort(403)
            blocked_value = 1 if bool(data.get("is_blocked")) else 0
            conn.execute("UPDATE users SET is_blocked = ? WHERE id = ?", (blocked_value, user_id))
            updates_made = True

        if "full_name" in data:
            if not can_edit_profile:
                abort(403)
            full_name = (data.get("full_name") or "").strip() or None
            conn.execute(
                "UPDATE users SET full_name = ? WHERE id = ?",
                (full_name, user_id),
            )
            updates_made = True

        if "photo" in data:
            if not can_edit_profile:
                abort(403)
            photo_value = (data.get("photo") or "").strip() or None
            conn.execute(
                "UPDATE users SET photo = ? WHERE id = ?",
                (photo_value, user_id),
            )
            updates_made = True

        if "birth_date" in data:
            if not can_edit_profile:
                abort(403)
            birth_date = (data.get("birth_date") or "").strip() or None
            conn.execute(
                "UPDATE users SET birth_date = ? WHERE id = ?",
                (birth_date, user_id),
            )
            updates_made = True

        if "email" in data:
            if not can_edit_profile:
                abort(403)
            email_value = (data.get("email") or "").strip() or None
            conn.execute(
                "UPDATE users SET email = ? WHERE id = ?",
                (email_value, user_id),
            )
            updates_made = True

        if "department" in data:
            if not can_edit_profile:
                abort(403)
            department = (data.get("department") or "").strip() or None
            conn.execute(
                "UPDATE users SET department = ? WHERE id = ?",
                (department, user_id),
            )
            updates_made = True

        if "phones" in data:
            if not can_edit_profile:
                abort(403)
            phones_payload = _normalize_phone_payload(data.get("phones"))
            phones_json = json.dumps(phones_payload, ensure_ascii=False) if phones_payload else None
            conn.execute(
                "UPDATE users SET phones = ? WHERE id = ?",
                (phones_json, user_id),
            )
            updates_made = True

        if updates_made:
            conn.commit()
        summary = _fetch_user_summary(conn, user_id)

    if not summary:
        return jsonify({"success": False, "error": "Пользователь не найден"}), 404

    if "department" in data:
        sync_org_structure_user_department(
            user_id,
            original_department,
            summary.get("department") if summary else None,
        )

    return jsonify({"success": True, "user": _serialize_user_row(summary), "updated": updates_made})


@app.route("/users/<int:user_id>", methods=["DELETE"])
@login_required
def delete_user(user_id):
    current_id = _current_user_id()
    if current_id == user_id:
        return jsonify({"success": False, "error": "Нельзя удалить свою учётную запись"}), 400
    if not has_field_edit_permission("user.delete"):
        abort(403)

    with get_users_db() as conn:
        row = conn.execute(
            "SELECT id, username, role, role_id FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Пользователь не найден"}), 404

        role_name = (row["role"] or "").lower()
        if role_name == "admin":
            remaining_admins = conn.execute(
                "SELECT COUNT(*) AS total FROM users WHERE LOWER(role) = 'admin' AND id <> ?",
                (user_id,),
            ).fetchone()["total"]
            if remaining_admins == 0:
                return jsonify({"success": False, "error": "Нельзя удалить последнего администратора"}), 400

        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()

    return jsonify({"success": True})


@app.route("/users/<int:user_id>/password", methods=["GET"])
@login_required
def get_user_password(user_id):
    current_id = _current_user_id()
    if user_id != current_id and not has_field_view_permission("user.password"):
        abort(403)

    row = _load_user_row(user_id=user_id)
    if not row:
        return jsonify({"success": False, "error": "Пользователь не найден"}), 404
    password_value = _row_value(row, "password", "") or ""
    if not password_value:
        return jsonify({"success": False, "error": "Пароль недоступен"}), 404
    return jsonify({"success": True, "password": password_value})


@app.route("/roles", methods=["GET"])
@login_required
def list_roles():
    if not (has_page_access("user_management") or has_page_access("settings")):
        abort(403)
    with get_users_db() as conn:
        rows = conn.execute(
            "SELECT id, name, description, permissions FROM roles ORDER BY LOWER(name)"
        ).fetchall()
    return jsonify([_serialize_role(row) for row in rows])


@app.route("/roles", methods=["POST"])
@login_required
def create_role():
    if not has_field_edit_permission("role.create"):
        abort(403)
    data = request.json or {}
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    permissions_payload = _prepare_permissions_payload(data.get("permissions"))

    if not name:
        return jsonify({"success": False, "error": "Название роли не может быть пустым"}), 400

    with get_users_db() as conn:
        existing = conn.execute(
            "SELECT id FROM roles WHERE LOWER(name) = LOWER(?)",
            (name,),
        ).fetchone()
        if existing:
            return jsonify({"success": False, "error": "Роль с таким названием уже существует"}), 400
        cursor = conn.execute(
            "INSERT INTO roles (name, description, permissions) VALUES (?, ?, ?)",
            (name, description, json.dumps(permissions_payload, ensure_ascii=False)),
        )
        conn.commit()
        new_role = conn.execute(
            "SELECT id, name, description, permissions FROM roles WHERE id = ?",
            (cursor.lastrowid,),
        ).fetchone()

    return jsonify({"success": True, "role": _serialize_role(new_role)})


@app.route("/roles/<int:role_id>", methods=["PUT", "PATCH"])
@login_required
def update_role(role_id):
    data = request.json or {}
    with get_users_db() as conn:
        row = conn.execute(
            "SELECT id, name, description, permissions FROM roles WHERE id = ?",
            (role_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Роль не найдена"}), 404

        role_name = _row_value(row, "name", "")
        is_admin_role = _is_builtin_role(role_name)
        updates = {}

        if "name" in data:
            if is_admin_role:
                return jsonify({"success": False, "error": "Нельзя переименовывать роль admin"}), 400
            if not has_field_edit_permission("role.name"):
                abort(403)
            new_name = (data.get("name") or "").strip()
            if not new_name:
                return jsonify({"success": False, "error": "Название роли не может быть пустым"}), 400
            exists = conn.execute(
                "SELECT id FROM roles WHERE LOWER(name) = LOWER(?) AND id <> ?",
                (new_name, role_id),
            ).fetchone()
            if exists:
                return jsonify({"success": False, "error": "Роль с таким названием уже существует"}), 400
            updates["name"] = new_name

        if "description" in data:
            if not has_field_edit_permission("role.description"):
                abort(403)
            updates["description"] = (data.get("description") or "").strip()

        if "permissions" in data:
            if is_admin_role:
                return jsonify({"success": False, "error": "Нельзя изменять права роли admin"}), 400
            permissions_data = data.get("permissions") or {}
            if isinstance(permissions_data, dict):
                if "pages" in permissions_data and not has_field_edit_permission("role.pages"):
                    abort(403)
                fields_part = permissions_data.get("fields") or {}
                if isinstance(fields_part, dict):
                    if "edit" in fields_part and not has_field_edit_permission("role.fields.edit"):
                        abort(403)
                    if "view" in fields_part and not has_field_edit_permission("role.fields.view"):
                        abort(403)
            permissions_payload = _prepare_permissions_payload(permissions_data)
            updates["permissions"] = json.dumps(permissions_payload, ensure_ascii=False)

        if updates:
            assignments = []
            params = []
            for column, value in updates.items():
                assignments.append(f"{column} = ?")
                params.append(value)
            params.append(role_id)
            conn.execute(
                f"UPDATE roles SET {', '.join(assignments)} WHERE id = ?",
                params,
            )
            conn.commit()

        updated_row = conn.execute(
            "SELECT id, name, description, permissions FROM roles WHERE id = ?",
            (role_id,),
        ).fetchone()

    return jsonify({"success": True, "role": _serialize_role(updated_row), "updated": bool(updates)})


@app.route("/roles/<int:role_id>", methods=["DELETE"])
@login_required
def delete_role(role_id):
    if not has_field_edit_permission("role.delete"):
        abort(403)
    with get_users_db() as conn:
        row = conn.execute(
            "SELECT id, name FROM roles WHERE id = ?",
            (role_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Роль не найдена"}), 404
        role_name = _row_value(row, "name", "")
        if _is_builtin_role(role_name):
            return jsonify({"success": False, "error": "Нельзя удалить системную роль"}), 400
        usage = conn.execute(
            "SELECT COUNT(*) AS total FROM users WHERE role_id = ?",
            (role_id,),
        ).fetchone()["total"]
        if usage:
            return jsonify({"success": False, "error": "Роль используется пользователями"}), 400
        conn.execute("DELETE FROM roles WHERE id = ?", (role_id,))
        conn.commit()

    return jsonify({"success": True})


@app.route("/auth/state", methods=["GET"])
@login_required
def get_auth_state():
    if not (has_page_access("user_management") or has_page_access("settings")):
        abort(403)
    with get_users_db() as conn:
        users_rows = conn.execute(
            """
            SELECT
                u.id,
                u.username,
                u.full_name,
                u.role_id,
                COALESCE(r.name, u.role) AS role_name,
                u.photo,
                u.registration_date,
                u.birth_date,
                u.email,
                u.department,
                u.phones,
                u.is_blocked
            FROM users u
            LEFT JOIN roles r ON r.id = u.role_id
            ORDER BY LOWER(u.username)
            """
        ).fetchall()
        role_rows = conn.execute(
            "SELECT id, name, description, permissions FROM roles ORDER BY LOWER(name)"
        ).fetchall()

    current_id = _current_user_id()
    can_view_password_globally = has_field_view_permission("user.password")
    can_edit_password_globally = has_field_edit_permission("user.password")
    can_edit_username = has_field_edit_permission("user.username")
    can_edit_role = has_field_edit_permission("user.role")
    can_delete_user = has_field_edit_permission("user.delete")
    can_block_user = has_field_edit_permission("user.block")

    users_payload = []
    for row in users_rows:
        user_id = row["id"]
        is_self = user_id == current_id
        role_name = (row["role_name"] or "").lower()
        users_payload.append(
            {
                **_serialize_user_row(row),
                "is_self": is_self,
                "role_is_admin": role_name == "admin",
                "can_view_password": is_self or can_view_password_globally,
                "can_edit_password": is_self or can_edit_password_globally,
                "can_edit_username": can_edit_username,
                "can_edit_role": can_edit_role,
                "can_delete": can_delete_user and not is_self and role_name != "admin",
                "can_block": can_block_user and not is_self,
            }
        )

    capabilities = {
        "fields": {
            "edit": {
                item["key"]: has_field_edit_permission(item["key"])
                for item in EDITABLE_FIELD_PERMISSIONS
            },
            "view": {
                item["key"]: has_field_view_permission(item["key"])
                for item in VIEWABLE_FIELD_PERMISSIONS
            },
        }
    }

    return jsonify(
        {
            "users": users_payload,
            "roles": [_serialize_role(row) for row in role_rows],
            "catalog": _get_permissions_catalog(),
            "capabilities": capabilities,
            "current_user_id": current_id,
            "org_structure": load_org_structure(),
        }
    )


@app.route("/auth/org-structure", methods=["POST"])
@login_required
def update_org_structure_state():
    if not (has_page_access("user_management") or has_page_access("settings")):
        abort(403)

    data = request.get_json(force=True, silent=True) or {}
    payload = data.get("org_structure") if isinstance(data, dict) else data
    sanitized = save_org_structure(payload)
    return jsonify({"success": True, "org_structure": sanitized})

@app.route("/api/channels/<int:channel_id>", methods=["GET"])
@login_required
def api_channels_get(channel_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM channels WHERE id = ?", (channel_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Канал не найден"}), 404
        d = _refresh_bot_identity_if_needed(conn, dict(row))
    # Попробуем распарсить JSON, но вернём как есть — фронт сам приведёт
    return jsonify({"success": True, "channel": d})

@app.route("/stats_data")
@login_required
def stats_data():
    conn = get_db()
    total     = conn.execute("SELECT COUNT(*) FROM tickets").fetchone()[0]
    resolved  = conn.execute("SELECT COUNT(*) FROM tickets WHERE status = 'resolved'").fetchone()[0]
    pending   = total - resolved
    by_channel= conn.execute("""
        SELECT c.channel_name AS name, COUNT(*) as total
        FROM tickets t
        LEFT JOIN channels c ON c.id = t.channel_id
        GROUP BY c.channel_name
        ORDER BY total DESC
    """).fetchall()
    conn.close()
    return jsonify({
        "total": total, 
        "resolved": resolved, 
        "pending": pending,
        "by_channel": [dict(r) for r in by_channel]
    })

# === история сообщений (совместимо со всеми Flask) ===
@app.route('/history', methods=['GET'])
@login_required
def history():
    ticket_id = request.args.get("ticket_id", type=str)
    if not ticket_id:
        return jsonify({"success": False, "error": "ticket_id обязателен"}), 400

    # channel_id опционален; фильтруем по нему только если он валидный (>0)
    ch_raw = request.args.get("channel_id", default=None)
    channel_id = None
    try:
        if ch_raw not in (None, "", "null", "undefined"):
            ch_val = int(ch_raw)
            if ch_val > 0:
                channel_id = ch_val
    except Exception:
        channel_id = None

    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        # Тащим все строки по тикету; если channel_id валиден — добавляем фильтр
        if channel_id is not None:
            cur.execute("""
                SELECT sender, message, timestamp, message_type, attachment,
                       tg_message_id, reply_to_tg_id, channel_id,
                       edited_at, deleted_at
                  FROM chat_history
                 WHERE ticket_id = ? AND channel_id = ?
                 ORDER BY
                  substr(timestamp,1,19) ASC,
                  COALESCE(tg_message_id, 0) ASC,
                  rowid ASC
            """, (ticket_id, channel_id))
        else:
            cur.execute("""
                SELECT sender, message, timestamp, message_type, attachment,
                       tg_message_id, reply_to_tg_id, channel_id,
                       edited_at, deleted_at
                  FROM chat_history
                 WHERE ticket_id = ?
                 ORDER BY
                  substr(timestamp,1,19) ASC,
                  COALESCE(tg_message_id, 0) ASC,
                  rowid ASC
            """, (ticket_id,))
        rows = cur.fetchall()
    finally:
        try:
            conn.close()
        except:
            pass

    # Построим карту (channel_id, tg_message_id) -> текст для превью
    by_tg = {}
    for r in rows:
        mid = r["tg_message_id"]
        if not mid:
            continue
        # текст для превью: сообщение или подпись к медиа; если пусто — ставить метку
        base = (r["message"] or "").strip()
        if not base:
            # можно подставить короткую метку по типу медиа
            mt = (r["message_type"] or "").lower()
            if mt and mt != "text":
                base = f"[{mt}]"
        # обрежем превью до 140 символов
        by_tg[(r["channel_id"], mid)] = (base[:140] + "…") if len(base) > 140 else base

    # Формируем ответ + reply_preview
    out = []
    for r in rows:
        preview = None
        if r["reply_to_tg_id"]:
            key_exact = (r["channel_id"], r["reply_to_tg_id"])
            preview = by_tg.get(key_exact)
            if preview is None:
                # на всякий случай попробуем без учёта channel_id
                for (ch, mid), txt in by_tg.items():
                    if mid == r["reply_to_tg_id"]:
                        preview = txt
                        break

        out.append({
            "sender":         r["sender"],
            "message":        r["message"],
            "timestamp":      r["timestamp"],
            "message_type":   r["message_type"],
            "attachment":     r["attachment"],
            "tg_message_id":  r["tg_message_id"],
            "reply_to_tg_id": r["reply_to_tg_id"],
            "reply_preview":  preview,
            "edited_at":      r["edited_at"],
            "deleted_at":     r["deleted_at"],
        })

    return jsonify({"success": True, "messages": out})

@app.get("/api/history")
@login_required
def api_history():
    ticket_id = request.args.get("ticket_id")
    channel_id = request.args.get("channel_id", type=int)
    if not ticket_id or channel_id is None:
        return jsonify({"success": False, "error": "ticket_id и channel_id обязательны"}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT sender, message, timestamp, message_type, attachment,
               tg_message_id, reply_to_tg_id, edited_at, deleted_at
        FROM chat_history
        WHERE ticket_id = ? AND channel_id = ?
        ORDER BY
          substr(timestamp,1,19) ASC,
          COALESCE(tg_message_id, 0) ASC,
          rowid ASC
    """, (ticket_id, channel_id))
    rows = cur.fetchall()
    conn.close()

    # строим словарь для быстрых превью исходных сообщений
    by_tg = {r["tg_message_id"]: (r["message"] or "") for r in rows if r["tg_message_id"]}

    out = []
    for r in rows:
        preview = None
        if r["reply_to_tg_id"]:
            # короткое превью исходного текста
            src = by_tg.get(r["reply_to_tg_id"])
            if src:
                preview = src[:200]  # обрежем на всякий случай
        out.append({
            "sender":         r["sender"],
            "message":        r["message"],
            "timestamp":      r["timestamp"],
            "message_type":   r["message_type"],
            "attachment":     r["attachment"],
            "tg_message_id":  r["tg_message_id"],
            "reply_to_tg_id": r["reply_to_tg_id"],
            "reply_preview":  preview,
            "edited_at":      r["edited_at"],
            "deleted_at":     r["deleted_at"],
        })

    return jsonify({"success": True, "messages": out})
    
@app.route('/reply', methods=['POST'])
@login_required
def reply():
    """
    Ответ оператора текстом.
      • лимит переоткрытий до 3 раз;
      • переоткрытие закрытого тикета;
      • уведомление клиента о переоткрытии;
      • запись в chat_history с channel_id и reply_to_tg_id;
      • возврат флага reopened для фронта.
    """
    try:
        data = request.get_json(force=True)
        user_id   = int(data['user_id'])
        ticket_id = str(data.get('ticket_id') or '')
        if not ticket_id:
            return jsonify(success=False, error="ticket_id is required"), 400

        admin    = (data.get('admin') or 'Bender').strip()
        text     = _fix_surrogates((data.get('text') or '').strip())
        reply_to = data.get('reply_to_tg_id')  # может быть None
        if not text:
            return jsonify(success=False, error='Пустой текст')

        # 0) если тикет был закрыт — переоткрываем (с лимитом: не более 3 раз)
        reopened = reopen_ticket_if_needed(ticket_id)  # True | "LIMIT_EXCEEDED" | False

        if reopened == "LIMIT_EXCEEDED":
            return jsonify(success=False, error="Лимит переоткрытий исчерпан (3 раза)."), 409

        # 1) шлём клиенту (reply_to — если был выбран пузырь)
        ok, info = send_telegram_message(
            chat_id=user_id,
            text=f"От: {admin}\n\n{text}",
            parse_mode='HTML',
            reply_to_message_id=int(reply_to) if reply_to else None,
            allow_sending_without_reply=True
        )
        if not ok:
            # аккуратно вытащим retry_after из текста ошибки, если он есть
            import re
            retry_after = None
            m = re.search(r"retry after\s+(\d+)", str(info), re.IGNORECASE)
            if m:
                retry_after = int(m.group(1))
            resp_payload = {"success": False, "error": f"Не удалось отправить: {info}"}
            if retry_after is not None:
                resp_payload["retry_after"] = retry_after
                return jsonify(resp_payload), 429  # явный сигнал фронту
            return jsonify(resp_payload), 502

        tg_msg_id = info.get('message_id') if isinstance(info, dict) else None

        # 1.1) если тикет переоткрыт — отдельное уведомление клиенту
        if reopened is True:
            send_telegram_message(
                chat_id=user_id,
                text="🔄 Ваша заявка переоткрыта. Можете продолжать диалог.",
                parse_mode='HTML'
            )

        # 2) сохраняем запись в chat_history (включая reply_to и channel_id)
        from datetime import datetime, timezone
        now_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00', 'Z')

        with get_db() as conn:
            cur = conn.cursor()
            row = cur.execute(
                "SELECT channel_id FROM tickets WHERE ticket_id = ?",
                (ticket_id,)
            ).fetchone()
            channel_id = row['channel_id'] if row else None

            cur.execute("""
                INSERT INTO chat_history (
                    user_id, ticket_id, sender, message, timestamp, message_type,
                    attachment, tg_message_id, reply_to_tg_id, channel_id
                ) VALUES (?, ?, ?, ?, ?, 'text', NULL, ?, ?, ?)
            """, (
                user_id,
                ticket_id,
                admin if admin else 'Bender',
                text,
                now_utc,
                tg_msg_id,
                reply_to,
                channel_id
            ))

            # 🟢 ОБНОВЛЕНИЕ СТАТУСА ТИКЕТА
            try:
                cur.execute("UPDATE tickets SET status = 'pending' WHERE ticket_id = ?", (ticket_id,))
                conn.commit()
                app.logger.info(f"🟢 Статус тикета {ticket_id} обновлён на 'pending'")
            except Exception as e:
                app.logger.warning(f"⚠️ Не удалось обновить статус тикета {ticket_id}: {e}")

        # 3) отдаём фронту message_id и факт переоткрытия
        return jsonify(
            success=True,
            message_type='text',
            attachment=None,
            tg_message_id=tg_msg_id,
            reopened=bool(reopened is True)
        )

    except Exception as e:
        app.logger.error(f"/reply failed: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.post('/reply_file')
@login_required
def reply_file():
    try:
        user_id   = int(request.form.get('user_id', 0))
        ticket_id = (request.form.get('ticket_id') or '').strip()
        admin     = (request.form.get('admin') or 'Bender').strip()
        reply_to  = request.form.get('reply_to_tg_id')  # может быть None

        f = request.files.get('file')
        if not user_id or not ticket_id or not f:
            return jsonify(success=False, error='bad params'), 400

        from zoneinfo import ZoneInfo
        now_utc = dt.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00', 'Z')

        # гарантируем счётчики и переоткрытие
        reopened = False
        with get_db() as conn:
            conn.row_factory = sqlite3.Row
            cols = {r["name"] for r in conn.execute("PRAGMA table_info(tickets)").fetchall()}
            if "reopen_count" not in cols:
                conn.execute("ALTER TABLE tickets ADD COLUMN reopen_count INTEGER DEFAULT 0")
            if "closed_count" not in cols:
                conn.execute("ALTER TABLE tickets ADD COLUMN closed_count INTEGER DEFAULT 0")
            if "work_time_total_sec" not in cols:
                conn.execute("ALTER TABLE tickets ADD COLUMN work_time_total_sec INTEGER DEFAULT 0")
            if "last_reopen_at" not in cols:
                conn.execute("ALTER TABLE tickets ADD COLUMN last_reopen_at TEXT")

            cur = conn.cursor()
            row = cur.execute(
                "SELECT status, COALESCE(reopen_count,0) AS rc FROM tickets WHERE ticket_id = ?",
                (ticket_id,)
            ).fetchone()
            if row and row["status"] == "resolved":
                if row["rc"] >= 3:
                    return jsonify(success=False, error="Лимит переоткрытий исчерпан (3/3)"), 409
                cur.execute("""
                    UPDATE tickets
                       SET status='pending',
                           resolved_by=NULL,
                           resolved_at=NULL,
                           reopen_count = COALESCE(reopen_count,0) + 1,
                           last_reopen_at = ?
                     WHERE ticket_id=?
                """, (now_utc, ticket_id))
                conn.commit()
                reopened = True

        if reopened:
            try:
                send_telegram_message(
                    chat_id=user_id,
                    text=f"⚠️ Обращение #{ticket_id} переоткрыто. Мы продолжим работу.",
                    parse_mode='HTML'
                )
            except Exception:
                pass

        # сохраняем файл
        os.makedirs(os.path.join(ATTACHMENTS_DIR, str(ticket_id)), exist_ok=True)
        fname = secure_filename(f.filename or 'file')
        save_path = os.path.join(ATTACHMENTS_DIR, str(ticket_id), fname)
        f.save(save_path)

        # выбираем метод Telegram по расширению
        ext = os.path.splitext(fname)[1].lower()
        method, files_key, message_type = 'sendDocument', 'document', 'document'
        if ext in {'.jpg', '.jpeg', '.png', '.webp'}:
            method, files_key, message_type = 'sendPhoto', 'photo', 'photo'
        elif ext == '.mp4':
            method, files_key, message_type = 'sendVideo', 'video', 'video'
        elif ext in {'.ogg', '.oga', '.opus'}:
            method, files_key, message_type = 'sendVoice', 'voice', 'voice'
        elif ext in {'.mp3', '.m4a', '.wav', '.aac', '.flac'}:
            method, files_key, message_type = 'sendAudio', 'audio', 'audio'
        elif ext == '.gif':
            method, files_key, message_type = 'sendAnimation', 'animation', 'animation'
        elif ext in {'.tgs', '.webm', '.webp'}:
            method, files_key, message_type = 'sendSticker', 'sticker', 'sticker'

        # шлём в Telegram
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/{method}"
        payload = {'chat_id': user_id}
        if method != 'sendSticker':
            payload['caption'] = f"От: {admin}"
            payload['parse_mode'] = 'HTML'
        if reply_to:
            payload['reply_to_message_id'] = int(reply_to)
            payload['allow_sending_without_reply'] = True

        with open(save_path, 'rb') as fh:
            resp = requests.post(url, data=payload, files={files_key: fh}, timeout=60)
        data = resp.json()
        if not data.get('ok'):
            return jsonify(success=False, error=data.get('description', 'Telegram error')), 502
        tg_msg_id = data['result']['message_id']

        # пишем в историю (с channel_id и reply_to)
        def _insert():
            with get_db() as conn:
                cur = conn.cursor()
                row = cur.execute("SELECT channel_id FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
                channel_id = row['channel_id'] if row else None
                cur.execute("""
                    INSERT INTO chat_history (
                        user_id, ticket_id, sender, message, timestamp, message_type,
                        attachment, tg_message_id, reply_to_tg_id, channel_id
                    ) VALUES (?, ?, 'support', ?, ?, ?, ?, ?, ?, ?)
                """, (user_id, ticket_id, f"От: {admin}", now_utc, message_type, save_path, tg_msg_id, reply_to, channel_id))
        exec_with_retry(_insert)

        # ✅ обновляем статус тикета (финальное подтверждение)
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE tickets SET status = 'pending' WHERE ticket_id = ?", (ticket_id,))
            conn.commit()
            app.logger.info(f"🟢 Статус тикета {ticket_id} обновлён на 'pending' (файл)")

        return jsonify(success=True, message_type=message_type, attachment=save_path,
                       tg_message_id=tg_msg_id, reopened=reopened)

    except Exception as e:
        app.logger.exception("reply_file failed")
        return jsonify(success=False, error=str(e)), 500

@app.route("/close_ticket", methods=["POST"])
@login_required
def close_ticket():
    from datetime import datetime as dt, timezone
    data = request.json
    user_id   = int(data["user_id"])
    ticket_id = data["ticket_id"]
    admin_name = (data.get("admin") or "Bender").strip()
    category   = data.get("category", "Без категории")

    try:
        # 1) перенос категории в messages
        conn = get_db()
        conn.execute("UPDATE messages SET category = ? WHERE ticket_id = ?", (category, ticket_id))

        # 2) гарантируем служебные колонки
        cols = {r["name"] for r in conn.execute("PRAGMA table_info(tickets)").fetchall()}
        if "reopen_count" not in cols:
            conn.execute("ALTER TABLE tickets ADD COLUMN reopen_count INTEGER DEFAULT 0")
        if "closed_count" not in cols:
            conn.execute("ALTER TABLE tickets ADD COLUMN closed_count INTEGER DEFAULT 0")
        if "work_time_total_sec" not in cols:
            conn.execute("ALTER TABLE tickets ADD COLUMN work_time_total_sec INTEGER DEFAULT 0")
        if "last_reopen_at" not in cols:
            conn.execute("ALTER TABLE tickets ADD COLUMN last_reopen_at TEXT")

        # 3) добиваем суммарное время: если есть last_reopen_at — прибавим дельту
        row = conn.execute("SELECT work_time_total_sec, last_reopen_at FROM tickets WHERE ticket_id=?",
                           (ticket_id,)).fetchone()
        total_sec = int((row["work_time_total_sec"] or 0) if row else 0)
        if row and row["last_reopen_at"]:
            try:
                dt_last = dt.fromisoformat(row["last_reopen_at"])
            except Exception:
                dt_last = None
            if dt_last:
                total_sec += int((dt.now() - dt_last).total_seconds())

        # 4) помечаем закрытым, увеличиваем closed_count, обнуляем last_reopen_at, фиксируем total_sec
        now_iso = dt.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        conn.execute("""
            UPDATE tickets
               SET status='resolved',
                   resolved_at=?,
                   resolved_by=?,
                   closed_count = COALESCE(closed_count,0) + 1,
                   work_time_total_sec = ?,
                   last_reopen_at = NULL
             WHERE ticket_id = ?
        """, (now_iso, admin_name, total_sec, ticket_id))
        conn.commit()
        conn.close()

                # 5) уведомления клиенту: закрытие + просьба оценить (пер-ботово из настроек)
        try:
            # определим channel_id тикета
            with get_db() as conn2:
                row_ch = conn2.execute("SELECT channel_id FROM tickets WHERE ticket_id=?", (ticket_id,)).fetchone()
                channel_id = row_ch["channel_id"] if row_ch else None

            close_msg = f"Ваше обращение #{ticket_id} закрыто. Для запуска нового диалога нажмите /start"
            send_telegram_message(chat_id=user_id, text=close_msg, parse_mode='HTML')

            # 5.1. ставим «ожидание оценки» на 24 часа
            try:
                with get_db() as conn2:
                    expires = (datetime.datetime.now() + datetime.timedelta(hours=24)).isoformat()
                    conn2.execute("""
                        INSERT OR REPLACE INTO pending_feedback_requests
                            (user_id, channel_id, ticket_id, source, created_at, expires_at)
                        VALUES (?,?,?,?,?,?)
                    """, (user_id, channel_id, ticket_id, 'operator_close', dt.now().isoformat(), expires))
            except Exception as e:
                print(f"pending_feedback_requests insert error: {e}")

            def send_sorry():
                with sqlite3.connect(os.path.join(BASE_DIR, "tickets.db")) as conn2:
                    row2 = conn2.execute("SELECT 1 FROM feedbacks WHERE user_id = ? LIMIT 1", (user_id,)).fetchone()
                if not row2:
                    send_telegram_message(chat_id=user_id, text="Очень жаль, что не получили вашей оценки.", parse_mode='HTML')

            from threading import Timer
            Timer(900, send_sorry).start()

        except Exception as e:
            logging.error(f"❌ Ошибка при отправке уведомлений: {e}")

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
        
if __name__ == "__main__":
    app.run(port=5000, debug=True)
